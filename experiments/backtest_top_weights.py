"""
experiments/backtest_top_weights.py

Детальне порівняння топових наборів ваг з калібрування.
Повний грід ODDS 1.65-3.0 × GAP 80/90/110 для кожного набору ваг.
Зберігає Excel з кольоровим форматуванням для зручного аналізу.

Запуск: python -m experiments.backtest_top_weights
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DATABASE_URL', 'postgresql://capper:capper@localhost:5432/capper')

import pandas as pd
from loguru import logger

from model.train import load_data_from_db
from model.features.builder import build_dataset
from model.weighted_score import _get_factors, AWAY_WEIGHTS, HOME_WEIGHTS

SPLIT_1    = pd.Timestamp("2025-01-01")
SPLIT_2    = pd.Timestamp("2025-07-01")
INITIAL_BR = 1000.0
FRAC       = 0.25
CAP        = 0.04
ODDS_MAX   = 3.0

TOP5 = {"Premier League", "La Liga", "Bundesliga", "Serie A", "Ligue 1"}

# ---------------------------------------------------------------------------
# Набори ваг для порівняння
# ---------------------------------------------------------------------------
# Змінюємо тільки 4 фактори, решта — з AWAY_WEIGHTS
WEIGHT_SETS = [
    {
        "label": "CURRENT (baseline)",
        "big_injury_adv_away":  10,
        "away_possession_edge": 10,
        "away_xg_regression":    8,
        "home_elo_weak":         15,
    },
    {
        "label": "A: inj↓7 (best 2.10+)",
        "big_injury_adv_away":   7,
        "away_possession_edge": 10,
        "away_xg_regression":    8,
        "home_elo_weak":         15,
    },
    {
        "label": "B: inj↓5, poss=0 (best 1.90)",
        "big_injury_adv_away":   5,
        "away_possession_edge":  0,
        "away_xg_regression":    8,
        "home_elo_weak":         15,
    },
    {
        "label": "C: inj↓5, poss↓5",
        "big_injury_adv_away":   5,
        "away_possession_edge":  5,
        "away_xg_regression":    8,
        "home_elo_weak":         15,
    },
    {
        "label": "D: structural pure (inj=3,poss=0,xgr=0)",
        "big_injury_adv_away":   3,
        "away_possession_edge":  0,
        "away_xg_regression":    0,
        "home_elo_weak":         15,
    },
    {
        "label": "E: inj=0,poss=3 (best 2.20)",
        "big_injury_adv_away":   0,
        "away_possession_edge":  3,
        "away_xg_regression":    8,
        "home_elo_weak":         10,
    },
    {
        "label": "F: balanced (inj=5,poss=7,xgr=5)",
        "big_injury_adv_away":   5,
        "away_possession_edge":  7,
        "away_xg_regression":    5,
        "home_elo_weak":         12,
    },
]

ODDS_MIN_VALS = [1.65, 1.75, 1.90, 2.00, 2.10, 2.20, 2.40, 2.60]
GAP_MIN_VALS  = [80, 90, 110]


def build_away_weights(overrides: dict) -> dict:
    w = dict(AWAY_WEIGHTS)
    w.update({k: v for k, v in overrides.items() if k != "label"})
    return w


def simulate(bets: list) -> dict:
    if not bets:
        return {"n": 0, "wr": 0.0, "flat_roi": 0.0,
                "kelly_roi": 0.0, "bankroll": INITIAL_BR, "max_dd": 0.0}
    bankroll = INITIAL_BR
    peak     = INITIAL_BR
    max_dd   = 0.0
    staked   = 0.0
    flat_pnl = 0.0
    wins     = 0
    for b in bets:
        p, odd = b["p_elo"], b["odds"]
        denom  = odd - 1
        kelly  = max(0.0, (p * denom - (1 - p)) / denom) * FRAC if denom > 0 else 0
        stake  = min(bankroll * kelly, bankroll * CAP)
        if stake > 0 and bankroll > 0:
            if b["won"]:
                bankroll += stake * denom; wins += 1
            else:
                bankroll -= stake
            staked += stake
            peak    = max(peak, bankroll)
            max_dd  = max(max_dd, (peak - bankroll) / peak * 100)
        flat_pnl += denom if b["won"] else -1.0
    n = len(bets)
    return {
        "n": n, "wr": wins / n,
        "flat_roi":  flat_pnl / n * 100,
        "kelly_roi": (bankroll - INITIAL_BR) / staked * 100 if staked > 0 else 0.0,
        "bankroll":  round(bankroll, 2),
        "max_dd":    round(max_dd, 1),
    }


def run():
    logger.info("Loading data...")
    matches_df, stats_df, odds_df, teams, injuries_df = load_data_from_db()
    logger.info("Building dataset...")
    df = build_dataset(matches_df, stats_df, odds_df, teams, injuries_df)
    df["date"] = pd.to_datetime(df["date"])
    logger.info(f"Dataset: {len(df)} rows")

    from db.session import engine
    from sqlalchemy import text
    with engine.connect() as conn:
        rows = conn.execute(text(
            "SELECT m.id, l.name FROM matches m JOIN leagues l ON l.id = m.league_id"
        )).fetchall()
    league_map = {r[0]: r[1] for r in rows}

    # Передобчислюємо активні фактори один раз
    logger.info("Pre-computing factors...")
    base = []
    for _, row in df.iterrows():
        r = row.to_dict()
        mid = int(r.get("match_id", 0))
        if league_map.get(mid, "") not in TOP5:
            continue
        a_odds = r.get("away_odds")
        if not a_odds or a_odds > ODDS_MAX:
            continue
        factors_h = {n: HOME_WEIGHTS.get(n, 1) for n, active in _get_factors(r, "home") if active}
        factors_a_raw = {n: v for n, v in {n: AWAY_WEIGHTS.get(n, 1)
                         for n, active in _get_factors(r, "away") if active}.items()}
        base.append({
            "date":       row["date"],
            "odds":       round(a_odds, 2),
            "p_elo":      1.0 - float(r.get("elo_home_win_prob", 0.5)),
            "won":        r.get("target") == "away",
            "active_h":   factors_h,
            "active_a":   factors_a_raw,
        })
    logger.info(f"Base records: {len(base)}")

    # ---------------------------------------------------------------------------
    # Основний цикл: по кожному набору ваг
    # ---------------------------------------------------------------------------
    all_results = []

    for ws in WEIGHT_SETS:
        label    = ws["label"]
        away_w   = build_away_weights(ws)

        # Обчислюємо WS і формуємо пул
        pool = []
        for rec in base:
            ws_h = sum(HOME_WEIGHTS.get(n, v) for n, v in rec["active_h"].items())
            ws_a = sum(away_w.get(n, v) for n, v in rec["active_a"].items())
            if ws_a <= ws_h:
                continue
            pool.append({
                "date":  rec["date"],
                "odds":  rec["odds"],
                "gap":   ws_a - ws_h,
                "p_elo": rec["p_elo"],
                "won":   rec["won"],
            })

        pool.sort(key=lambda x: x["date"])
        pool_df = pd.DataFrame(pool) if pool else pd.DataFrame(
            columns=["date", "odds", "gap", "p_elo", "won"])

        f1_df = pool_df[(pool_df["date"] >= SPLIT_1) & (pool_df["date"] < SPLIT_2)] \
                if len(pool_df) else pool_df
        f2_df = pool_df[pool_df["date"] >= SPLIT_2] if len(pool_df) else pool_df

        for odds_min in ODDS_MIN_VALS:
            for gap_min in GAP_MIN_VALS:
                def filt(d, g=gap_min, o=odds_min):
                    if d.empty: return []
                    return d[(d["gap"] >= g) & (d["odds"] >= o)].to_dict("records")

                sa = simulate(filt(pool_df))
                sf1 = simulate(filt(f1_df))
                sf2 = simulate(filt(f2_df))

                if sa["n"] < 15:
                    continue

                avg = (sf1["flat_roi"] + sf2["flat_roi"]) / 2

                all_results.append({
                    "label":       label,
                    "odds_min":    odds_min,
                    "gap_min":     gap_min,
                    "N_all":       sa["n"],
                    "N_F1":        sf1["n"],
                    "N_F2":        sf2["n"],
                    "WR_all_%":    round(sa["wr"] * 100, 1),
                    "F1_ROI_%":    round(sf1["flat_roi"], 1),
                    "F2_ROI_%":    round(sf2["flat_roi"], 1),
                    "KellyF1_%":   round(sf1["kelly_roi"], 1),
                    "KellyF2_%":   round(sf2["kelly_roi"], 1),
                    "MaxDD_F2_%":  sf2["max_dd"],
                    "BK_all_$":    sa["bankroll"],
                    "Avg_ROI_%":   round(avg, 1),
                })

        logger.info(f"Done: {label}")

    res_df = pd.DataFrame(all_results)

    # ---------------------------------------------------------------------------
    # Консольний вивід: по кожному набору ваг — найкращі конфіги
    # ---------------------------------------------------------------------------
    for ws in WEIGHT_SETS:
        label = ws["label"]
        sub   = res_df[res_df["label"] == label].sort_values("Avg_ROI_%", ascending=False)
        good  = sub[(sub["N_F1"] >= 15) & (sub["N_F2"] >= 15) &
                    (sub["F1_ROI_%"] > 0) & (sub["F2_ROI_%"] > 0)]

        print(f"\n{'='*100}")
        print(f"  {label}")
        if ws != WEIGHT_SETS[0]:
            changed = {k: f"{WEIGHT_SETS[0][k]}→{ws[k]}" for k in
                       ["big_injury_adv_away","away_possession_edge","away_xg_regression","home_elo_weak"]
                       if ws[k] != WEIGHT_SETS[0][k]}
            print(f"  Зміни: {changed}")
        print(f"  {'odds':>5} {'gap':>4} | {'N_all':>6} {'N_F1':>5} {'N_F2':>5} | "
              f"{'F1':>8} {'F2':>8} {'Avg':>8} | {'KellyF2':>8} {'MaxDD':>6} {'BK_all':>8}")
        print(f"  {'-'*95}")
        for _, r in good.head(10).iterrows():
            print(f"  {r['odds_min']:>5.2f} {int(r['gap_min']):>4} | "
                  f"{int(r['N_all']):>6} {int(r['N_F1']):>5} {int(r['N_F2']):>5} | "
                  f"{r['F1_ROI_%']:>+8.1f}% {r['F2_ROI_%']:>+8.1f}% {r['Avg_ROI_%']:>+8.1f}% | "
                  f"{r['KellyF2_%']:>+8.1f}% {r['MaxDD_F2_%']:>5.1f}% ${r['BK_all_$']:>7,.0f}")

    # ---------------------------------------------------------------------------
    # Зведена таблиця: для топ-конфігів порівняти всі набори ваг
    # ---------------------------------------------------------------------------
    KEY_CONFIGS = [(2.20, 90), (2.10, 90), (2.10, 110), (2.00, 110), (1.90, 110), (1.90, 90)]
    print(f"\n\n{'='*100}")
    print(f"  ЗВЕДЕННЯ: F2 ROI по всіх наборах ваг для ключових конфігів")
    header = f"  {'Config':<14} |"
    for ws in WEIGHT_SETS:
        header += f" {ws['label'][:12]:>13}"
    print(header)
    print(f"  {'-'*95}")
    for omin, gmin in KEY_CONFIGS:
        row_str = f"  {omin:.2f}/GAP{gmin:<4} |"
        for ws in WEIGHT_SETS:
            sub = res_df[(res_df["label"] == ws["label"]) &
                         (res_df["odds_min"] == omin) &
                         (res_df["gap_min"] == gmin)]
            if sub.empty:
                row_str += f"        {'N/A':>5}"
            else:
                f2 = sub.iloc[0]["F2_ROI_%"]
                n  = sub.iloc[0]["N_F2"]
                row_str += f" {f2:>+7.1f}%(n{n:>2})"
        print(row_str)

    # ---------------------------------------------------------------------------
    # Збереження Excel
    # ---------------------------------------------------------------------------
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # Один аркуш на кожен набір ваг
        for wi, ws in enumerate(WEIGHT_SETS):
            label = ws["label"]
            sub   = res_df[res_df["label"] == label].sort_values("Avg_ROI_%", ascending=False)

            sheet = wb.active if wi == 0 else wb.create_sheet()
            sheet.title = label[:28].replace(":", "-").replace("/", "-")

            hfill = PatternFill("solid", fgColor="1F3864")
            hfont = Font(bold=True, color="FFFFFF", size=9)
            cols  = [c for c in sub.columns if c != "label"]

            for ci, h in enumerate(cols, 1):
                c = sheet.cell(row=1, column=ci, value=h)
                c.fill = hfill; c.font = hfont
                c.alignment = Alignment(horizontal="center")

            for ri, (_, row_data) in enumerate(sub[cols].iterrows(), 2):
                f2 = row_data["F2_ROI_%"]
                if f2 >= 15:
                    rf = PatternFill("solid", fgColor="C6EFCE"); rfo = Font(color="006100", size=9)
                elif f2 >= 5:
                    rf = PatternFill("solid", fgColor="EBFCD8"); rfo = Font(color="2D6A04", size=9)
                elif f2 >= 0:
                    rf = PatternFill("solid", fgColor="FFFFE0"); rfo = Font(color="7D6608", size=9)
                elif f2 >= -10:
                    rf = PatternFill("solid", fgColor="FFE5CC"); rfo = Font(color="8B3A00", size=9)
                else:
                    rf = PatternFill("solid", fgColor="FFC7CE"); rfo = Font(color="9C0006", size=9)
                for ci, val in enumerate(row_data, 1):
                    c = sheet.cell(row=ri, column=ci, value=val)
                    c.fill = rf; c.font = rfo
                    c.alignment = Alignment(horizontal="center")

            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
            for ci, h in enumerate(cols, 1):
                sheet.column_dimensions[get_column_letter(ci)].width = max(len(h)+2, 9)

        xlsx_path = "experiments/results/top_weights.xlsx"
        wb.save(xlsx_path)
        logger.info(f"Excel saved: {xlsx_path}")
    except ImportError:
        logger.warning("openpyxl not installed")

    csv_path = "experiments/results/top_weights.csv"
    res_df.to_csv(csv_path, index=False)
    logger.info(f"CSV saved: {csv_path}")


if __name__ == "__main__":
    run()
