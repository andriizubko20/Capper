"""
experiments/backtest_weight_calibration.py

Калібрування ваг підозрілих факторів × повний грід odds/gap.
Перебирає комбінації ваг для 4 факторів і тестує на ODDS 1.65-3.0.

Запуск: python -m experiments.backtest_weight_calibration
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DATABASE_URL', 'postgresql://capper:capper@localhost:5432/capper')

import pandas as pd
import itertools
from loguru import logger

from model.train import load_data_from_db
from model.features.builder import build_dataset
from model.weighted_score import _get_factors, AWAY_WEIGHTS, HOME_WEIGHTS

SPLIT_1    = pd.Timestamp("2025-01-01")
SPLIT_2    = pd.Timestamp("2025-07-01")
INITIAL_BR = 1000.0
FRAC       = 0.25
CAP        = 0.04
ODDS_MAX   = 3.0

TOP5 = {"Premier League", "La Liga", "Bundesliga", "Serie A", "Ligue 1"}

# ---------------------------------------------------------------------------
# Фактори для калібрування і їхні тест-значення
# ---------------------------------------------------------------------------
CALIB = {
    "big_injury_adv_away":  [0, 3, 5, 7, 10],   # поточне: 10
    "away_possession_edge": [0, 3, 5, 7, 10],   # поточне: 10
    "away_xg_regression":   [0, 3, 5, 8],       # поточне: 8
    "home_elo_weak":        [8, 10, 12, 15],     # поточне: 15
}

# ---------------------------------------------------------------------------
# Повний конфіг-грід: odds_min × gap_min
# ---------------------------------------------------------------------------
ODDS_MIN_VALS = [1.65, 1.75, 1.90, 2.00, 2.10, 2.20, 2.40, 2.60]
GAP_MIN_VALS  = [80, 90, 110]


def simulate(bets: list) -> dict:
    if not bets:
        return {"n": 0, "wr": 0.0, "flat_roi": 0.0,
                "kelly_roi": 0.0, "bankroll": INITIAL_BR, "max_dd": 0.0}
    bankroll = INITIAL_BR
    peak     = INITIAL_BR
    max_dd   = 0.0
    staked   = 0.0
    flat_pnl = 0.0
    wins     = 0
    for b in bets:   # already sorted by date
        p, odd = b["p_elo"], b["odds"]
        denom  = odd - 1
        kelly  = max(0.0, (p * denom - (1 - p)) / denom) * FRAC if denom > 0 else 0
        stake  = min(bankroll * kelly, bankroll * CAP)
        if stake > 0 and bankroll > 0:
            if b["won"]:
                bankroll += stake * denom; wins += 1
            else:
                bankroll -= stake
            staked += stake
            peak    = max(peak, bankroll)
            max_dd  = max(max_dd, (peak - bankroll) / peak * 100)
        flat_pnl += denom if b["won"] else -1.0
    n = len(bets)
    return {
        "n": n, "wr": wins / n,
        "flat_roi": flat_pnl / n * 100,
        "kelly_roi": (bankroll - INITIAL_BR) / staked * 100 if staked > 0 else 0.0,
        "bankroll": round(bankroll, 2),
        "max_dd": round(max_dd, 1),
    }


def run():
    logger.info("Loading data...")
    matches_df, stats_df, odds_df, teams, injuries_df = load_data_from_db()
    logger.info("Building dataset...")
    df = build_dataset(matches_df, stats_df, odds_df, teams, injuries_df)
    df["date"] = pd.to_datetime(df["date"])
    logger.info(f"Dataset: {len(df)} rows")

    from db.session import engine
    from sqlalchemy import text
    with engine.connect() as conn:
        rows = conn.execute(text(
            "SELECT m.id, l.name FROM matches m JOIN leagues l ON l.id = m.league_id"
        )).fetchall()
    league_map = {r[0]: r[1] for r in rows}

    # -----------------------------------------------------------------------
    # Крок 1: Передобчислюємо активні фактори для кожного рядка (ОДИН РАЗ)
    # -----------------------------------------------------------------------
    logger.info("Pre-computing factor activity for all records...")
    base_records = []
    for _, row in df.iterrows():
        r = row.to_dict()
        mid    = int(r.get("match_id", 0))
        league = league_map.get(mid, "")
        if league not in TOP5:
            continue
        a_odds = r.get("away_odds")
        if not a_odds or a_odds > ODDS_MAX:
            continue

        factors_h = _get_factors(r, "home")
        factors_a = _get_factors(r, "away")

        # Зберігаємо як dict {name: weight_contributes} лише активні
        active_h = {n: HOME_WEIGHTS.get(n, 1) for n, active in factors_h if active}
        active_a = {n: AWAY_WEIGHTS.get(n, 1) for n, active in factors_a if active}

        base_records.append({
            "date":     row["date"],
            "odds":     round(a_odds, 2),
            "p_elo":    1.0 - float(r.get("elo_home_win_prob", 0.5)),
            "won":      r.get("target") == "away",
            "active_h": active_h,
            "active_a": active_a,
        })

    logger.info(f"Base records (Top-5, odds≤{ODDS_MAX}): {len(base_records)}")

    # -----------------------------------------------------------------------
    # Крок 2: Грід ваг × грід конфігів
    # -----------------------------------------------------------------------
    calib_keys   = list(CALIB.keys())
    calib_values = [CALIB[k] for k in calib_keys]
    weight_combos = list(itertools.product(*calib_values))

    total = len(weight_combos) * len(ODDS_MIN_VALS) * len(GAP_MIN_VALS)
    logger.info(f"Weight combos: {len(weight_combos)} | Configs: {len(ODDS_MIN_VALS)*len(GAP_MIN_VALS)} | Total: {total}")

    results = []

    for wi, combo in enumerate(weight_combos):
        # Модифіковані ваги для AWAY (HOME лишаємо без змін)
        w_override = dict(zip(calib_keys, combo))
        away_w = dict(AWAY_WEIGHTS)
        away_w.update(w_override)

        # Обчислюємо WS для кожного запису з новими вагами
        pool = []
        for rec in base_records:
            ws_h = sum(HOME_WEIGHTS.get(n, v) for n, v in rec["active_h"].items())
            ws_a = sum(away_w.get(n, v) for n, v in rec["active_a"].items())
            if ws_a <= ws_h:
                continue
            pool.append({
                "date":   rec["date"],
                "odds":   rec["odds"],
                "gap":    ws_a - ws_h,
                "p_elo":  rec["p_elo"],
                "won":    rec["won"],
            })

        pool.sort(key=lambda x: x["date"])
        pool_df = pd.DataFrame(pool) if pool else pd.DataFrame(
            columns=["date", "odds", "gap", "p_elo", "won"])

        f1_pool = pool_df[(pool_df["date"] >= SPLIT_1) & (pool_df["date"] < SPLIT_2)] if len(pool_df) else pool_df
        f2_pool = pool_df[pool_df["date"] >= SPLIT_2] if len(pool_df) else pool_df

        for odds_min in ODDS_MIN_VALS:
            for gap_min in GAP_MIN_VALS:
                def filt(d):
                    if d.empty:
                        return []
                    return d[(d["gap"] >= gap_min) & (d["odds"] >= odds_min)].to_dict("records")

                bets_all = filt(pool_df)
                bets_f1  = filt(f1_pool)
                bets_f2  = filt(f2_pool)

                s_all = simulate(bets_all)
                s_f1  = simulate(bets_f1)
                s_f2  = simulate(bets_f2)

                if s_all["n"] < 20:
                    continue

                avg = (s_f1["flat_roi"] + s_f2["flat_roi"]) / 2

                results.append({
                    # ваги
                    "w_inj":  w_override["big_injury_adv_away"],
                    "w_poss": w_override["away_possession_edge"],
                    "w_xgr":  w_override["away_xg_regression"],
                    "w_elo_weak": w_override["home_elo_weak"],
                    # конфіг
                    "odds_min": odds_min,
                    "gap_min":  gap_min,
                    # результати
                    "N_all":  s_all["n"],
                    "N_F1":   s_f1["n"],
                    "N_F2":   s_f2["n"],
                    "WR_all_%":    round(s_all["wr"] * 100, 1),
                    "F1_ROI_%":    round(s_f1["flat_roi"], 1),
                    "F2_ROI_%":    round(s_f2["flat_roi"], 1),
                    "KellyF2_%":   round(s_f2["kelly_roi"], 1),
                    "MaxDD_F2_%":  s_f2["max_dd"],
                    "Avg_ROI_%":   round(avg, 1),
                })

        if (wi + 1) % 100 == 0:
            logger.info(f"  {wi+1}/{len(weight_combos)} weight combos done...")

    logger.info(f"Total results: {len(results)}")
    res_df = pd.DataFrame(results).sort_values("Avg_ROI_%", ascending=False)

    # -----------------------------------------------------------------------
    # Збереження
    # -----------------------------------------------------------------------
    csv_path = "experiments/results/weight_calibration.csv"
    res_df.to_csv(csv_path, index=False)
    logger.info(f"CSV saved: {csv_path}")

    # Excel з кольорами
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "WeightCalib"

        hfill = PatternFill("solid", fgColor="1F3864")
        hfont = Font(bold=True, color="FFFFFF", size=9)
        headers = list(res_df.columns)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = hfill; c.font = hfont
            c.alignment = Alignment(horizontal="center")

        for ri, (_, row_data) in enumerate(res_df.iterrows(), 2):
            f2 = row_data["F2_ROI_%"]
            if f2 >= 15:
                rf = PatternFill("solid", fgColor="C6EFCE"); rfo = Font(color="006100", size=9)
            elif f2 >= 5:
                rf = PatternFill("solid", fgColor="EBFCD8"); rfo = Font(color="2D6A04", size=9)
            elif f2 >= 0:
                rf = PatternFill("solid", fgColor="FFFFE0"); rfo = Font(color="7D6608", size=9)
            elif f2 >= -10:
                rf = PatternFill("solid", fgColor="FFE5CC"); rfo = Font(color="8B3A00", size=9)
            else:
                rf = PatternFill("solid", fgColor="FFC7CE"); rfo = Font(color="9C0006", size=9)

            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = rf; c.font = rfo
                c.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        for ci, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(ci)].width = max(len(h) + 2, 8)

        xlsx_path = "experiments/results/weight_calibration.xlsx"
        wb.save(xlsx_path)
        logger.info(f"Excel saved: {xlsx_path}")
    except ImportError:
        logger.warning("openpyxl not installed")

    # -----------------------------------------------------------------------
    # Топ результати — по групах
    # -----------------------------------------------------------------------
    good = res_df[
        (res_df["N_F1"] >= 20) & (res_df["N_F2"] >= 20) &
        (res_df["F1_ROI_%"] > 0) & (res_df["F2_ROI_%"] > 0)
    ]

    print(f"\n{'='*110}")
    print(f"  TOP-20 (обидва фолди > 0%, N>=20 кожен)")
    print(f"  {'inj':>4} {'poss':>5} {'xgr':>5} {'elo_w':>6} | "
          f"{'odds':>5} {'gap':>4} | {'N_all':>6} {'N_F2':>5} | "
          f"{'F1':>7} {'F2':>7} {'Avg':>7} | {'KellyF2':>8} {'MaxDD':>6}")
    print(f"  {'-'*105}")
    seen = set()
    count = 0
    for _, r in good.iterrows():
        key = (r["w_inj"], r["w_poss"], r["w_xgr"], r["w_elo_weak"])
        if key in seen:
            continue
        seen.add(key)
        print(f"  {int(r['w_inj']):>4} {int(r['w_poss']):>5} {int(r['w_xgr']):>5} {int(r['w_elo_weak']):>6} | "
              f"{r['odds_min']:>5.2f} {int(r['gap_min']):>4} | "
              f"{int(r['N_all']):>6} {int(r['N_F2']):>5} | "
              f"{r['F1_ROI_%']:>+7.1f}% {r['F2_ROI_%']:>+7.1f}% {r['Avg_ROI_%']:>+7.1f}% | "
              f"{r['KellyF2_%']:>+8.1f}% {r['MaxDD_F2_%']:>5.1f}%")
        count += 1
        if count >= 20:
            break

    # Найкращі ваги окремо по кожному odds_min
    print(f"\n{'='*85}")
    print(f"  BEST WEIGHTS per odds_min (F2>0, N_F2>=20)")
    print(f"  {'odds':>5} | {'inj':>4} {'poss':>5} {'xgr':>5} {'elo_w':>6} | "
          f"{'gap':>4} {'N_F2':>5} {'F2':>8} {'Avg':>8}")
    print(f"  {'-'*75}")
    for omin in ODDS_MIN_VALS:
        sub = good[good["odds_min"] == omin]
        if sub.empty:
            continue
        best = sub.iloc[0]
        print(f"  {omin:>5.2f} | "
              f"{int(best['w_inj']):>4} {int(best['w_poss']):>5} {int(best['w_xgr']):>5} {int(best['w_elo_weak']):>6} | "
              f"{int(best['gap_min']):>4} {int(best['N_F2']):>5} "
              f"{best['F2_ROI_%']:>+8.1f}% {best['Avg_ROI_%']:>+8.1f}%")


if __name__ == "__main__":
    run()
