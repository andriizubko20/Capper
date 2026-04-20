"""
experiments/backtest_full_grid.py

Повний грід: ODDS_MIN × GAP × EV
Зберігає Excel з кольоровим форматуванням.

Запуск: python -m experiments.backtest_full_grid
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DATABASE_URL', 'postgresql://capper:capper@localhost:5432/capper')

import pandas as pd
from loguru import logger

from model.train import load_data_from_db
from model.features.builder import build_dataset
from model.weighted_score import _get_factors, AWAY_WEIGHTS, HOME_WEIGHTS

SPLIT_1    = pd.Timestamp("2025-01-01")
SPLIT_2    = pd.Timestamp("2025-07-01")
INITIAL_BR = 1000.0
FRAC       = 0.25
CAP        = 0.04
ODDS_MAX   = 3.0

TOP5 = {"Premier League", "La Liga", "Bundesliga", "Serie A", "Ligue 1"}

ODDS_MIN_VALS = [1.65, 1.75, 1.90, 2.00, 2.10, 2.20, 2.40, 2.60]
GAP_VALS      = [80, 90, 100, 110, 120]
EV_MIN_VALS   = [0.03, 0.05, 0.07, 0.10, 0.12, 0.15, 0.18, 0.20, 0.25]


def compute_ws(features: dict, outcome: str) -> float:
    weights = HOME_WEIGHTS if outcome == "home" else AWAY_WEIGHTS
    return sum(weights.get(n, 1) for n, active in _get_factors(features, outcome) if active)


def simulate(bets: list) -> dict:
    if not bets:
        return {"n": 0, "wins": 0, "wr": 0.0, "flat_roi": 0.0,
                "kelly_roi": 0.0, "bankroll": INITIAL_BR, "max_dd": 0.0}
    bankroll = INITIAL_BR
    peak = INITIAL_BR
    max_dd = 0.0
    staked = 0.0
    flat_pnl = 0.0
    wins = 0
    for b in bets:
        p, odd = b["p_elo"], b["odds"]
        q = 1 - p
        denom = odd - 1
        kelly = max(0.0, (p * denom - q) / denom) * FRAC if denom > 0 else 0
        stake = min(bankroll * kelly, bankroll * CAP)
        if stake > 0 and bankroll > 0:
            if b["won"]:
                bankroll += stake * (odd - 1); wins += 1
            else:
                bankroll -= stake
            staked += stake
            peak = max(peak, bankroll)
            max_dd = max(max_dd, (peak - bankroll) / peak * 100)
        flat_pnl += (odd - 1) if b["won"] else -1.0
    n = len(bets)
    return {
        "n": n, "wins": wins, "wr": wins / n if n else 0,
        "flat_roi": flat_pnl / n * 100 if n else 0,
        "kelly_roi": (bankroll - INITIAL_BR) / staked * 100 if staked > 0 else 0.0,
        "bankroll": round(bankroll, 2),
        "max_dd": round(max_dd, 1),
    }


def run():
    logger.info("Loading data...")
    matches_df, stats_df, odds_df, teams, injuries_df = load_data_from_db()
    logger.info("Building dataset...")
    df = build_dataset(matches_df, stats_df, odds_df, teams, injuries_df)
    df["date"] = pd.to_datetime(df["date"])
    logger.info(f"Dataset: {len(df)} rows")

    from db.session import engine
    from sqlalchemy import text
    with engine.connect() as conn:
        rows = conn.execute(text("SELECT m.id, l.name FROM matches m JOIN leagues l ON l.id = m.league_id")).fetchall()
    league_map = {r[0]: r[1] for r in rows}

    logger.info("Pre-computing WS scores...")
    records = []
    for _, row in df.iterrows():
        r = row.to_dict()
        ws_h = compute_ws(r, "home")
        ws_a = compute_ws(r, "away")
        if ws_a <= ws_h:
            continue

        gap = ws_a - ws_h
        a_odds = r.get("away_odds")
        if not a_odds or a_odds > ODDS_MAX:
            continue

        mid = int(r.get("match_id", 0))
        league = league_map.get(mid, "")
        if league not in TOP5:
            continue

        p_elo = 1.0 - float(r.get("elo_home_win_prob", 0.5))
        ev = round(p_elo * a_odds - 1.0, 4)

        records.append({
            "date":   row["date"],
            "odds":   round(a_odds, 2),
            "ws_gap": gap,
            "p_elo":  p_elo,
            "ev":     ev,
            "won":    r.get("target") == "away",
        })

    all_df = pd.DataFrame(records).sort_values("date")
    f1_df  = all_df[(all_df["date"] >= SPLIT_1) & (all_df["date"] < SPLIT_2)]
    f2_df  = all_df[all_df["date"] >= SPLIT_2]
    logger.info(f"Pool: {len(all_df)} records")

    # --- GRID ---
    results = []
    total = len(ODDS_MIN_VALS) * len(GAP_VALS) * len(EV_MIN_VALS)
    logger.info(f"Running {total} combinations...")

    for odds_min in ODDS_MIN_VALS:
        for gap_min in GAP_VALS:
            for ev_min in EV_MIN_VALS:

                def filt(d):
                    return d[
                        (d["ws_gap"] >= gap_min) &
                        (d["odds"]   >= odds_min) &
                        (d["ev"]     >= ev_min)
                    ].to_dict("records")

                bets_all = filt(all_df)
                bets_f1  = filt(f1_df)
                bets_f2  = filt(f2_df)

                s_all = simulate(bets_all)
                s_f1  = simulate(bets_f1)
                s_f2  = simulate(bets_f2)

                avg_roi = (s_f1["flat_roi"] + s_f2["flat_roi"]) / 2

                results.append({
                    "ODDS_MIN":   odds_min,
                    "GAP_MIN":    gap_min,
                    "EV_MIN_%":   int(ev_min * 100),
                    # All data
                    "N_all":      s_all["n"],
                    "WR_all_%":   round(s_all["wr"] * 100, 1),
                    "FlatROI_all_%": round(s_all["flat_roi"], 1),
                    "BK_all_$":   s_all["bankroll"],
                    # F1
                    "N_F1":       s_f1["n"],
                    "WR_F1_%":    round(s_f1["wr"] * 100, 1),
                    "FlatROI_F1_%": round(s_f1["flat_roi"], 1),
                    "KellyROI_F1_%": round(s_f1["kelly_roi"], 1),
                    # F2
                    "N_F2":       s_f2["n"],
                    "WR_F2_%":    round(s_f2["wr"] * 100, 1),
                    "FlatROI_F2_%": round(s_f2["flat_roi"], 1),
                    "KellyROI_F2_%": round(s_f2["kelly_roi"], 1),
                    "MaxDD_F2_%": s_f2["max_dd"],
                    # Summary
                    "Avg_F1_F2_%": round(avg_roi, 1),
                })

    out = pd.DataFrame(results).sort_values("Avg_F1_F2_%", ascending=False)

    # --- Save CSV ---
    csv_path = "experiments/results/full_grid.csv"
    out.to_csv(csv_path, index=False)
    logger.info(f"CSV: {csv_path}")

    # --- Save Excel with color ---
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.formatting.rule import ColorScaleRule

        wb = Workbook()
        ws = wb.active
        ws.title = "Grid"

        header_fill = PatternFill("solid", fgColor="1F3864")
        header_font = Font(bold=True, color="FFFFFF", size=10)

        headers = list(out.columns)
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Color scale fill ranges
        GREEN_DARK  = "006100"
        GREEN_MED   = "C6EFCE"
        YELLOW      = "FFEB9C"
        RED_MED     = "FFC7CE"
        RED_DARK    = "9C0006"

        for ri, (_, row_data) in enumerate(out.iterrows(), 2):
            f2_roi = row_data["FlatROI_F2_%"]
            avg    = row_data["Avg_F1_F2_%"]

            # Row background by F2 ROI
            if f2_roi >= 15:
                row_fill = PatternFill("solid", fgColor="C6EFCE")
                row_font = Font(color="006100")
            elif f2_roi >= 5:
                row_fill = PatternFill("solid", fgColor="EBFCD8")
                row_font = Font(color="2D6A04")
            elif f2_roi >= 0:
                row_fill = PatternFill("solid", fgColor="FFFFE0")
                row_font = Font(color="7D6608")
            elif f2_roi >= -10:
                row_fill = PatternFill("solid", fgColor="FFE5CC")
                row_font = Font(color="8B3A00")
            else:
                row_fill = PatternFill("solid", fgColor="FFC7CE")
                row_font = Font(color="9C0006")

            for ci, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.fill = row_fill
                cell.font = row_font
                cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        col_widths = {
            "ODDS_MIN": 10, "GAP_MIN": 9, "EV_MIN_%": 9,
            "N_all": 7, "WR_all_%": 9, "FlatROI_all_%": 13, "BK_all_$": 10,
            "N_F1": 6, "WR_F1_%": 8, "FlatROI_F1_%": 12, "KellyROI_F1_%": 13,
            "N_F2": 6, "WR_F2_%": 8, "FlatROI_F2_%": 12, "KellyROI_F2_%": 13,
            "MaxDD_F2_%": 11, "Avg_F1_F2_%": 12,
        }
        for ci, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 10)

        xlsx_path = "experiments/results/full_grid.xlsx"
        wb.save(xlsx_path)
        logger.info(f"Excel: {xlsx_path}")

    except ImportError:
        logger.warning("openpyxl not installed")

    logger.info(f"Done. Top-5 configs:")
    for _, r in out.head(5).iterrows():
        logger.info(
            f"  ODDS≥{r['ODDS_MIN']:.2f}  GAP≥{int(r['GAP_MIN'])}  EV≥{int(r['EV_MIN_%'])}%  "
            f"N={int(r['N_all'])}  F1={r['FlatROI_F1_%']:+.1f}%  F2={r['FlatROI_F2_%']:+.1f}%  "
            f"avg={r['Avg_F1_F2_%']:+.1f}%"
        )


if __name__ == "__main__":
    run()
