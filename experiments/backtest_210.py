"""
experiments/backtest_210.py

Бектест з ODDS_MIN=2.10 (замість 2.20).
GAP≥80, ODDS 2.10-2.60, AWAY only.
Генерує CSV + Excel з підсвіткою WIN/LOSS.

Запуск: python -m experiments.backtest_210
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DATABASE_URL', 'postgresql://capper:capper@localhost:5432/capper')

import pandas as pd
from loguru import logger

from model.train import load_data_from_db
from model.features.builder import build_dataset
from model.weighted_score import _get_factors, AWAY_WEIGHTS, HOME_WEIGHTS

SPLIT_1    = pd.Timestamp("2025-01-01")
SPLIT_2    = pd.Timestamp("2025-07-01")
INITIAL_BR = 1000.0
FRAC       = 0.25
CAP        = 0.04

WS_GAP_MIN = 80
ODDS_MIN   = 2.10
ODDS_MAX   = 2.60

TOP5_LEAGUES = {"Premier League", "La Liga", "Bundesliga", "Serie A", "Ligue 1"}


def compute_ws(features: dict, outcome: str) -> float:
    weights = HOME_WEIGHTS if outcome == "home" else AWAY_WEIGHTS
    return sum(weights.get(n, 1) for n, active in _get_factors(features, outcome) if active)


def _load_match_info():
    """Returns dict {match_id: (date, league_name, home_name, away_name)}"""
    from db.session import engine
    from sqlalchemy import text
    with engine.connect() as conn:
        rows = conn.execute(text("""
            SELECT m.id, m.date::date, l.name, ht.name, at.name,
                   m.home_score, m.away_score
            FROM matches m
            JOIN teams ht ON ht.id = m.home_team_id
            JOIN teams at ON at.id = m.away_team_id
            JOIN leagues l  ON l.id  = m.league_id
        """)).fetchall()
    return {r[0]: {"date": r[1], "league": r[2], "home": r[3], "away": r[4],
                   "hs": r[5], "as_": r[6]} for r in rows}


def run():
    logger.info("Loading data...")
    matches_df, stats_df, odds_df, teams, injuries_df = load_data_from_db()
    logger.info("Building dataset...")
    df = build_dataset(matches_df, stats_df, odds_df, teams, injuries_df)
    df["date"] = pd.to_datetime(df["date"])
    logger.info(f"Dataset: {len(df)} rows")

    logger.info("Loading match info (team names)...")
    match_info = _load_match_info()

    # --- collect bets ---
    bets = []
    for _, row in df.iterrows():
        r = row.to_dict()
        ws_h = compute_ws(r, "home")
        ws_a = compute_ws(r, "away")

        if ws_a <= ws_h:
            continue  # AWAY only

        gap = ws_a - ws_h
        if gap < WS_GAP_MIN:
            continue

        odds_val = r.get("away_odds")
        if not odds_val or not (ODDS_MIN <= odds_val <= ODDS_MAX):
            continue

        won = r.get("target") == "away"
        result = r.get("target", "?").upper()

        mid = int(r.get("match_id", 0))
        info = match_info.get(mid, {})
        league_name = info.get("league", "")

        if league_name not in TOP5_LEAGUES:
            continue

        p_elo = 1.0 - float(r.get("elo_home_win_prob", 0.5))

        bets.append({
            "match_id":   mid,
            "date":       row["date"],
            "league":     info.get("league", ""),
            "home_team":  info.get("home", ""),
            "away_team":  info.get("away", ""),
            "odds":       round(odds_val, 2),
            "ws_gap":     int(gap),
            "ws_away":    int(ws_a),
            "p_elo":      round(p_elo, 3),
            "result":     result,
            "won":        won,
        })

    bets = sorted(bets, key=lambda x: x["date"])
    logger.info(f"Bets collected: {len(bets)}")

    # --- simulate Kelly ---
    bankroll = INITIAL_BR
    rows_out = []
    for i, b in enumerate(bets, 1):
        p, odd = b["p_elo"], b["odds"]
        q = 1 - p
        kelly = max(0.0, (p * (odd - 1) - q) / (odd - 1)) * FRAC
        stake = round(min(bankroll * kelly, bankroll * CAP), 2)
        stake = max(stake, 1.0)

        bk_before = round(bankroll, 2)
        if b["won"]:
            pnl = round(stake * (odd - 1), 2)
        else:
            pnl = round(-stake, 2)
        bankroll += pnl

        rows_out.append({
            "N":             i,
            "Дата":          b["date"].strftime("%Y-%m-%d") if hasattr(b["date"], "strftime") else str(b["date"])[:10],
            "Ліга":          b["league"],
            "Ставка_на":     b["away_team"],
            "Проти":         b["home_team"],
            "Коеф":          b["odds"],
            "WS_Gap":        b["ws_gap"],
            "P_Elo":         b["p_elo"],
            "Результат":     b["result"],
            "WIN/LOSS":      "WIN" if b["won"] else "LOSS",
            "Банкрол_до":    bk_before,
            "Стейк_$":       stake,
            "PnL_$":         pnl,
            "Банкрол_$":     round(bankroll, 2),
            "Зміна_%":       round(pnl / bk_before * 100, 1),
            "Загальний_ROI_%": round((bankroll - INITIAL_BR) / INITIAL_BR * 100, 1),
        })

    out_df = pd.DataFrame(rows_out)

    # --- stats by fold ---
    all_bets = pd.DataFrame(bets)
    f1 = all_bets[(all_bets["date"] >= SPLIT_1) & (all_bets["date"] < SPLIT_2)]
    f2 = all_bets[all_bets["date"] >= SPLIT_2]

    def fold_stats(fb):
        if fb.empty: return "0 bets"
        n = len(fb); wr = fb["won"].mean()
        roi = ((fb["won"] * (fb["odds"] - 1)) - (~fb["won"])).sum() / n * 100
        return f"N={n}  WR={wr:.1%}  FlatROI={roi:+.1f}%"

    logger.info(f"=== GAP≥{WS_GAP_MIN}, ODDS {ODDS_MIN}-{ODDS_MAX}, AWAY, Top-5 only ===")
    logger.info(f"Всього: {len(bets)} бетів  WR={all_bets['won'].mean():.1%}  BK_final=${bankroll:.0f}")
    logger.info(f"F1 (Jan-Jun 2025): {fold_stats(f1)}")
    logger.info(f"F2 (Jul 2025+):    {fold_stats(f2)}")

    # by league
    logger.info("По лігах:")
    for lg, g in all_bets.groupby("league"):
        n = len(g); wr = g["won"].mean()
        roi = ((g["won"] * (g["odds"] - 1)) - (~g["won"])).sum() / n * 100
        logger.info(f"  {lg:<25} N={n:3d}  WR={wr:.1%}  FlatROI={roi:+.1f}%")

    # --- save CSV ---
    csv_path = "experiments/results/backtest_210_top5_bets.csv"
    out_df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    logger.info(f"CSV saved: {csv_path}")

    # --- save Excel with colors ---
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Backtest 2.10"

        header_fill = PatternFill("solid", fgColor="2F5496")
        header_font = Font(bold=True, color="FFFFFF")
        win_fill  = PatternFill("solid", fgColor="C6EFCE")
        win_font  = Font(bold=True, color="276221")
        loss_fill = PatternFill("solid", fgColor="FFC7CE")
        loss_font = Font(bold=True, color="9C0006")

        headers = list(out_df.columns)
        for col_i, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_i, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_i, row_data in out_df.iterrows():
            is_win = row_data["WIN/LOSS"] == "WIN"
            fill = win_fill if is_win else loss_fill
            font = win_font if is_win else loss_font
            for col_i, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_i + 2, column=col_i, value=val)
                cell.fill = fill
                cell.font = font

        ws.freeze_panes = "A2"
        for col_i, col in enumerate(out_df.columns, 1):
            ws.column_dimensions[get_column_letter(col_i)].width = max(len(str(col)) + 2, 10)

        xlsx_path = "experiments/results/backtest_210_top5_bets.xlsx"
        wb.save(xlsx_path)
        logger.info(f"Excel saved: {xlsx_path}")
    except ImportError:
        logger.warning("openpyxl not found, skipping Excel")

    logger.info(f"Done. $1000 → ${bankroll:.0f} ({(bankroll/INITIAL_BR - 1)*100:+.1f}%)")


if __name__ == "__main__":
    run()
