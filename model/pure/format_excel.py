"""
model/pure/format_excel.py

Adds color-coding, sorting, freeze panes, and conditional formatting to
profitable_niches.xlsx.

Color rules:
  wr           — green if >=70%, light-green 65-70%, light-red <60%
  wr_lower_95  — green if >=60%, light-green 55-60%
  roi          — green if >=30%, light-green 20-30%, red if <0
  ev_lower_95  — green if >0, red if <=0
  stable       — green TRUE / red FALSE
  n            — bold if >=50

Sort: each sheet by ROI descending (already done in mass_search, but enforce).
Freeze: top row + first 3 columns.
"""
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

REPORTS = Path(__file__).parent / "reports"

GREEN_DARK   = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
GREEN_MED    = PatternFill(start_color="74C69D", end_color="74C69D", fill_type="solid")
GREEN_LIGHT  = PatternFill(start_color="D8F3DC", end_color="D8F3DC", fill_type="solid")
YELLOW       = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")
RED_LIGHT    = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
RED_MED      = PatternFill(start_color="E63946", end_color="E63946", fill_type="solid")

HEADER_FILL  = PatternFill(start_color="264653", end_color="264653", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)

KEY_HEADER_FILL = PatternFill(start_color="E76F51", end_color="E76F51", fill_type="solid")  # accent for key cols

WHITE_FONT = Font(color="FFFFFF", bold=True)
BLACK_BOLD = Font(bold=True)

THIN = Side(border_style="thin", color="BBBBBB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# Columns we want to call out (will get accent fill on header)
KEY_COLS = {"wr", "wr_lower_95", "roi", "ev_lower_95", "n", "stable"}


def color_wr(value: float | None):
    if value is None:
        return None
    if value >= 0.75: return GREEN_DARK, WHITE_FONT
    if value >= 0.70: return GREEN_MED, None
    if value >= 0.65: return GREEN_LIGHT, None
    if value >= 0.60: return YELLOW, None
    return RED_LIGHT, None


def color_wr_lower(value):
    if value is None:
        return None
    if value >= 0.65: return GREEN_DARK, WHITE_FONT
    if value >= 0.60: return GREEN_MED, None
    if value >= 0.55: return GREEN_LIGHT, None
    if value >= 0.50: return YELLOW, None
    return RED_LIGHT, None


def color_roi(value):
    if value is None:
        return None
    if value >= 0.50: return GREEN_DARK, WHITE_FONT
    if value >= 0.30: return GREEN_MED, None
    if value >= 0.20: return GREEN_LIGHT, None
    if value >= 0.0:  return YELLOW, None
    return RED_LIGHT, None


def color_ev_lower(value):
    if value is None:
        return None
    if value >= 0.10: return GREEN_DARK, WHITE_FONT
    if value >= 0.05: return GREEN_MED, None
    if value > 0:     return GREEN_LIGHT, None
    return RED_LIGHT, None


def color_n(value):
    if value is None:
        return None
    if value >= 100: return GREEN_DARK, WHITE_FONT
    if value >= 50:  return GREEN_MED, None
    if value >= 30:  return GREEN_LIGHT, None
    return None


def color_stable(value):
    if value is True or value == "True":
        return GREEN_MED, None
    if value is False or value == "False":
        return RED_LIGHT, None
    return None


COL_FORMATTERS = {
    "wr":          (color_wr,       "0.0%"),
    "wr_lower_95": (color_wr_lower, "0.0%"),
    "roi":         (color_roi,      "+0.0%;-0.0%"),
    "ev_point":    (color_roi,      "+0.0%;-0.0%"),
    "ev_lower_95": (color_ev_lower, "+0.0%;-0.0%"),
    "n":           (color_n,        "0"),
    "wins":        (None,           "0"),
    "stable":      (color_stable,   None),
    "h1_wr":       (color_wr,       "0.0%"),
    "h2_wr":       (color_wr,       "0.0%"),
    "h1_n":        (None,           "0"),
    "h2_n":        (None,           "0"),
    "avg_odds":    (None,           "0.00"),
    "n_factors":   (None,           "0"),
}


def format_sheet(ws):
    if ws.max_row < 2:
        return

    # ── Header ───────────────────────────────────────────────────────────
    headers = {}
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        col_name = (cell.value or "").strip()
        headers[col_name] = c
        cell.fill = KEY_HEADER_FILL if col_name in KEY_COLS else HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    # Row height for header
    ws.row_dimensions[1].height = 28

    # ── Body cells ───────────────────────────────────────────────────────
    for r in range(2, ws.max_row + 1):
        for col_name, col_idx in headers.items():
            cell = ws.cell(row=r, column=col_idx)
            fmt = COL_FORMATTERS.get(col_name)
            if fmt is None:
                continue
            color_fn, num_fmt = fmt
            if num_fmt:
                cell.number_format = num_fmt
            if color_fn is not None and cell.value is not None:
                style = color_fn(cell.value)
                if style is not None:
                    fill, font = style
                    cell.fill = fill
                    if font is not None:
                        cell.font = font

    # ── Column widths ────────────────────────────────────────────────────
    width_overrides = {
        "league":      18,
        "side":         6,
        "niche_id":    44,
        "factors":     32,
        "n_factors":    8,
        "n":            7,
        "wins":         7,
        "wr":           8,
        "wr_lower_95": 11,
        "avg_odds":    10,
        "roi":          9,
        "ev_point":     9,
        "ev_lower_95": 11,
        "h1_n":         7,
        "h1_wr":        9,
        "h2_n":         7,
        "h2_wr":        9,
        "stable":       8,
    }
    for col_name, col_idx in headers.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width_overrides.get(col_name, 12)

    # ── Freeze first row + first 3 columns (league/side/niche_id) ───────
    ws.freeze_panes = "D2"

    # ── Auto-filter ──────────────────────────────────────────────────────
    last_col_letter = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"


def run() -> None:
    path = REPORTS / "profitable_niches.xlsx"
    wb = load_workbook(path)
    print(f"Loaded {path.name} — {len(wb.sheetnames)} sheets")
    for name in wb.sheetnames:
        ws = wb[name]
        format_sheet(ws)
        print(f"  Formatted: {name}  ({ws.max_row - 1} rows)")
    wb.save(path)
    print(f"\n✅ Saved with formatting → {path}")


if __name__ == "__main__":
    run()
