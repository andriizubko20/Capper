"""
model/pure/add_efficiency.py

Augment profitable_niches.xlsx with practical efficiency metrics:
  annual_units                 — flat-stake yearly profit (n × roi / years)
  annual_units_conservative    — same but with Wilson lower-95 EV (worst case)
  efficiency_score             — annual_units_cons × (1.0 if stable else 0.5)

Re-sort each sheet by efficiency_score descending.
"""
from pathlib import Path

import pandas as pd
from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPORTS = Path(__file__).parent / "reports"

# Years of historical data (matches start 2023-08, end 2026-04 ≈ 32 months)
HISTORY_YEARS = 32 / 12.0

GREEN_DARK   = PatternFill(start_color="2D6A4F", end_color="2D6A4F", fill_type="solid")
GREEN_MED    = PatternFill(start_color="74C69D", end_color="74C69D", fill_type="solid")
GREEN_LIGHT  = PatternFill(start_color="D8F3DC", end_color="D8F3DC", fill_type="solid")
YELLOW       = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")
RED_LIGHT    = PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid")
HEADER_FILL  = PatternFill(start_color="264653", end_color="264653", fill_type="solid")
EFF_HEADER   = PatternFill(start_color="9D4EDD", end_color="9D4EDD", fill_type="solid")  # purple for efficiency cols
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
WHITE_BOLD   = Font(color="FFFFFF", bold=True)


def color_annual(value):
    if value is None or pd.isna(value):
        return None
    if value >= 10:  return GREEN_DARK, WHITE_BOLD
    if value >= 5:   return GREEN_MED, None
    if value >= 2:   return GREEN_LIGHT, None
    if value >= 0:   return YELLOW, None
    return RED_LIGHT, None


def augment_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["bets_per_year"]            = (df["n"] / HISTORY_YEARS).round(2)
    df["annual_units"]             = (df["n"] / HISTORY_YEARS * df["roi"]).round(3)
    df["annual_units_conservative"] = (df["n"] / HISTORY_YEARS * df["ev_lower_95"]).round(3)
    df["efficiency_score"] = (
        df["annual_units_conservative"]
        * df["stable"].apply(lambda s: 1.0 if (s is True or s == "True" or s == True) else 0.5)
    ).round(3)
    return df


def reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    preferred = [
        "league", "side", "niche_id",
        "efficiency_score", "annual_units", "annual_units_conservative", "bets_per_year",
        "n", "wins", "wr", "wr_lower_95", "avg_odds",
        "roi", "ev_point", "ev_lower_95",
        "h1_n", "h1_wr", "h2_n", "h2_wr", "stable",
        "n_factors", "factors",
    ]
    cols = [c for c in preferred if c in df.columns] + [c for c in df.columns if c not in preferred]
    return df[cols]


def write_sheet(writer, sheet_name: str, df: pd.DataFrame):
    df = augment_df(df)
    df = reorder_columns(df)
    df = df.sort_values("efficiency_score", ascending=False).reset_index(drop=True)
    df.to_excel(writer, sheet_name=sheet_name, index=False)


def style_workbook(path: Path) -> None:
    """Add formatting to all sheets (colors, freeze, filter)."""
    wb = load_workbook(path)
    EFF_COLS = {"efficiency_score", "annual_units", "annual_units_conservative"}
    PCT_COLS = {"wr", "wr_lower_95", "h1_wr", "h2_wr", "roi", "ev_point", "ev_lower_95"}
    INT_COLS = {"n", "wins", "h1_n", "h2_n", "n_factors"}
    NUM_COLS = {"avg_odds", "bets_per_year", "annual_units", "annual_units_conservative", "efficiency_score"}

    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        # Header
        headers = {}
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=c)
            name = (cell.value or "").strip()
            headers[name] = c
            cell.fill = EFF_HEADER if name in EFF_COLS else HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28

        # Body
        for r in range(2, ws.max_row + 1):
            for name, col_idx in headers.items():
                cell = ws.cell(row=r, column=col_idx)
                if cell.value is None:
                    continue
                if name in PCT_COLS:
                    cell.number_format = "+0.0%;-0.0%" if "ev" in name or name == "roi" else "0.0%"
                elif name in NUM_COLS:
                    cell.number_format = "0.00"
                elif name in INT_COLS:
                    cell.number_format = "0"

                # Color code key columns
                if name in EFF_COLS:
                    style = color_annual(cell.value)
                    if style:
                        fill, font = style
                        cell.fill = fill
                        if font: cell.font = font
                elif name == "wr":
                    if   cell.value >= 0.75: cell.fill = GREEN_DARK; cell.font = WHITE_BOLD
                    elif cell.value >= 0.70: cell.fill = GREEN_MED
                    elif cell.value >= 0.65: cell.fill = GREEN_LIGHT
                    elif cell.value >= 0.60: cell.fill = YELLOW
                elif name == "wr_lower_95":
                    if   cell.value >= 0.65: cell.fill = GREEN_DARK; cell.font = WHITE_BOLD
                    elif cell.value >= 0.60: cell.fill = GREEN_MED
                    elif cell.value >= 0.55: cell.fill = GREEN_LIGHT
                elif name == "roi":
                    if   cell.value >= 0.50: cell.fill = GREEN_DARK; cell.font = WHITE_BOLD
                    elif cell.value >= 0.30: cell.fill = GREEN_MED
                    elif cell.value >= 0.20: cell.fill = GREEN_LIGHT
                    elif cell.value < 0:     cell.fill = RED_LIGHT
                elif name == "ev_lower_95":
                    if   cell.value >= 0.10: cell.fill = GREEN_DARK; cell.font = WHITE_BOLD
                    elif cell.value >= 0.05: cell.fill = GREEN_MED
                    elif cell.value > 0:     cell.fill = GREEN_LIGHT
                    else:                    cell.fill = RED_LIGHT
                elif name == "stable":
                    if cell.value is True or cell.value == "True":
                        cell.fill = GREEN_MED
                    else:
                        cell.fill = RED_LIGHT

        # Widths
        for name, col_idx in headers.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = {
                "league": 18, "side": 6, "niche_id": 44,
                "efficiency_score": 12, "annual_units": 11, "annual_units_conservative": 14,
                "bets_per_year": 9, "n": 7, "wins": 7,
                "wr": 8, "wr_lower_95": 11, "avg_odds": 10,
                "roi": 9, "ev_point": 9, "ev_lower_95": 11,
                "h1_n": 7, "h1_wr": 9, "h2_n": 7, "h2_wr": 9,
                "stable": 8, "n_factors": 8, "factors": 30,
            }.get(name, 12)

        ws.freeze_panes = "D2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    wb.save(path)


def run() -> None:
    in_path = REPORTS / "profitable_niches.xlsx"
    out_path = REPORTS / "profitable_niches.xlsx"  # overwrite

    # Read all sheets
    xl = pd.ExcelFile(in_path)
    sheet_dfs = {name: xl.parse(name) for name in xl.sheet_names}

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for name, df in sheet_dfs.items():
            write_sheet(writer, name, df)

    style_workbook(out_path)
    logger.info(f"Saved with efficiency metrics → {out_path}")


if __name__ == "__main__":
    run()
