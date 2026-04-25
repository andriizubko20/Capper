"""
model/pure/mark_selected.py

Mark niches that are already in user-curated production list (selected_niches.py)
in the Excel file with:
  - "in_prod" column: ✅ if matched
  - Cell highlighted yellow on niche_id column
"""
from pathlib import Path

from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

from model.pure.selected_niches import RAW_NICHES

REPORTS = Path(__file__).parent / "reports"

YELLOW_HIGHLIGHT = PatternFill(start_color="FFD60A", end_color="FFD60A", fill_type="solid")
GOLD_FONT = Font(color="000000", bold=True)


def selected_set() -> set[tuple[str, str]]:
    """Returns set of (league, niche_id) for niches in selected_niches.py."""
    out = set()
    for league, niches in RAW_NICHES.items():
        for n in niches:
            out.add((league, n.strip()))
    return out


def run() -> None:
    src = REPORTS / "profitable_niches_recent.xlsx"
    if not src.exists():
        logger.error(f"Source missing: {src}")
        return

    wb = load_workbook(src)
    sel = selected_set()

    for ws in wb.worksheets:
        if ws.max_row < 2:
            continue
        # Find niche_id column + league column
        headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        niche_col = headers.get("niche_id")
        league_col = headers.get("league")
        if niche_col is None:
            continue

        # Insert "in_prod" column at the very end
        new_col = ws.max_column + 1
        cell = ws.cell(row=1, column=new_col, value="in_prod")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFD60A", end_color="FFD60A", fill_type="solid")
        cell.font = Font(bold=True, color="000000")

        marked = 0
        for r in range(2, ws.max_row + 1):
            niche_id_val = ws.cell(row=r, column=niche_col).value
            if niche_id_val is None:
                continue
            league_val = (
                ws.cell(row=r, column=league_col).value if league_col
                else ws.title  # per-league sheets use title
            )
            key = (str(league_val).strip(), str(niche_id_val).strip())
            if key in sel:
                # Highlight niche_id cell
                ws.cell(row=r, column=niche_col).fill = YELLOW_HIGHLIGHT
                ws.cell(row=r, column=niche_col).font = GOLD_FONT
                # Mark in_prod column
                c = ws.cell(row=r, column=new_col, value="✅")
                c.fill = YELLOW_HIGHLIGHT
                c.font = GOLD_FONT
                marked += 1

        # Update auto-filter range to include new column
        from openpyxl.utils import get_column_letter
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        ws.column_dimensions[get_column_letter(new_col)].width = 8
        logger.info(f"  {ws.title}: {marked} niches marked as in_prod")

    wb.save(src)
    logger.info(f"Saved → {src}")


if __name__ == "__main__":
    run()
