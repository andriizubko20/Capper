import warnings; warnings.filterwarnings('ignore')
import pandas as pd
import numpy as np
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
sys.path.insert(0, '.')
from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix

OUT        = '/app/BETA/monthly_backtest_v4.xlsx'
FLAT_STAKE = 50.0
OOS_START  = pd.Timestamp('2025-08-01')
OVERLAP_THRESHOLD = 0.85

LEAGUES = [
    'Premier League','Bundesliga','Serie A','La Liga','Ligue 1',
    'Primeira Liga','Serie B','Eredivisie','Jupiler Pro League','Champions League'
]
ODDS_RANGES = [(1.30,1.55),(1.55,1.80),(1.70,2.00),(1.80,2.20),(2.00,2.50),(2.20,2.80),(2.50,3.50)]
XG = [0.0,1.0,1.2,1.5,1.8]
EH = [0,30,75,150]; EA = [0,-30,-75,-150]
FT = [0.0,1.5,1.8,2.2]; MH = [0.0,0.45,0.50,0.55]; MA = [0.0,0.35,0.40,0.45]

print('Loading data...')
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result']=='H').astype(int)
df['away_win']  = (df['result']=='A').astype(int)
df['ym'] = df['date'].dt.to_period('M')
df['_id'] = np.arange(len(df), dtype=np.int32)

all_months = pd.period_range(df['date'].min().to_period('M'),
                              df['date'].max().to_period('M'), freq='M')
oos_p = OOS_START.to_period('M')
print(f'Matches: {len(df)}  |  Months: {len(all_months)}  |  OOS from: {oos_p}')

def niche_label(side,lo,hi,xg,elo,fm,mk):
    s=f'{side}[{lo},{hi})'
    if xg>0:   s+=f' xg>={xg}'
    if elo!=0: s+=f' elo>={elo}' if elo>0 else f' elo<={elo}'
    if fm>0:   s+=f' form>={fm}'
    if mk>0:   s+=f' mkt>={mk}'
    return s

def apply_mask(data,side,lo,hi,xg,elo,fm,mk,league):
    oc='home_odds_val' if side=='home' else 'away_odds_val'
    m=(data['league_name']==league)&(data[oc]>=lo)&(data[oc]<hi)
    if xg>0:  m&=data['xg_ratio_home_5' if side=='home' else 'xg_ratio_away_5'].fillna(0)>=xg
    if side=='home' and elo>0: m&=data['elo_diff'].fillna(0)>=elo
    if side=='away' and elo<0: m&=data['elo_diff'].fillna(0)<=elo
    if fm>0:  m&=data['home_pts_5' if side=='home' else 'away_pts_5'].fillna(0)>=fm
    if mk>0:  m&=data['mkt_home_prob' if side=='home' else 'mkt_away_prob'].fillna(0)>=mk
    return data[m]

def calc_stats(bets, wc, oc):
    monthly={}
    for ym,grp in bets.groupby('ym'):
        pl=(grp[wc]*(grp[oc]-1)-(1-grp[wc])).sum()*FLAT_STAKE
        n=len(grp); w=int(grp[wc].sum())
        monthly[ym]={'n':n,'w':w,'pl':round(pl,1),'roi':round(pl/(n*FLAT_STAKE)*100,1)}
    total_pl=sum(v['pl'] for v in monthly.values())
    total_n =sum(v['n']  for v in monthly.values())
    total_w =sum(v['w']  for v in monthly.values())
    ins_items=[(m,v) for m,v in monthly.items() if m<oos_p]
    oos_items=[(m,v) for m,v in monthly.items() if m>=oos_p]
    ins_pl=sum(v['pl'] for _,v in ins_items); ins_n=sum(v['n'] for _,v in ins_items); ins_w=sum(v['w'] for _,v in ins_items)
    oos_pl=sum(v['pl'] for _,v in oos_items); oos_n=sum(v['n'] for _,v in oos_items); oos_w=sum(v['w'] for _,v in oos_items)
    return dict(
        total_n=total_n, total_w=total_w, total_pl=total_pl,
        roi_pct=total_pl/(total_n*FLAT_STAKE)*100 if total_n else 0,
        wr_pct=total_w/total_n*100 if total_n else 0,
        ins_roi=ins_pl/(ins_n*FLAT_STAKE)*100 if ins_n else 0,
        ins_wr=ins_w/ins_n*100 if ins_n else 0,
        oos_roi=oos_pl/(oos_n*FLAT_STAKE)*100 if oos_n else 0,
        oos_wr=oos_w/oos_n*100 if oos_n else 0,
        oos_n=oos_n, ins_n=ins_n, monthly=monthly,
    )

print('Running combinations with group-level dedup...')
kept = []

for side in ('home','away'):
    et=EH if side=='home' else EA; mt=MH if side=='home' else MA
    wc='home_win' if side=='home' else 'away_win'
    oc='home_odds_val' if side=='home' else 'away_odds_val'
    for lo,hi in ODDS_RANGES:
        for lg in LEAGUES:
            # Collect all profitable niches in this (side, odds, league) group
            group = []
            for xg,elo,fm,mk in product(XG,et,FT,mt):
                bets = apply_mask(df,side,lo,hi,xg,elo,fm,mk,lg)
                if len(bets)<5: continue
                s = calc_stats(bets,wc,oc)
                if s['total_pl']<=0: continue
                n_f = niche_label(side,lo,hi,xg,elo,fm,mk).count('>=') + \
                      niche_label(side,lo,hi,xg,elo,fm,mk).count('<=')
                group.append({
                    'league':lg,'side':side,'lo':lo,'hi':hi,
                    'niche':niche_label(side,lo,hi,xg,elo,fm,mk),
                    'n_filters':n_f,
                    'ids': bets['_id'].values,  # numpy array, not frozenset
                    **s,
                })

            if not group: continue

            # Dedup within group: best OOS ROI first
            group.sort(key=lambda x:(-x['oos_roi'],-x['total_n']))
            kept_ids = []   # list of numpy arrays
            for r in group:
                is_dup = False
                for kid in kept_ids:
                    common = np.intersect1d(r['ids'], kid, assume_unique=True)
                    coverage = len(common)/len(r['ids']) if len(r['ids'])>0 else 0
                    if coverage >= OVERLAP_THRESHOLD:
                        is_dup = True; break
                if not is_dup:
                    kept_ids.append(r['ids'])
                    del r['ids']
                    kept.append(r)
                else:
                    del r['ids']

print(f'After dedup: {len(kept)}')
kept.sort(key=lambda x:(x['league'],-x['oos_roi']))

# ── Styles ─────────────────────────────────────────────────────────────────────
fill_hdr   = PatternFill('solid',fgColor='1F3864')
fill_oos_h = PatternFill('solid',fgColor='2E4057')
fill_lg    = PatternFill('solid',fgColor='D6E4F0')
fill_ps    = PatternFill('solid',fgColor='70AD47')
fill_p     = PatternFill('solid',fgColor='C6EFCE')
fill_neg   = PatternFill('solid',fgColor='FFCCCC')
fill_zero  = PatternFill('solid',fgColor='F5F5F5')
thin       = Side(style='thin',color='CCCCCC')
border     = Border(left=thin,right=thin,top=thin,bottom=thin)

META = ['Niche','Side','Total n','Total P&L','ROI%','WR%',
        'IS ROI%','IS WR%','OOS ROI%','OOS WR%','OOS n']
N_META = len(META)

def write_league_sheet(wb, league_name, rows):
    ws = wb.create_sheet(league_name[:20])
    for ci,col in enumerate(META,1):
        c=ws.cell(1,ci,col); c.fill=fill_hdr
        c.font=Font(color='FFFFFF',bold=True,size=10)
        c.alignment=Alignment(horizontal='center',wrap_text=True)
    for mi,m in enumerate(all_months):
        base=N_META+1+mi*3; lbl=m.strftime('%b-%y')
        hf=fill_oos_h if m>=oos_p else fill_hdr
        for off,sub in enumerate(['n','P&L','ROI%']):
            c=ws.cell(1,base+off,f'{lbl}\n{sub}')
            c.fill=hf; c.font=Font(color='FFFFFF',bold=True,size=8)
            c.alignment=Alignment(horizontal='center',wrap_text=True)
    ws.row_dimensions[1].height=28
    row_idx=2
    for r in rows:
        roi=r['roi_pct']; oos_roi=r['oos_roi']; ins_roi=r['ins_roi']
        mv=[r['niche'],r['side'],r['total_n'],round(r['total_pl'],0),
            round(roi,1),round(r['wr_pct'],1),
            round(ins_roi,1),round(r['ins_wr'],1),
            round(oos_roi,1),round(r['oos_wr'],1),int(r['oos_n'])]
        for ci,val in enumerate(mv,1):
            c=ws.cell(row_idx,ci,val); c.border=border
            c.alignment=Alignment(horizontal='center')
            if ci==1: c.alignment=Alignment(horizontal='left')
            if ci in (5,6): c.number_format='+0.0;-0.0;0'
            if ci==7:
                c.number_format='+0.0;-0.0;0'
                c.fill=fill_ps if ins_roi>=15 else (fill_p if ins_roi>0 else fill_neg)
            if ci==8: c.number_format='0.0'
            if ci==9:
                c.number_format='+0.0;-0.0;0'
                c.fill=fill_ps if oos_roi>=15 else (fill_p if oos_roi>0 else fill_neg)
                c.font=Font(bold=True)
            if ci==10: c.number_format='0.0'; c.font=Font(bold=True)
        for mi,m in enumerate(all_months):
            base=N_META+1+mi*3; md=r['monthly'].get(m)
            if md is None:
                for off in range(3):
                    c=ws.cell(row_idx,base+off,''); c.fill=fill_zero; c.border=border
            else:
                c=ws.cell(row_idx,base,md['n']); c.border=border
                c.alignment=Alignment(horizontal='center'); c.font=Font(size=8)
                c=ws.cell(row_idx,base+1,md['pl']); c.border=border
                c.number_format='+0;-0;0'; c.font=Font(size=8); c.alignment=Alignment(horizontal='center')
                c.fill=fill_ps if md['pl']>=100 else (fill_p if md['pl']>0 else (fill_neg if md['pl']<0 else fill_zero))
                c=ws.cell(row_idx,base+2,md['roi']); c.border=border
                c.number_format='+0;-0;0'; c.font=Font(size=8); c.alignment=Alignment(horizontal='center')
                c.fill=fill_ps if md['roi']>=20 else (fill_p if md['roi']>0 else (fill_neg if md['roi']<0 else fill_zero))
        row_idx+=1
    ws.column_dimensions['A'].width=42; ws.column_dimensions['B'].width=6
    for ci in range(3,N_META+1): ws.column_dimensions[get_column_letter(ci)].width=9
    for mi in range(len(all_months)):
        for off in range(3): ws.column_dimensions[get_column_letter(N_META+1+mi*3+off)].width=4.5
    ws.freeze_panes=f'{get_column_letter(N_META+1)}2'
    ws.auto_filter.ref=f'A1:{get_column_letter(N_META)}1'

print('Writing Excel...')
wb=Workbook(); wb.remove(wb.active)
by_league={}
for r in kept: by_league.setdefault(r['league'],[]).append(r)
for lg in LEAGUES:
    rows=by_league.get(lg,[])
    if not rows: continue
    write_league_sheet(wb,lg,rows)
    print(f'  {lg}: {len(rows)} niches')

ws_sum=wb.create_sheet('SUMMARY',0)
sum_cols=['League','Niche','Side','Total n','ROI%','WR%','IS ROI%','IS WR%','OOS ROI%','OOS WR%','OOS n']
for ci,col in enumerate(sum_cols,1):
    c=ws_sum.cell(1,ci,col); c.fill=fill_hdr
    c.font=Font(color='FFFFFF',bold=True,size=10)
    c.alignment=Alignment(horizontal='center',wrap_text=True)
top=[r for r in kept if r['oos_n']>=5]; top.sort(key=lambda x:-x['oos_roi']); top=top[:150]
prev_lg=None; row_idx=2
for r in top:
    if r['league']!=prev_lg:
        c=ws_sum.cell(row_idx,1,r['league']); c.fill=fill_lg; c.font=Font(bold=True,size=11)
        ws_sum.merge_cells(start_row=row_idx,start_column=1,end_row=row_idx,end_column=len(sum_cols))
        row_idx+=1; prev_lg=r['league']
    oos_roi=r['oos_roi']; ins_roi=r['ins_roi']
    vals=[r['league'],r['niche'],r['side'],r['total_n'],round(r['roi_pct'],1),round(r['wr_pct'],1),
          round(ins_roi,1),round(r['ins_wr'],1),round(oos_roi,1),round(r['oos_wr'],1),int(r['oos_n'])]
    for ci,val in enumerate(vals,1):
        c=ws_sum.cell(row_idx,ci,val); c.border=border; c.alignment=Alignment(horizontal='center')
        if ci==2: c.alignment=Alignment(horizontal='left')
        if ci in (5,6): c.number_format='+0.0;-0.0;0'
        if ci==7:
            c.number_format='+0.0;-0.0;0'
            c.fill=fill_ps if ins_roi>=15 else (fill_p if ins_roi>0 else fill_neg)
        if ci==8: c.number_format='0.0'
        if ci==9:
            c.number_format='+0.0;-0.0;0'
            c.fill=fill_ps if oos_roi>=20 else (fill_p if oos_roi>0 else fill_neg)
            c.font=Font(bold=True)
        if ci==10: c.number_format='0.0'; c.font=Font(bold=True)
    row_idx+=1
ws_sum.column_dimensions['A'].width=16; ws_sum.column_dimensions['B'].width=44
for ci in range(3,len(sum_cols)+1): ws_sum.column_dimensions[get_column_letter(ci)].width=10
ws_sum.freeze_panes='A2'; ws_sum.auto_filter.ref=f'A1:{get_column_letter(len(sum_cols))}1'

wb.save(OUT)
print(f'\nSaved: {OUT}')
print('✅ Done!')
