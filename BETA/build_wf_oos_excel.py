"""
BETA/build_wf_oos_excel.py

WF windows: 2023-08 → 2025-10 (9 вікон, строго до OOS — без leakage).
OOS period: 2025-11 → 2026-04 (реальний holdout).

Структура таблиці:
  Filter | Side | WF: n_tot | Avg_ROI% | Avg_WR% | Win% | W+/Wtot | Stable
  | OOS: n | WR% | ROI% | EV% | ΔROI
  | Per WF window: ROI% | n  (×9)
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import numpy as np
import pandas as pd
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix

# ── Config ────────────────────────────────────────────────────────────────────
WF_START  = pd.Timestamp('2023-08-01')
WF_END    = pd.Timestamp('2025-10-31')   # строго до OOS
OOS_START = pd.Timestamp('2025-11-01')

# ── Load data ─────────────────────────────────────────────────────────────────
print("Loading data...")
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result'] == 'H').astype(int)
df['away_win']  = (df['result'] == 'A').astype(int)

disc_df = df[df['date'] <= WF_END].copy()
oos_df  = df[df['date'] >= OOS_START].copy()
print(f"Discovery: {len(disc_df)} matches ({disc_df['date'].min().date()} – {disc_df['date'].max().date()})")
print(f"OOS:       {len(oos_df)} matches ({oos_df['date'].min().date()} – {oos_df['date'].max().date()})")

LEAGUES = [
    'Premier League','Bundesliga','Serie A','La Liga','Ligue 1',
    'Primeira Liga','Serie B','Eredivisie','Jupiler Pro League','Champions League'
]
LEAGUE_SHORT = {
    'Premier League':'EPL','Bundesliga':'Bundesliga','Serie A':'Serie A',
    'La Liga':'La Liga','Ligue 1':'Ligue 1','Primeira Liga':'Portugal',
    'Serie B':'Serie B','Eredivisie':'Eredivisie',
    'Jupiler Pro League':'Jupiler','Champions League':'UCL'
}

# ── WF Windows (2023-08 → 2025-10) ───────────────────────────────────────────
WINDOWS = []
cur = WF_START
while cur <= WF_END:
    w_end = cur + pd.DateOffset(months=3) - pd.Timedelta(days=1)
    if w_end > WF_END: w_end = WF_END
    WINDOWS.append((cur.strftime('%Y-%m'), cur, w_end))
    cur += pd.DateOffset(months=3)
WIN_LABELS = [w[0] for w in WINDOWS]
print(f"WF windows: {len(WINDOWS)} ({WIN_LABELS[0]} → {WIN_LABELS[-1]})")

win_slices = {wlbl: disc_df[(disc_df['date'] >= ws) & (disc_df['date'] <= we)]
              for wlbl, ws, we in WINDOWS}

# ── Filter space ──────────────────────────────────────────────────────────────
ODDS_RANGES = [
    (1.30,1.55),(1.55,1.80),(1.70,2.00),
    (1.80,2.20),(2.00,2.50),(2.20,2.80),(2.50,3.50),
]
XG_THRS    = [0.0,1.0,1.2,1.5,1.8]
ELO_H_THRS = [0,30,75,150]
ELO_A_THRS = [0,-30,-75,-150]
FORM_THRS  = [0.0,1.5,1.8,2.2]
MKT_H_THRS = [0.0,0.45,0.50,0.55]
MKT_A_THRS = [0.0,0.35,0.40,0.45]
MIN_N_WIN  = 3

def calc_stats(sub, side):
    wc = 'home_win' if side == 'home' else 'away_win'
    oc = 'home_odds_val' if side == 'home' else 'away_odds_val'
    if len(sub) < MIN_N_WIN:
        return None
    wr  = sub[wc].mean()
    ao  = sub[oc].mean()
    roi = (sub[wc] * (sub[oc] - 1) - (1 - sub[wc])).mean() * 100
    ev  = (wr * ao - 1) * 100
    return {'n': len(sub), 'wr': round(wr*100,1), 'roi': round(roi,1), 'ev': round(ev,1)}

def apply_mask(d, side, lo, hi, xg_t, elo_t, form_t, mkt_t):
    oc  = 'home_odds_val'   if side == 'home' else 'away_odds_val'
    xgc = 'xg_ratio_home_5' if side == 'home' else 'xg_ratio_away_5'
    fc  = 'home_pts_5'      if side == 'home' else 'away_pts_5'
    mc  = 'mkt_home_prob'   if side == 'home' else 'mkt_away_prob'
    m = (d[oc] >= lo) & (d[oc] < hi)
    if xg_t > 0:                          m &= (d[xgc].fillna(0) >= xg_t)
    if side == 'home' and elo_t > 0:      m &= (d['elo_diff'].fillna(0) >= elo_t)
    if side == 'away' and elo_t < 0:      m &= (d['elo_diff'].fillna(0) <= elo_t)
    if form_t > 0:                        m &= (d[fc].fillna(0) >= form_t)
    if mkt_t > 0:                         m &= (d[mc].fillna(0) >= mkt_t)
    return d[m]

# ── Sweep ─────────────────────────────────────────────────────────────────────
print("Running sweep (WF + OOS)...")
# results[lg][label] = {meta, wlbl: stats, 'oos': stats_or_None}
results = {lg: {} for lg in LEAGUES}

for side in ('home', 'away'):
    et = ELO_H_THRS if side == 'home' else ELO_A_THRS
    mt = MKT_H_THRS if side == 'home' else MKT_A_THRS

    for lo, hi in ODDS_RANGES:
        for xg_t, elo_t, form_t, mkt_t in product(XG_THRS, et, FORM_THRS, mt):
            parts = [f'{side}[{lo},{hi})']
            if xg_t > 0:                     parts.append(f'xg>={xg_t}')
            if side == 'home' and elo_t > 0: parts.append(f'elo>={elo_t}')
            if side == 'away' and elo_t < 0: parts.append(f'elo<={elo_t}')
            if form_t > 0:                   parts.append(f'form>={form_t}')
            if mkt_t > 0:                    parts.append(f'mkt>={mkt_t}')
            label = ' '.join(parts)

            # WF windows
            for wlbl, _, _ in WINDOWS:
                sub_w = apply_mask(win_slices[wlbl], side, lo, hi, xg_t, elo_t, form_t, mkt_t)
                for lg in LEAGUES:
                    sub_l = sub_w[sub_w['league_name'] == lg]
                    m = calc_stats(sub_l, side)
                    if m:
                        if label not in results[lg]:
                            results[lg][label] = {
                                'side': side, 'lo': lo, 'hi': hi,
                                'xg_t': xg_t, 'elo_t': elo_t,
                                'form_t': form_t, 'mkt_t': mkt_t
                            }
                        results[lg][label][wlbl] = m

            # OOS
            sub_oos = apply_mask(oos_df, side, lo, hi, xg_t, elo_t, form_t, mkt_t)
            for lg in LEAGUES:
                sub_l = sub_oos[sub_oos['league_name'] == lg]
                m = calc_stats(sub_l, side)
                if label not in results[lg]:
                    if m:
                        results[lg][label] = {
                            'side': side, 'lo': lo, 'hi': hi,
                            'xg_t': xg_t, 'elo_t': elo_t,
                            'form_t': form_t, 'mkt_t': mkt_t
                        }
                if label in results[lg]:
                    results[lg][label]['oos'] = m  # None if < MIN_N_WIN

print("Sweep done.")

# ── Aggregate WF stats ────────────────────────────────────────────────────────
def agg_wf(data):
    wins = [(wl, data[wl]) for wl in WIN_LABELS if wl in data]
    if not wins:
        return None
    n_tot   = sum(w['n']   for _, w in wins)
    avg_roi = round(np.mean([w['roi'] for _, w in wins]), 1)
    avg_wr  = round(np.mean([w['wr']  for _, w in wins]), 1)
    n_pos   = sum(1 for _, w in wins if w['roi'] > 0)
    win_pct = round(n_pos / len(wins) * 100, 0)
    return {
        'n_tot': n_tot, 'avg_roi': avg_roi, 'avg_wr': avg_wr,
        'win_pct': win_pct, 'n_wins': n_pos, 'n_windows': len(wins)
    }

# ── Excel styles ──────────────────────────────────────────────────────────────
GREEN_LIGHT = PatternFill('solid', start_color='D5F5E3')
RED_LIGHT   = PatternFill('solid', start_color='FADBD8')
GREY_FILL   = PatternFill('solid', start_color='F2F3F4')
BLUE_LIGHT  = PatternFill('solid', start_color='D6EAF8')
YELLOW_FILL = PatternFill('solid', start_color='FEF9E7')
HEADER_FILL = PatternFill('solid', start_color='2C3E50')
OOS_FILL    = PatternFill('solid', start_color='1A5276')
WHITE_FILL  = PatternFill('solid', start_color='FFFFFF')

HDR_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
BODY_FONT = Font(name='Arial', size=9)
CENTER    = Alignment(horizontal='center', vertical='center')
LEFT      = Alignment(horizontal='left',   vertical='center')

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def hc(cell, val, fill=HEADER_FILL):
    cell.value = val; cell.font = HDR_FONT
    cell.fill = fill; cell.alignment = CENTER; cell.border = thin_border()

def bc(cell, val, align=CENTER, fill=None, bold=False):
    cell.value = val
    cell.font = Font(name='Arial', size=9, bold=bold)
    cell.alignment = align; cell.border = thin_border()
    if fill: cell.fill = fill

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ── SUMMARY sheet ─────────────────────────────────────────────────────────────
ws_sum = wb.create_sheet('SUMMARY')
SUM_HDRS = ['Ліга','Side','Найкращий фільтр',
            'WF n_tot','WF Avg_ROI%','WF WR%','WF Win%','WF W+/tot',
            'OOS n','OOS WR%','OOS ROI%','OOS EV%','Δ ROI','Stable?']
for ci, h in enumerate(SUM_HDRS, 1):
    hc(ws_sum.cell(1, ci), h)

sum_row = 2
for lg in LEAGUES:
    lname = LEAGUE_SHORT[lg]
    for side in ('home', 'away'):
        best = None
        for label, data in results[lg].items():
            if data.get('side') != side: continue
            a = agg_wf(data)
            if not a or a['n_tot'] < 10 or a['n_windows'] < 4: continue
            if best is None or a['avg_roi'] > best['avg_roi']:
                best = {**a, 'label': label, 'oos': data.get('oos')}
        if not best: continue

        oos = best['oos']
        delta = round((oos['roi'] - best['avg_roi']), 1) if oos else None
        stable = ('✅ ТАК' if best['avg_roi'] > 0 and best['win_pct'] >= 60 else
                  '⚠️ MIXED' if best['avg_roi'] > 0 else '❌ НІ')

        row = [
            lname, side, best['label'],
            best['n_tot'], best['avg_roi']/100, best['avg_wr']/100,
            best['win_pct']/100, f"{best['n_wins']}/{best['n_windows']}",
            oos['n']   if oos else '—',
            oos['wr']/100  if oos else '—',
            oos['roi']/100 if oos else '—',
            oos['ev']/100  if oos else '—',
            delta/100 if delta is not None else '—',
            stable
        ]
        fmts = [None,None,None, None,'+0.0%;-0.0%','0.0%','0%',None,
                None,'0.0%','+0.0%;-0.0%','+0.0%;-0.0%','+0.0%;-0.0%',None]
        for ci, (val, fmt) in enumerate(zip(row, fmts), 1):
            c = ws_sum.cell(sum_row, ci)
            bc(c, val, LEFT if ci in (1,2,3,14) else CENTER)
            if fmt and isinstance(val, (int, float)): c.number_format = fmt
            if ci == 5:  c.fill = GREEN_LIGHT if best['avg_roi'] > 0 else RED_LIGHT
            if ci == 11 and oos:
                c.fill = GREEN_LIGHT if oos['roi'] > 0 else RED_LIGHT
            if ci == 13 and delta is not None:
                c.fill = GREEN_LIGHT if delta >= 0 else RED_LIGHT
        sum_row += 1

ws_sum.column_dimensions['A'].width = 12
ws_sum.column_dimensions['B'].width = 6
ws_sum.column_dimensions['C'].width = 52
for ci in range(4, len(SUM_HDRS)+1):
    ws_sum.column_dimensions[get_column_letter(ci)].width = 11
ws_sum.freeze_panes = 'A2'
ws_sum.auto_filter.ref = f'A1:{get_column_letter(len(SUM_HDRS))}1'

# ── Per-league sheets ─────────────────────────────────────────────────────────
# Column layout:
#   [Filter][Side] | [WF: n_tot][Avg_ROI%][WF WR%][Win%][W+/tot][Stable]
#   | [OOS: n][WR%][ROI%][EV%][ΔROI]
#   | per window: [ROI%][n] ×9

STATIC_WF  = ['Filter','Side','WF n_tot','WF Avg_ROI%','WF WR%','WF Win%','W+/tot','Stable']
OOS_COLS   = ['OOS n','OOS WR%','OOS ROI%','OOS EV%','Δ ROI']
N_STATIC   = len(STATIC_WF)
N_OOS      = len(OOS_COLS)
N_PREFIX   = N_STATIC + N_OOS   # columns before per-window data

for lg in LEAGUES:
    lname = LEAGUE_SHORT[lg]
    print(f"  Sheet: {lname}...")
    ws = wb.create_sheet(lname)

    total_cols = N_PREFIX + len(WINDOWS) * 2

    # Row 1: static headers
    for ci, h in enumerate(STATIC_WF, 1):
        hc(ws.cell(1, ci), h)
    # Row 1: OOS group header (merged)
    oos_col_start = N_STATIC + 1
    ws.merge_cells(start_row=1, start_column=oos_col_start,
                   end_row=1,   end_column=oos_col_start + N_OOS - 1)
    hc(ws.cell(1, oos_col_start), '── OOS: 2025-11 → 2026-04 ──', fill=OOS_FILL)
    # Row 1: WF window group headers
    for wi, (wlbl, _, _) in enumerate(WINDOWS):
        cs = N_PREFIX + wi*2 + 1
        ws.merge_cells(start_row=1, start_column=cs, end_row=1, end_column=cs+1)
        hc(ws.cell(1, cs), wlbl)

    # Row 2: sub-headers
    for ci in range(1, N_STATIC+1):
        hc(ws.cell(2, ci), '', fill=PatternFill('solid', start_color='5D6D7E'))
    for ci, h in enumerate(OOS_COLS, oos_col_start):
        hc(ws.cell(2, ci), h, fill=PatternFill('solid', start_color='1F618D'))
    for wi in range(len(WINDOWS)):
        cs = N_PREFIX + wi*2 + 1
        hc(ws.cell(2, cs),   'ROI%', fill=PatternFill('solid', start_color='5D6D7E'))
        hc(ws.cell(2, cs+1), 'n',    fill=PatternFill('solid', start_color='5D6D7E'))

    # Collect rows
    all_rows = []
    for label, data in results[lg].items():
        if 'side' not in data: continue
        a = agg_wf(data)
        if not a: continue
        stable = ('✅' if a['avg_roi'] > 0 and a['win_pct'] >= 60 else
                  '⚠️'  if a['avg_roi'] > 0 else '❌')
        all_rows.append({'label': label, 'data': data, **a, 'stable': stable})

    def sort_key(r):
        g = (0 if r['avg_roi'] > 0 and r['win_pct'] >= 60 else
             1 if r['avg_roi'] > 0 else 2)
        return (g, -r['avg_roi'])
    all_rows.sort(key=sort_key)

    excel_row = 3
    prev_grp = None
    for r in all_rows:
        grp = (0 if r['avg_roi'] > 0 and r['win_pct'] >= 60 else
               1 if r['avg_roi'] > 0 else 2)

        if grp != prev_grp:
            titles = {
                0: '▶ СТАБІЛЬНО ПРИБУТКОВІ (avg_roi>0, win%≥60%)',
                1: '▶ ПОМІРНО ПРИБУТКОВІ (avg_roi>0, win%<60%)',
                2: '▶ ЗБИТКОВІ (avg_roi≤0)'
            }
            grp_colors = {0: '1A5276', 1: '7D6608', 2: '78281F'}
            ws.merge_cells(start_row=excel_row, start_column=1,
                           end_row=excel_row, end_column=total_cols)
            c = ws.cell(excel_row, 1)
            c.value = titles[grp]
            c.font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            c.fill  = PatternFill('solid', start_color=grp_colors[grp])
            c.alignment = LEFT
            excel_row += 1
            prev_grp = grp

        row_fill = (GREEN_LIGHT if grp == 0 else
                    YELLOW_FILL if grp == 1 else
                    PatternFill('solid', start_color='FDEDEC'))

        oos = r['data'].get('oos')
        delta = round(oos['roi'] - r['avg_roi'], 1) if oos else None

        # Static WF cells
        static_vals = [
            r['label'], r['data']['side'], r['n_tot'],
            r['avg_roi']/100, r['avg_wr']/100,
            r['win_pct']/100, f"{r['n_wins']}/{r['n_windows']}",
            r['stable']
        ]
        static_fmts = [None,None,None,'+0.0%;-0.0%','0.0%','0%',None,None]
        for ci, (val, fmt) in enumerate(zip(static_vals, static_fmts), 1):
            c = ws.cell(excel_row, ci)
            bc(c, val, LEFT if ci in (1,2,8) else CENTER, row_fill)
            if fmt and isinstance(val, (int, float)): c.number_format = fmt
            if ci == 4:
                c.fill = (GREEN_LIGHT if r['avg_roi'] > 0 else
                          RED_LIGHT   if r['avg_roi'] < 0 else WHITE_FILL)
                c.font = Font(name='Arial', size=9, bold=(abs(r['avg_roi']) > 15))

        # OOS cells
        oos_vals = [
            oos['n']        if oos else '—',
            oos['wr']/100   if oos else '—',
            oos['roi']/100  if oos else '—',
            oos['ev']/100   if oos else '—',
            delta/100       if delta is not None else '—',
        ]
        oos_fmts = [None,'0.0%','+0.0%;-0.0%','+0.0%;-0.0%','+0.0%;-0.0%']
        for i, (val, fmt) in enumerate(zip(oos_vals, oos_fmts)):
            ci = oos_col_start + i
            c  = ws.cell(excel_row, ci)
            bc(c, val, CENTER, BLUE_LIGHT)
            if fmt and isinstance(val, (int, float)): c.number_format = fmt
            if i == 2 and oos:  # OOS ROI
                c.fill = GREEN_LIGHT if oos['roi'] > 0 else RED_LIGHT
                c.font = Font(name='Arial', size=9, bold=(abs(oos['roi']) > 15))
            if i == 4 and delta is not None:  # Δ ROI
                c.fill = GREEN_LIGHT if delta >= 0 else RED_LIGHT

        # Per-window cells
        for wi, (wlbl, _, _) in enumerate(WINDOWS):
            cs = N_PREFIX + wi*2 + 1
            if wlbl in r['data']:
                d = r['data'][wlbl]
                c_roi = ws.cell(excel_row, cs)
                c_roi.value = d['roi']/100
                c_roi.font  = Font(name='Arial', size=9, bold=(abs(d['roi']) > 20))
                c_roi.number_format = '+0.0%;-0.0%;0.0%'
                c_roi.alignment = CENTER; c_roi.border = thin_border()
                c_roi.fill = GREEN_LIGHT if d['roi'] > 0 else RED_LIGHT
                c_n = ws.cell(excel_row, cs+1)
                bc(c_n, d['n'], fill=row_fill)
            else:
                bc(ws.cell(excel_row, cs),   '—', fill=GREY_FILL)
                bc(ws.cell(excel_row, cs+1), '—', fill=GREY_FILL)

        excel_row += 1

    # Column widths
    ws.column_dimensions['A'].width = 52
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 7
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 7
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 7
    for i in range(N_OOS):
        ws.column_dimensions[get_column_letter(oos_col_start + i)].width = 9
    for wi in range(len(WINDOWS)):
        cs = N_PREFIX + wi*2 + 1
        ws.column_dimensions[get_column_letter(cs)].width   = 8
        ws.column_dimensions[get_column_letter(cs+1)].width = 4

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(N_PREFIX)}2'
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18

outpath = os.path.join(os.path.dirname(__file__), 'wf_oos_results.xlsx')
wb.save(outpath)
print(f"\nSaved: {outpath}  ({os.path.getsize(outpath)//1024}KB)")
print("Done!")
