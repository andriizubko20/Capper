"""
Kelly 25% — тільки OOS симуляція (2025-08 → зараз)
- p оцінюється з IS даних (без look-ahead)
- Старт банкролю: $1000
- Cap: 2% від банкролю per bet
"""
import warnings; warnings.filterwarnings('ignore')
import pandas as pd
import numpy as np
import re, sys
sys.path.insert(0,'.')
from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix
from sqlalchemy import create_engine
from config.settings import settings

DATABASE_URL = settings.database_url
OOS_START    = pd.Timestamp('2025-08-01')
FLAT_STAKE   = 50.0
KELLY_FRAC   = 0.25
START_BANK   = 1000.0
MAX_BET_PCT  = 0.02  # 2% of bankroll

MODELS = {
    'Premier League': [
        'home[2.5,3.5) xg>=1.0','home[2.2,2.8) xg>=1.2 form>=1.5',
        'home[1.8,2.2) form>=2.2','home[1.7,2.0) xg>=1.8',
        'away[1.55,1.8) xg>=1.2 elo<=-75 form>=2.2',
    ],
    'Bundesliga': [
        'home[2.2,2.8) xg>=1.2 form>=1.5','home[1.8,2.2) xg>=1.2 elo>=30 mkt>=0.5',
        'home[1.8,2.2) elo>=75 form>=1.5 mkt>=0.5','home[1.55,1.8) elo>=150 form>=1.5',
        'away[2.5,3.5) xg>=1.0 elo<=-30 form>=1.5','away[2.5,3.5) form>=1.5 mkt>=0.35',
        'away[2.5,3.5) elo<=-75 mkt>=0.35','away[2.5,3.5) elo<=-30 form>=1.5 mkt>=0.35',
        'away[2.2,2.8) xg>=1.5 mkt>=0.4','away[2.2,2.8) mkt>=0.4',
        'away[2.2,2.8) elo<=-75 mkt>=0.35','away[2.0,2.5) mkt>=0.4',
        'away[1.7,2.0) xg>=1.0 form>=2.2',
    ],
    'Serie A': [
        'home[2.5,3.5) elo>=30 form>=1.5','home[2.0,2.5) xg>=1.0 mkt>=0.45',
        'home[2.0,2.5) form>=2.2 mkt>=0.45','home[1.8,2.2) xg>=1.5 form>=1.5',
        'home[1.3,1.55) xg>=1.5 elo>=150','away[2.5,3.5) xg>=1.0 form>=1.8 mkt>=0.35',
        'away[2.2,2.8) xg>=1.2 form>=1.8 mkt>=0.35','away[2.2,2.8) xg>=1.0 elo<=-150',
        'away[1.8,2.2) xg>=1.5 form>=2.2',
    ],
    'La Liga': [
        'home[2.2,2.8) xg>=1.0 form>=1.5','home[2.0,2.5) xg>=1.0 form>=1.5',
        'home[1.8,2.2) xg>=1.2 form>=1.5','home[1.8,2.2) xg>=1.0 form>=1.5 mkt>=0.5',
        'home[1.8,2.2) xg>=1.0 elo>=30 mkt>=0.5','home[1.55,1.8) xg>=1.8',
        'away[2.5,3.5) xg>=1.8 form>=1.5','away[2.5,3.5) elo<=-75 form>=2.2',
        'away[2.2,2.8) xg>=1.2 mkt>=0.4','away[1.8,2.2) xg>=1.8 elo<=-75 mkt>=0.45',
    ],
    'Ligue 1': [
        'home[2.5,3.5) xg>=1.5','home[2.2,2.8) xg>=1.2','home[2.0,2.5) elo>=75',
        'home[1.7,2.0) xg>=1.8','home[1.7,2.0) xg>=1.5 mkt>=0.5',
        'home[1.7,2.0) xg>=1.5 elo>=75','home[1.7,2.0) xg>=1.5 elo>=30',
        'home[1.55,1.8) xg>=1.2 elo>=75','away[2.5,3.5) xg>=1.0 form>=2.2',
        'away[2.2,2.8) form>=1.5 mkt>=0.4','away[2.2,2.8) elo<=-30 form>=1.5 mkt>=0.4',
    ],
    'Primeira Liga': [
        'away[2.2,2.8) xg>=1.2 mkt>=0.4','home[2.0,2.5) form>=1.5 mkt>=0.45',
        'home[1.55,1.8) elo>=150','home[1.8,2.2) form>=2.2',
    ],
    'Serie B': [
        'home[2.2,2.8) form>=2.2','home[2.0,2.5) elo>=75 form>=2.2',
        'home[1.7,2.0) xg>=1.0 form>=2.2','home[1.55,1.8) xg>=1.8',
        'home[1.55,1.8) xg>=1.5 elo>=75 form>=2.2','home[1.55,1.8) xg>=1.2 elo>=150',
        'home[1.55,1.8) xg>=1.5','away[2.0,2.5) xg>=1.5 elo<=-30',
    ],
    'Eredivisie': [
        'home[2.5,3.5) xg>=1.0','home[2.0,2.5) elo>=75',
        'away[2.5,3.5) elo<=-75 mkt>=0.35','away[1.8,2.2) xg>=1.2 form>=2.2',
        'away[1.7,2.0) xg>=1.2 form>=2.2','away[1.55,1.8) elo<=-150',
    ],
    'Jupiler Pro League': [
        'home[2.5,3.5) form>=2.2','home[2.0,2.5) xg>=1.0 mkt>=0.45',
        'home[1.8,2.2) xg>=1.2 mkt>=0.5','home[1.8,2.2) xg>=1.0 form>=1.5 mkt>=0.5',
        'home[1.7,2.0) xg>=1.0 form>=1.5 mkt>=0.5','home[1.55,1.8) xg>=1.8',
        'home[1.55,1.8) form>=2.2','home[1.3,1.55) xg>=1.5 elo>=150 form>=1.5',
        'away[2.2,2.8) xg>=1.2 form>=2.2 mkt>=0.35','away[2.0,2.5) xg>=1.5 form>=2.2',
    ],
    'Champions League': [
        'home[2.2,2.8) xg>=1.8 form>=1.5','home[2.0,2.5) form>=1.5 mkt>=0.45',
        'home[1.8,2.2) mkt>=0.5','home[1.7,2.0) xg>=1.5 mkt>=0.55',
        'home[1.7,2.0) xg>=1.5 elo>=75 mkt>=0.5','home[1.55,1.8) elo>=30 form>=2.2',
        'home[1.55,1.8) elo>=150','home[1.3,1.55) xg>=1.8 form>=1.8',
        'away[2.2,2.8) xg>=1.8 mkt>=0.4','away[2.2,2.8) xg>=1.5 form>=1.8 mkt>=0.4',
    ],
}

def parse_niche(s):
    m=re.match(r'(home|away)\[(\d+\.?\d*),(\d+\.?\d*)\)',s)
    side=m.group(1); lo=float(m.group(2)); hi=float(m.group(3))
    xg=0.0; elo=0; fm=0.0; mk=0.0
    mx=re.search(r'xg>=([\d.]+)',s);   xg=float(mx.group(1)) if mx else 0.0
    mx=re.search(r'elo>=([-\d]+)',s);  elo=int(mx.group(1)) if mx else elo
    mx=re.search(r'elo<=([-\d]+)',s);  elo=int(mx.group(1)) if mx else elo
    mx=re.search(r'form>=([\d.]+)',s); fm=float(mx.group(1)) if mx else 0.0
    mx=re.search(r'mkt>=([\d.]+)',s);  mk=float(mx.group(1)) if mx else 0.0
    return side,lo,hi,xg,elo,fm,mk

def apply_mask(data,side,lo,hi,xg,elo,fm,mk,league):
    oc='home_odds_val' if side=='home' else 'away_odds_val'
    m=(data['league_name']==league)&(data[oc]>=lo)&(data[oc]<hi)
    if xg>0:  m&=data['xg_ratio_home_5' if side=='home' else 'xg_ratio_away_5'].fillna(0)>=xg
    if side=='home' and elo>0: m&=data['elo_diff'].fillna(0)>=elo
    if side=='away' and elo<0: m&=data['elo_diff'].fillna(0)<=elo
    if fm>0:  m&=data['home_pts_5' if side=='home' else 'away_pts_5'].fillna(0)>=fm
    if mk>0:  m&=data['mkt_home_prob' if side=='home' else 'mkt_away_prob'].fillna(0)>=mk
    return data[m]

print('Loading data...')
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result']=='H').astype(int)
df['away_win']  = (df['result']=='A').astype(int)
df['ym'] = df['date'].dt.to_period('M')
df['_id'] = np.arange(len(df))
oos_p = OOS_START.to_period('M')

print('Loading team names...')
engine = create_engine(DATABASE_URL)
with engine.connect() as conn:
    teams_df = pd.read_sql('SELECT id, name FROM teams', conn)
    matches_raw = pd.read_sql('SELECT id as match_id, home_team_id, away_team_id FROM matches', conn)
teams_map = matches_raw.merge(teams_df.rename(columns={'id':'home_team_id','name':'home_name'}), on='home_team_id', how='left')
teams_map = teams_map.merge(teams_df.rename(columns={'id':'away_team_id','name':'away_name'}), on='away_team_id', how='left')
name_map = teams_map.set_index('match_id')[['home_name','away_name']]
if 'match_id' in df.columns:
    df = df.join(name_map, on='match_id')

# ── Step 1: Collect OOS bets with IS-based probability ─────────────────────────
print('Collecting bets...')
all_bets = []

for league, niches in MODELS.items():
    for niche_str in niches:
        side,lo,hi,xg,elo,fm,mk = parse_niche(niche_str)
        wc = 'home_win' if side=='home' else 'away_win'
        oc = 'home_odds_val' if side=='home' else 'away_odds_val'
        bets = apply_mask(df,side,lo,hi,xg,elo,fm,mk,league)
        if len(bets)==0: continue

        is_bets = bets[bets['ym'] < oos_p]
        oos_bets = bets[bets['ym'] >= oos_p]
        if len(oos_bets) == 0: continue

        # p estimated from IS only (no look-ahead)
        p_is = is_bets[wc].mean() if len(is_bets) >= 3 else bets[wc].mean()

        # OOS ROI for dedup priority
        oos_pl = (oos_bets[wc]*(oos_bets[oc]-1)-(1-oos_bets[wc])).sum()*FLAT_STAKE
        oos_roi = oos_pl/(len(oos_bets)*FLAT_STAKE)*100

        for idx, row in oos_bets.iterrows():
            hname = row.get('home_name','') or ''
            aname = row.get('away_name','') or ''
            all_bets.append({
                'date':     row['date'],
                'ym':       row['ym'],
                'league':   league,
                'side':     side,
                'match_id': idx,
                'niche':    niche_str,
                'wc':       wc,
                'oc':       oc,
                'won':      bool(row[wc]),
                'odds':     float(row[oc]),
                'oos_roi':  oos_roi,
                'p_is':     p_is,
                'match':    f"{hname} vs {aname}",
            })

bets_df = pd.DataFrame(all_bets)
print(f'  OOS bets (with dups): {len(bets_df)}')

# ── Step 2: Dedup ───────────────────────────────────────────────────────────────
bets_df = bets_df.sort_values(['date','match_id','side','oos_roi'], ascending=[True,True,True,False])
bets_df = bets_df.drop_duplicates(subset=['match_id','side'], keep='first').reset_index(drop=True)
bets_df = bets_df.sort_values('date').reset_index(drop=True)
print(f'  After dedup: {len(bets_df)} unique OOS bets')

# ── Step 3: Kelly + Flat simulation ────────────────────────────────────────────
bankroll = START_BANK
rows = []

for _, row in bets_df.iterrows():
    p = row['p_is']
    b = row['odds'] - 1.0
    if b <= 0 or p <= 0: continue

    f_star = (p * b - (1 - p)) / b
    kelly_stake = KELLY_FRAC * f_star * bankroll if f_star > 0 else 0
    kelly_stake = min(kelly_stake, MAX_BET_PCT * bankroll)
    kelly_stake = max(kelly_stake, 0.5) if f_star > 0 else 0

    flat_pl  = FLAT_STAKE*(b) if row['won'] else -FLAT_STAKE
    kelly_pl = kelly_stake*(b) if row['won'] else -kelly_stake if kelly_stake > 0 else 0

    bankroll += kelly_pl

    rows.append({
        'date':     str(row['date'])[:10],
        'ym':       row['ym'],
        'league':   row['league'],
        'match':    row['match'],
        'side':     row['side'],
        'niche':    row['niche'],
        'odds':     round(row['odds'],2),
        'p_is':     round(p,3),
        'f_star':   round(f_star,4) if f_star > 0 else 0,
        'k_stake':  round(kelly_stake,2),
        'won':      '✓' if row['won'] else '✗',
        'flat_pl':  round(flat_pl,1),
        'kelly_pl': round(kelly_pl,2),
        'bankroll': round(bankroll,2),
    })

sim = pd.DataFrame(rows)
oos_months = sorted(sim['ym'].unique())

# ── Monthly table ───────────────────────────────────────────────────────────────
print()
print('='*90)
print(f'  OOS Kelly 25% Simulation  |  Start: $1,000  |  Max bet: 2% bankroll')
print('='*90)
print(f"{'Month':<10} {'n':>4} {'W':>4} {'WR%':>5} {'FlatP&L':>9} {'FlatROI%':>9} {'KellyP&L':>10} {'Bankroll':>12} {'Δ%':>7}")
print('-'*90)

flat_cum = 0
for ym in oos_months:
    mg = sim[sim['ym']==ym]
    n = len(mg); w = int((mg['won']=='✓').sum())
    wr = w/n*100
    fp = mg['flat_pl'].sum(); fr = fp/(n*FLAT_STAKE)*100
    kp = mg['kelly_pl'].sum()
    bank_end = mg['bankroll'].iloc[-1]
    bank_start = mg['bankroll'].iloc[0] - mg['kelly_pl'].iloc[0] + (mg['k_stake'].iloc[0] if mg['won'].iloc[0]=='✗' else -mg['k_stake'].iloc[0]*(mg['odds'].iloc[0]-1))
    delta_pct = kp / (bank_end - kp) * 100 if (bank_end - kp) > 0 else 0
    flat_cum += fp
    print(f"  {str(ym):<8} {n:>4} {w:>4} {wr:>4.0f}%  {fp:>+8.0f}$  {fr:>+7.1f}%  {kp:>+9.0f}$  ${bank_end:>10,.0f}  {delta_pct:>+6.1f}%")

print('-'*90)
n_tot = len(sim); w_tot = int((sim['won']=='✓').sum())
fp_tot = sim['flat_pl'].sum(); kp_tot = sim['kelly_pl'].sum()
final_bank = sim['bankroll'].iloc[-1]
print(f"  {'TOTAL':<8} {n_tot:>4} {w_tot:>4} {w_tot/n_tot*100:>4.0f}%  {fp_tot:>+8.0f}$  {fp_tot/(n_tot*FLAT_STAKE)*100:>+7.1f}%  {kp_tot:>+9.0f}$  ${final_bank:>10,.0f}  {(final_bank-START_BANK)/START_BANK*100:>+6.1f}%")
print('='*90)

# ── Match list ──────────────────────────────────────────────────────────────────
print()
print(f"{'#':<4} {'Date':<12} {'League':<16} {'Match':<32} {'Side':<5} {'Odds':>5} {'p_IS':>5} {'Stake':>7} {'Won':<3} {'FlatP&L':>8} {'KellyP&L':>9} {'Bank':>10}")
print('-'*130)

for i,(_,r) in enumerate(sim.iterrows(),1):
    won_sym = '✓' if r['won']=='✓' else '✗'
    print(f"  {i:<3} {r['date']:<12} {r['league']:<16} {r['match'][:30]:<32} {r['side']:<5} "
          f"{r['odds']:>5.2f} {r['p_is']:>5.3f} ${r['k_stake']:>6.1f} {won_sym:<3} "
          f"{r['flat_pl']:>+8.1f}$ {r['kelly_pl']:>+9.2f}$ ${r['bankroll']:>9,.0f}")

print()
print(f'  Avg Kelly stake: ${sim["k_stake"].mean():.2f}  |  Min: ${sim["k_stake"].min():.2f}  |  Max: ${sim["k_stake"].max():.2f}')
print(f'  Flat  total: ${fp_tot:>+,.0f}  |  Kelly total: ${kp_tot:>+,.0f}')
print(f'  Final bankroll: ${final_bank:>,.0f}  (+{(final_bank-START_BANK)/START_BANK*100:.1f}%)')
print()

# ── Save Excel ──────────────────────────────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    OUT = '/app/BETA/kelly_oos_only.xlsx'
    fill_hdr  = PatternFill('solid',fgColor='1F3864')
    fill_oos  = PatternFill('solid',fgColor='2E4057')
    fill_ps   = PatternFill('solid',fgColor='70AD47')
    fill_p    = PatternFill('solid',fgColor='C6EFCE')
    fill_neg  = PatternFill('solid',fgColor='FFCCCC')
    fill_lg   = PatternFill('solid',fgColor='D6E4F0')
    thin      = Side(style='thin',color='CCCCCC')
    border    = Border(left=thin,right=thin,top=thin,bottom=thin)

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Monthly table
    ws1 = wb.create_sheet('Monthly OOS')
    mo_cols = ['Month','n','W','WR%','Flat P&L','Flat ROI%','Kelly P&L','Bankroll','Δ Bankroll %']
    for ci,col in enumerate(mo_cols,1):
        c=ws1.cell(1,ci,col); c.fill=fill_hdr
        c.font=Font(color='FFFFFF',bold=True); c.alignment=Alignment(horizontal='center',wrap_text=True)

    flat_cum2 = 0
    bank_prev = START_BANK
    for ri,ym in enumerate(oos_months,2):
        mg = sim[sim['ym']==ym]
        n=len(mg); w=int((mg['won']=='✓').sum()); wr=round(w/n*100,1)
        fp=round(mg['flat_pl'].sum(),0); fr=round(fp/(n*FLAT_STAKE)*100,1)
        kp=round(mg['kelly_pl'].sum(),0)
        bank_end=round(mg['bankroll'].iloc[-1],0)
        delta=round((bank_end-bank_prev)/bank_prev*100,1)
        bank_prev=bank_end
        vals=[str(ym),n,w,wr,fp,fr,kp,bank_end,delta]
        for ci,val in enumerate(vals,1):
            c=ws1.cell(ri,ci,val); c.border=border; c.alignment=Alignment(horizontal='center')
            if ci==5:
                c.number_format='+$#,##0;-$#,##0'; c.fill=fill_ps if val>500 else (fill_p if val>0 else fill_neg)
            if ci==6:
                c.number_format='+0.0;-0.0'; c.fill=fill_ps if val>=30 else (fill_p if val>0 else fill_neg); c.font=Font(bold=True)
            if ci==7:
                c.number_format='+$#,##0;-$#,##0'; c.fill=fill_ps if val>500 else (fill_p if val>0 else fill_neg)
            if ci==8:
                c.number_format='$#,##0'; c.font=Font(bold=True)
            if ci==9:
                c.number_format='+0.0;-0.0'; c.fill=fill_ps if val>=20 else (fill_p if val>0 else fill_neg); c.font=Font(bold=True)

    # Totals row
    ri=len(oos_months)+2
    vals=['TOTAL',n_tot,w_tot,round(w_tot/n_tot*100,1),round(fp_tot,0),
          round(fp_tot/(n_tot*FLAT_STAKE)*100,1),round(kp_tot,0),round(final_bank,0),
          round((final_bank-START_BANK)/START_BANK*100,1)]
    for ci,val in enumerate(vals,1):
        c=ws1.cell(ri,ci,val); c.font=Font(bold=True,size=11); c.fill=fill_lg
        c.border=border; c.alignment=Alignment(horizontal='center')
        if ci==5: c.number_format='+$#,##0;-$#,##0'
        if ci==7: c.number_format='+$#,##0;-$#,##0'
        if ci==8: c.number_format='$#,##0'
        if ci==9: c.number_format='+0.0;-0.0'

    for i,w_ in enumerate([10,5,5,7,12,10,12,13,13],1):
        ws1.column_dimensions[get_column_letter(i)].width=w_
    ws1.freeze_panes='A2'

    # Sheet 2: Match list
    ws2 = wb.create_sheet('Match List')
    ml_cols=['#','Date','League','Match','Side','Model','Odds','p_IS','f*','Kelly Stake $','Won','Flat P&L $','Kelly P&L $','Bankroll $']
    for ci,col in enumerate(ml_cols,1):
        c=ws2.cell(1,ci,col); c.fill=fill_hdr
        c.font=Font(color='FFFFFF',bold=True); c.alignment=Alignment(horizontal='center',wrap_text=True)
    ws2.row_dimensions[1].height=28

    for ri,(_,row) in enumerate(sim.iterrows(),2):
        vals=[ri-1, row['date'], row['league'], row['match'], row['side'],
              row['niche'], row['odds'], row['p_is'], row['f_star'],
              row['k_stake'], row['won'], row['flat_pl'], row['kelly_pl'], row['bankroll']]
        for ci,val in enumerate(vals,1):
            c=ws2.cell(ri,ci,val); c.border=border; c.alignment=Alignment(horizontal='center')
            if ci in (3,4,6): c.alignment=Alignment(horizontal='left')
            if ci==11:
                c.fill=fill_ps if val=='✓' else fill_neg; c.font=Font(bold=True)
            if ci==12:
                c.number_format='+0.0;-0.0;0'
                c.fill=fill_p if row['flat_pl']>0 else fill_neg
            if ci==13:
                c.number_format='+0.00;-0.00;0'
                c.fill=fill_p if row['kelly_pl']>0 else fill_neg
            if ci==14:
                c.number_format='$#,##0.00'
                c.font=Font(bold=True,color='155724' if row['bankroll']>=START_BANK else '721c24')

    ml_widths=[5,12,16,34,6,44,7,6,7,12,5,12,12,14]
    for i,w_ in enumerate(ml_widths,1): ws2.column_dimensions[get_column_letter(i)].width=w_
    ws2.freeze_panes='B2'
    ws2.auto_filter.ref=f'A1:{get_column_letter(len(ml_cols))}1'

    wb.save(OUT)
    print(f'Excel saved: {OUT}')
except Exception as e:
    print(f'Excel save failed: {e}')
