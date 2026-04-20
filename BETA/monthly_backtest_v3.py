import warnings; warnings.filterwarnings('ignore')
import pandas as pd
import numpy as np
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
sys.path.insert(0, '.')
from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix

OUT        = '/app/BETA/monthly_backtest_v3.xlsx'
FLAT_STAKE = 50.0
OOS_START  = pd.Timestamp('2025-08-01')
OVERLAP_THRESHOLD = 0.85   # якщо B ⊆ A на 85%+ → дублікат

LEAGUES = [
    'Premier League','Bundesliga','Serie A','La Liga','Ligue 1',
    'Primeira Liga','Serie B','Eredivisie','Jupiler Pro League','Champions League'
]
ODDS_RANGES = [(1.30,1.55),(1.55,1.80),(1.70,2.00),(1.80,2.20),(2.00,2.50),(2.20,2.80),(2.50,3.50)]
XG = [0.0,1.0,1.2,1.5,1.8]
EH = [0,30,75,150]; EA = [0,-30,-75,-150]
FT = [0.0,1.5,1.8,2.2]; MH = [0.0,0.45,0.50,0.55]; MA = [0.0,0.35,0.40,0.45]

print('Loading data...')
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result']=='H').astype(int)
df['away_win']  = (df['result']=='A').astype(int)
df['ym'] = df['date'].dt.to_period('M')
df['_id'] = np.arange(len(df))   # unique match index for set ops

all_months = pd.period_range(df['date'].min().to_period('M'),
                              df['date'].max().to_period('M'), freq='M')
oos_p = OOS_START.to_period('M')
print(f'Months: {len(all_months)}  |  OOS from: {oos_p}')

def niche_label(side,lo,hi,xg,elo,fm,mk):
    s=f'{side}[{lo},{hi})'
    if xg>0:   s+=f' xg>={xg}'
    if elo!=0: s+=f' elo>={elo}' if elo>0 else f' elo<={elo}'
    if fm>0:   s+=f' form>={fm}'
    if mk>0:   s+=f' mkt>={mk}'
    return s

def apply_mask(data,side,lo,hi,xg,elo,fm,mk,league):
    oc='home_odds_val' if side=='home' else 'away_odds_val'
    m=(data['league_name']==league)&(data[oc]>=lo)&(data[oc]<hi)
    if xg>0:  m&=data['xg_ratio_home_5' if side=='home' else 'xg_ratio_away_5'].fillna(0)>=xg
    if side=='home' and elo>0: m&=data['elo_diff'].fillna(0)>=elo
    if side=='away' and elo<0: m&=data['elo_diff'].fillna(0)<=elo
    if fm>0:  m&=data['home_pts_5' if side=='home' else 'away_pts_5'].fillna(0)>=fm
    if mk>0:  m&=data['mkt_home_prob' if side=='home' else 'mkt_away_prob'].fillna(0)>=mk
    return data[m]

print('Running combinations...')
results = []
for side in ('home','away'):
    et=EH if side=='home' else EA; mt=MH if side=='home' else MA
    wc='home_win' if side=='home' else 'away_win'
    oc='home_odds_val' if side=='home' else 'away_odds_val'
    for lo,hi in ODDS_RANGES:
        for xg,elo,fm,mk in product(XG,et,FT,mt):
            for lg in LEAGUES:
                bets = apply_mask(df,side,lo,hi,xg,elo,fm,mk,lg)
                if len(bets)<5: continue
                monthly={}
                for ym,grp in bets.groupby('ym'):
                    pl=(grp[wc]*(grp[oc]-1)-(1-grp[wc])).sum()*FLAT_STAKE
                    n=len(grp)
                    roi=pl/(n*FLAT_STAKE)*100 if n>0 else 0
                    monthly[ym]={'n':n,'pl':round(pl,1),'roi':round(roi,1)}
                total_pl=sum(v['pl'] for v in monthly.values())
                total_n =sum(v['n']  for v in monthly.values())
                if total_pl<=0: continue
                ins_pl=sum(v['pl'] for m,v in monthly.items() if m<oos_p)
                oos_pl=sum(v['pl'] for m,v in monthly.items() if m>=oos_p)
                ins_n =sum(v['n']  for m,v in monthly.items() if m<oos_p)
                oos_n =sum(v['n']  for m,v in monthly.items() if m>=oos_p)
                results.append({
                    'league':lg,'side':side,'lo':lo,'hi':hi,
                    'niche':niche_label(side,lo,hi,xg,elo,fm,mk),
                    'n_filters':niche_label(side,lo,hi,xg,elo,fm,mk).count('>=') +
                                niche_label(side,lo,hi,xg,elo,fm,mk).count('<='),
                    'total_n':total_n,'total_pl':total_pl,
                    'roi_pct':total_pl/(total_n*FLAT_STAKE)*100,
                    'ins_roi':ins_pl/(ins_n*FLAT_STAKE)*100 if ins_n>0 else 0,
                    'oos_roi':oos_pl/(oos_n*FLAT_STAKE)*100 if oos_n>0 else 0,
                    'oos_n':oos_n,'ins_n':ins_n,
                    'monthly':monthly,
                    'match_ids': frozenset(bets['_id'].tolist()),
                })

print(f'Profitable before dedup: {len(results)}')

# ── Dedup by match-set overlap ────────────────────────────────────────────────
# Group by (league, side, lo, hi) — only compare within same odds bucket
# Within each group: sort by OOS ROI desc, then greedily keep niches
# whose match set is NOT 85%+ covered by any already-kept niche

from collections import defaultdict
groups = defaultdict(list)
for r in results:
    groups[(r['league'], r['side'], r['lo'], r['hi'])].append(r)

kept = []
for key, group in groups.items():
    # Sort: best OOS ROI first; ties broken by more matches (broader)
    group.sort(key=lambda x: (-x['oos_roi'], -x['total_n']))
    kept_in_group = []
    for r in group:
        is_dup = False
        for k in kept_in_group:
            # Check if r's matches are 85%+ contained in k's matches
            intersection = len(r['match_ids'] & k['match_ids'])
            coverage = intersection / len(r['match_ids']) if r['match_ids'] else 0
            if coverage >= OVERLAP_THRESHOLD:
                is_dup = True
                break
        if not is_dup:
            kept_in_group.append(r)
    kept.extend(kept_in_group)

# Remove match_ids before export (not needed)
for r in kept:
    del r['match_ids']

print(f'After match-set dedup: {len(kept)}')
kept.sort(key=lambda x:(x['league'],-x['oos_roi']))

# ── Excel ─────────────────────────────────────────────────────────────────────
fill_hdr   = PatternFill('solid',fgColor='1F3864')
fill_oos_h = PatternFill('solid',fgColor='2E4057')
fill_lg    = PatternFill('solid',fgColor='D6E4F0')
fill_ps    = PatternFill('solid',fgColor='70AD47')
fill_p     = PatternFill('solid',fgColor='C6EFCE')
fill_neg   = PatternFill('solid',fgColor='FFCCCC')
fill_zero  = PatternFill('solid',fgColor='F5F5F5')
thin       = Side(style='thin',color='CCCCCC')
border     = Border(left=thin,right=thin,top=thin,bottom=thin)

META = ['Niche','Side','Total n','Total P&L','ROI%','In-Sample ROI%','OOS ROI%','OOS n']
N_META = len(META)

def write_league_sheet(wb, league_name, rows):
    ws = wb.create_sheet(league_name[:20])
    for ci,col in enumerate(META,1):
        c=ws.cell(1,ci,col)
        c.fill=fill_hdr; c.font=Font(color='FFFFFF',bold=True,size=10)
        c.alignment=Alignment(horizontal='center',wrap_text=True)
    for mi,m in enumerate(all_months):
        base=N_META+1+mi*3
        lbl=m.strftime('%b-%y')
        hf=fill_oos_h if m>=oos_p else fill_hdr
        for off,sub in enumerate(['n','P&L','ROI%']):
            c=ws.cell(1,base+off,f'{lbl}\n{sub}')
            c.fill=hf; c.font=Font(color='FFFFFF',bold=True,size=8)
            c.alignment=Alignment(horizontal='center',wrap_text=True)
    ws.row_dimensions[1].height=28

    row_idx=2
    for r in rows:
        roi=r['roi_pct']; oos=r['oos_roi']; ins=r['ins_roi']
        mv=[r['niche'],r['side'],r['total_n'],round(r['total_pl'],0),
            round(roi,1),round(ins,1),round(oos,1),int(r['oos_n'])]
        for ci,val in enumerate(mv,1):
            c=ws.cell(row_idx,ci,val); c.border=border
            c.alignment=Alignment(horizontal='center')
            if ci==1: c.alignment=Alignment(horizontal='left')
            if ci==5: c.number_format='+0.0;-0.0;0'
            if ci==6:
                c.number_format='+0.0;-0.0;0'
                c.fill=fill_ps if ins>=15 else (fill_p if ins>0 else fill_neg)
            if ci==7:
                c.number_format='+0.0;-0.0;0'
                c.fill=fill_ps if oos>=15 else (fill_p if oos>0 else fill_neg)
                c.font=Font(bold=True)
        for mi,m in enumerate(all_months):
            base=N_META+1+mi*3
            md=r['monthly'].get(m)
            if md is None:
                for off in range(3):
                    c=ws.cell(row_idx,base+off,''); c.fill=fill_zero; c.border=border
            else:
                c=ws.cell(row_idx,base,md['n']); c.border=border
                c.alignment=Alignment(horizontal='center'); c.font=Font(size=8)
                c=ws.cell(row_idx,base+1,md['pl']); c.border=border
                c.number_format='+0;-0;0'; c.font=Font(size=8)
                c.alignment=Alignment(horizontal='center')
                c.fill=fill_ps if md['pl']>=100 else (fill_p if md['pl']>0 else (fill_neg if md['pl']<0 else fill_zero))
                c=ws.cell(row_idx,base+2,md['roi']); c.border=border
                c.number_format='+0;-0;0'; c.font=Font(size=8)
                c.alignment=Alignment(horizontal='center')
                c.fill=fill_ps if md['roi']>=20 else (fill_p if md['roi']>0 else (fill_neg if md['roi']<0 else fill_zero))
        row_idx+=1

    ws.column_dimensions['A'].width=42
    ws.column_dimensions['B'].width=6
    for ci in range(3,N_META+1):
        ws.column_dimensions[get_column_letter(ci)].width=10
    for mi in range(len(all_months)):
        for off in range(3):
            ws.column_dimensions[get_column_letter(N_META+1+mi*3+off)].width=4.5
    ws.freeze_panes=f'{get_column_letter(N_META+1)}2'
    ws.auto_filter.ref=f'A1:{get_column_letter(N_META)}1'

print('Writing Excel...')
wb = Workbook()
wb.remove(wb.active)

by_league = {}
for r in kept:
    by_league.setdefault(r['league'],[]).append(r)

for lg in LEAGUES:
    rows = by_league.get(lg,[])
    if not rows: continue
    write_league_sheet(wb, lg, rows)
    print(f'  {lg}: {len(rows)} niches')

# Summary — топ 150 по OOS ROI де OOS n >= 5
ws_sum = wb.create_sheet('SUMMARY', 0)
sum_cols=['League','Niche','Side','Total n','Total P&L','ROI%','In-Sample ROI%','OOS ROI%','OOS n']
for ci,col in enumerate(sum_cols,1):
    c=ws_sum.cell(1,ci,col)
    c.fill=fill_hdr; c.font=Font(color='FFFFFF',bold=True,size=10)
    c.alignment=Alignment(horizontal='center',wrap_text=True)

top=[r for r in kept if r['oos_n']>=5]
top.sort(key=lambda x:-x['oos_roi'])
top=top[:150]

prev_lg=None; row_idx=2
for r in top:
    if r['league']!=prev_lg:
        c=ws_sum.cell(row_idx,1,r['league'])
        c.fill=fill_lg; c.font=Font(bold=True,size=11)
        ws_sum.merge_cells(start_row=row_idx,start_column=1,end_row=row_idx,end_column=len(sum_cols))
        row_idx+=1; prev_lg=r['league']
    oos=r['oos_roi']; ins=r['ins_roi']
    vals=[r['league'],r['niche'],r['side'],r['total_n'],round(r['total_pl'],0),
          round(r['roi_pct'],1),round(ins,1),round(oos,1),int(r['oos_n'])]
    for ci,val in enumerate(vals,1):
        c=ws_sum.cell(row_idx,ci,val); c.border=border
        c.alignment=Alignment(horizontal='center')
        if ci==2: c.alignment=Alignment(horizontal='left')
        if ci==6: c.number_format='+0.0;-0.0;0'
        if ci==7:
            c.number_format='+0.0;-0.0;0'
            c.fill=fill_ps if ins>=15 else (fill_p if ins>0 else fill_neg)
        if ci==8:
            c.number_format='+0.0;-0.0;0'
            c.fill=fill_ps if oos>=20 else (fill_p if oos>0 else fill_neg)
            c.font=Font(bold=True)
    row_idx+=1

ws_sum.column_dimensions['A'].width=16
ws_sum.column_dimensions['B'].width=44
for ci in range(3,len(sum_cols)+1):
    ws_sum.column_dimensions[get_column_letter(ci)].width=12
ws_sum.freeze_panes='A2'
ws_sum.auto_filter.ref=f'A1:{get_column_letter(len(sum_cols))}1'

wb.save(OUT)
print(f'\nSaved: {OUT}')
print('✅ Done!')
