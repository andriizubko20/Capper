"""
Builds BETA/walkforward_results.xlsx
One sheet per league, ALL filter combinations (stable + unstable),
all 11 quarterly windows, ROI% + n per cell.
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import numpy as np
import pandas as pd
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix

# ── Load data ─────────────────────────────────────────────────────────────────
print("Loading data...")
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result'] == 'H').astype(int)
df['away_win']  = (df['result'] == 'A').astype(int)
print(f"Total: {len(df)} | {df['date'].min().date()} – {df['date'].max().date()}")

LEAGUES = [
    'Premier League','Bundesliga','Serie A','La Liga','Ligue 1',
    'Primeira Liga','Serie B','Eredivisie','Jupiler Pro League','Champions League'
]
LEAGUE_SHORT = {
    'Premier League':'EPL','Bundesliga':'Bundesliga','Serie A':'Serie A',
    'La Liga':'La Liga','Ligue 1':'Ligue 1','Primeira Liga':'Portugal',
    'Serie B':'Serie B','Eredivisie':'Eredivisie',
    'Jupiler Pro League':'Jupiler','Champions League':'UCL'
}

# ── Windows ───────────────────────────────────────────────────────────────────
WINDOWS = []
cur = pd.Timestamp('2023-08-01')
end = pd.Timestamp('2026-04-30')
while cur < end:
    w_end = cur + pd.DateOffset(months=3) - pd.Timedelta(days=1)
    if w_end > end: w_end = end
    WINDOWS.append((cur.strftime('%Y-%m'), cur, w_end))
    cur += pd.DateOffset(months=3)
WIN_LABELS = [w[0] for w in WINDOWS]
print(f"Windows: {len(WINDOWS)} ({WIN_LABELS[0]} → {WIN_LABELS[-1]})")

# ── Filter space ──────────────────────────────────────────────────────────────
ODDS_RANGES = [
    (1.30,1.55),(1.55,1.80),(1.70,2.00),
    (1.80,2.20),(2.00,2.50),(2.20,2.80),(2.50,3.50),
]
XG_THRS    = [0.0,1.0,1.2,1.5,1.8]
ELO_H_THRS = [0,30,75,150]
ELO_A_THRS = [0,-30,-75,-150]
FORM_THRS  = [0.0,1.5,1.8,2.2]
MKT_H_THRS = [0.0,0.45,0.50,0.55]
MKT_A_THRS = [0.0,0.35,0.40,0.45]
MIN_N_WIN  = 3

# Pre-slice by window
win_slices = {wlbl: df[(df['date']>=ws)&(df['date']<=we)]
              for wlbl,ws,we in WINDOWS}

# ── Sweep ─────────────────────────────────────────────────────────────────────
print("Running sweep...")
# results[lg][label] = {meta..., wlbl: {'roi':..,'n':..,'wr':..,'ev':..}}
results = {lg: {} for lg in LEAGUES}

for side in ('home','away'):
    oc  = 'home_odds_val'   if side=='home' else 'away_odds_val'
    xgc = 'xg_ratio_home_5' if side=='home' else 'xg_ratio_away_5'
    fc  = 'home_pts_5'      if side=='home' else 'away_pts_5'
    mc  = 'mkt_home_prob'   if side=='home' else 'mkt_away_prob'
    wc  = 'home_win'        if side=='home' else 'away_win'
    mt  = MKT_H_THRS if side=='home' else MKT_A_THRS
    et  = ELO_H_THRS if side=='home' else ELO_A_THRS

    for lo,hi in ODDS_RANGES:
        for xg_t,elo_t,form_t,mkt_t in product(XG_THRS,et,FORM_THRS,mt):
            parts = [f'{side}[{lo},{hi})']
            if xg_t  > 0: parts.append(f'xg>={xg_t}')
            if side=='home' and elo_t>0: parts.append(f'elo>={elo_t}')
            if side=='away' and elo_t<0: parts.append(f'elo<={elo_t}')
            if form_t > 0: parts.append(f'form>={form_t}')
            if mkt_t  > 0: parts.append(f'mkt>={mkt_t}')
            label = ' '.join(parts)

            for wlbl,_,_ in WINDOWS:
                wdf  = win_slices[wlbl]
                mask = (wdf[oc]>=lo)&(wdf[oc]<hi)
                if xg_t  > 0: mask &= (wdf[xgc].fillna(0)>=xg_t)
                if side=='home' and elo_t>0: mask &= (wdf['elo_diff'].fillna(0)>=elo_t)
                if side=='away' and elo_t<0: mask &= (wdf['elo_diff'].fillna(0)<=elo_t)
                if form_t > 0: mask &= (wdf[fc].fillna(0)>=form_t)
                if mkt_t  > 0: mask &= (wdf[mc].fillna(0)>=mkt_t)
                sub = wdf[mask]

                for lg in LEAGUES:
                    sub_l = sub[sub['league_name']==lg]
                    n = len(sub_l)
                    if n < MIN_N_WIN: continue
                    wr    = sub_l[wc].mean()
                    avg_o = sub_l[oc].mean()
                    roi   = (sub_l[wc]*(sub_l[oc]-1)-(1-sub_l[wc])).mean()*100
                    ev    = (wr*avg_o-1)*100
                    if label not in results[lg]:
                        results[lg][label] = {'side':side,'lo':lo,'hi':hi,
                                               'xg_t':xg_t,'elo_t':elo_t,
                                               'form_t':form_t,'mkt_t':mkt_t}
                    results[lg][label][wlbl] = {
                        'roi':round(roi,1),'n':n,
                        'wr':round(wr*100,1),'ev':round(ev,1)
                    }

print("Sweep done. Building Excel...")

# ── Excel helpers ─────────────────────────────────────────────────────────────
GREEN_DARK  = PatternFill('solid', start_color='1E8449')   # dark green header
GREEN_LIGHT = PatternFill('solid', start_color='D5F5E3')   # pos ROI cell
RED_LIGHT   = PatternFill('solid', start_color='FADBD8')   # neg ROI cell
GREY_FILL   = PatternFill('solid', start_color='F2F3F4')   # missing
HEADER_FILL = PatternFill('solid', start_color='2C3E50')   # header bg
SUBHDR_FILL = PatternFill('solid', start_color='5D6D7E')   # sub-header
WHITE_FILL  = PatternFill('solid', start_color='FFFFFF')

HDR_FONT    = Font(name='Arial', bold=True, color='FFFFFF', size=10)
SUBHDR_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=9)
BODY_FONT   = Font(name='Arial', size=9)
BOLD_FONT   = Font(name='Arial', bold=True, size=9)
CENTER      = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left',   vertical='center')

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def set_header(cell, val):
    cell.value = val
    cell.font  = HDR_FONT
    cell.fill  = HEADER_FILL
    cell.alignment = CENTER
    cell.border = thin_border()

def set_subhdr(cell, val):
    cell.value = val
    cell.font  = SUBHDR_FONT
    cell.fill  = SUBHDR_FILL
    cell.alignment = CENTER
    cell.border = thin_border()

def set_body(cell, val, align=CENTER):
    cell.value = val
    cell.font  = BODY_FONT
    cell.alignment = align
    cell.border = thin_border()

def agg_stats(data):
    wins = [(wl, data[wl]) for wl in WIN_LABELS if wl in data]
    if not wins: return None
    n_tot   = sum(w['n']   for _,w in wins)
    avg_roi = round(np.mean([w['roi'] for _,w in wins]),1)
    avg_wr  = round(np.mean([w['wr']  for _,w in wins]),1)
    avg_ev  = round(np.mean([w['ev']  for _,w in wins]),1)
    n_pos   = sum(1 for _,w in wins if w['roi']>0)
    win_pct = round(n_pos/len(wins)*100,0)
    return {'n_tot':n_tot,'avg_roi':avg_roi,'avg_wr':avg_wr,
            'avg_ev':avg_ev,'win_pct':win_pct,
            'n_wins':n_pos,'n_windows':len(wins)}

# ── Build workbook ────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws_sum = wb.create_sheet('SUMMARY')

sum_headers = ['Ліга','Side','Найкращий фільтр','n_tot','Avg_ROI%',
               'Avg_WR%','Win%','Windows(+/tot)','Stable?']
for ci, h in enumerate(sum_headers, 1):
    set_header(ws_sum.cell(1, ci), h)

sum_row = 2
for lg in LEAGUES:
    lname = LEAGUE_SHORT[lg]
    for side in ('home','away'):
        best = None
        for label, data in results[lg].items():
            if data['side'] != side: continue
            a = agg_stats(data)
            if not a or a['n_tot']<10 or a['n_windows']<4: continue
            if best is None or a['avg_roi'] > best['avg_roi']:
                best = {**a, 'label':label}
        if best:
            stable = '✅ ТАК' if best['avg_roi']>0 and best['win_pct']>=60 else (
                     '⚠️ MIXED' if best['avg_roi']>0 else '❌ НІ')
            row = [lname, side, best['label'], best['n_tot'],
                   best['avg_roi']/100, best['avg_wr']/100,
                   best['win_pct']/100,
                   f"{best['n_wins']}/{best['n_windows']}",
                   stable]
            for ci, val in enumerate(row, 1):
                c = ws_sum.cell(sum_row, ci)
                set_body(c, val, LEFT if ci in (1,2,3,9) else CENTER)
                if ci == 5:  # Avg_ROI
                    c.number_format = '+0.0%;-0.0%;0.0%'
                    c.fill = GREEN_LIGHT if (best['avg_roi'] or 0)>0 else RED_LIGHT
                elif ci == 6:
                    c.number_format = '0.0%'
                elif ci == 7:
                    c.number_format = '0%'
            sum_row += 1

ws_sum.column_dimensions['A'].width = 12
ws_sum.column_dimensions['B'].width = 6
ws_sum.column_dimensions['C'].width = 55
for ci in range(4, 10):
    ws_sum.column_dimensions[get_column_letter(ci)].width = 14
ws_sum.freeze_panes = 'A2'
ws_sum.auto_filter.ref = f'A1:{get_column_letter(len(sum_headers))}1'

# ── Per-league sheets ─────────────────────────────────────────────────────────
for lg in LEAGUES:
    lname = LEAGUE_SHORT[lg]
    print(f"  Building sheet: {lname}...")
    ws = wb.create_sheet(lname)

    # Static columns: Filter | Side | n_tot | Avg_ROI% | Avg_WR% | Avg_EV% | Win% | W+/Wtot | Stable
    STATIC_COLS = ['Filter','Side','n_tot','Avg_ROI%','Avg_WR%','Avg_EV%','Win%','W+/Wtot','Stable']
    n_static = len(STATIC_COLS)

    # Dynamic cols: per window → ROI% | n
    # Row 1: group header (window label spanning 2 cols)
    # Row 2: sub-header ROI% | n

    total_cols = n_static + len(WINDOWS)*2

    # Row 1 headers — static
    for ci, h in enumerate(STATIC_COLS, 1):
        set_header(ws.cell(1, ci), h)
    # Row 1 — window group headers (merged pairs)
    for wi, (wlbl,_,_) in enumerate(WINDOWS):
        col_start = n_static + wi*2 + 1
        ws.merge_cells(start_row=1, start_column=col_start,
                       end_row=1,   end_column=col_start+1)
        set_header(ws.cell(1, col_start), wlbl)

    # Row 2 sub-headers
    for ci in range(1, n_static+1):
        set_subhdr(ws.cell(2, ci), '')
    for wi in range(len(WINDOWS)):
        col_start = n_static + wi*2 + 1
        set_subhdr(ws.cell(2, col_start),   'ROI%')
        set_subhdr(ws.cell(2, col_start+1), 'n')

    # Collect all rows
    all_rows = []
    for label, data in results[lg].items():
        a = agg_stats(data)
        if not a: continue
        stable = ('✅' if a['avg_roi']>0 and a['win_pct']>=60 else
                  '⚠️'  if a['avg_roi']>0 else '❌')
        all_rows.append({'label':label,'data':data,**a,'stable':stable})

    # Sort: stable first (avg_roi desc), then mixed, then losing
    def sort_key(r):
        if r['avg_roi']>0 and r['win_pct']>=60: grp=0
        elif r['avg_roi']>0:                    grp=1
        else:                                    grp=2
        return (grp, -r['avg_roi'])
    all_rows.sort(key=sort_key)

    # Write rows
    excel_row = 3
    prev_grp = None
    for r in all_rows:
        grp = (0 if r['avg_roi']>0 and r['win_pct']>=60 else
               1 if r['avg_roi']>0 else 2)
        # Group separator row
        if grp != prev_grp:
            titles = {0:'▶ СТАБІЛЬНО ПРИБУТКОВІ (avg_roi>0, win%≥60%)',
                      1:'▶ ПОМІРНО ПРИБУТКОВІ (avg_roi>0, win%<60%)',
                      2:'▶ ЗБИТКОВІ (avg_roi≤0)'}
            grp_colors = {0:'1A5276', 1:'7D6608', 2:'78281F'}
            ws.merge_cells(start_row=excel_row, start_column=1,
                           end_row=excel_row,   end_column=total_cols)
            c = ws.cell(excel_row, 1)
            c.value = titles[grp]
            c.font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            c.fill  = PatternFill('solid', start_color=grp_colors[grp])
            c.alignment = LEFT
            excel_row += 1
            prev_grp = grp

        # Static cells
        static_vals = [
            r['label'], r['data']['side'], r['n_tot'],
            r['avg_roi']/100, r['avg_wr']/100, r['avg_ev']/100,
            r['win_pct']/100,
            f"{r['n_wins']}/{r['n_windows']}",
            r['stable']
        ]
        row_fill = (GREEN_LIGHT if grp==0 else
                    PatternFill('solid', start_color='FEF9E7') if grp==1 else
                    PatternFill('solid', start_color='FDEDEC'))

        for ci, val in enumerate(static_vals, 1):
            c = ws.cell(excel_row, ci)
            set_body(c, val, LEFT if ci in (1,2,9) else CENTER)
            c.fill = row_fill
            if ci == 4:   c.number_format = '+0.0%;-0.0%;0.0%'
            elif ci == 5: c.number_format = '0.0%'
            elif ci == 6: c.number_format = '+0.0%;-0.0%;0.0%'
            elif ci == 7: c.number_format = '0%'
            # Colour avg_roi cell
            if ci == 4:
                c.fill = (GREEN_LIGHT if r['avg_roi']>0 else
                          RED_LIGHT   if r['avg_roi']<0 else WHITE_FILL)
                c.font = Font(name='Arial', size=9,
                              bold=(abs(r['avg_roi'])>15))

        # Window cells
        for wi, (wlbl,_,_) in enumerate(WINDOWS):
            col_roi = n_static + wi*2 + 1
            col_n   = n_static + wi*2 + 2
            if wlbl in r['data']:
                d = r['data'][wlbl]
                # ROI cell
                c_roi = ws.cell(excel_row, col_roi)
                c_roi.value  = d['roi']/100
                c_roi.font   = Font(name='Arial', size=9,
                                    bold=(abs(d['roi'])>20))
                c_roi.number_format = '+0.0%;-0.0%;0.0%'
                c_roi.alignment = CENTER
                c_roi.border = thin_border()
                c_roi.fill = (GREEN_LIGHT if d['roi']>0 else
                              RED_LIGHT   if d['roi']<0 else WHITE_FILL)
                # n cell
                c_n = ws.cell(excel_row, col_n)
                set_body(c_n, d['n'])
                c_n.fill = row_fill
            else:
                c_roi = ws.cell(excel_row, col_roi)
                set_body(c_roi, '—')
                c_roi.fill = GREY_FILL
                c_n = ws.cell(excel_row, col_n)
                set_body(c_n, '—')
                c_n.fill = GREY_FILL

        excel_row += 1

    # Column widths
    ws.column_dimensions['A'].width = 55   # Filter
    ws.column_dimensions['B'].width = 6    # Side
    ws.column_dimensions['C'].width = 7    # n_tot
    ws.column_dimensions['D'].width = 9    # Avg_ROI
    ws.column_dimensions['E'].width = 8    # Avg_WR
    ws.column_dimensions['F'].width = 8    # Avg_EV
    ws.column_dimensions['G'].width = 7    # Win%
    ws.column_dimensions['H'].width = 9    # W+/Wtot
    ws.column_dimensions['I'].width = 7    # Stable
    for wi in range(len(WINDOWS)):
        ws.column_dimensions[get_column_letter(n_static+wi*2+1)].width = 8
        ws.column_dimensions[get_column_letter(n_static+wi*2+2)].width = 4

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(n_static)}2'
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18

outpath = os.path.join(os.path.dirname(__file__), 'walkforward_results.xlsx')
wb.save(outpath)
print(f"\nSaved: {outpath}  ({os.path.getsize(outpath)//1024}KB)")
print("Done!")
