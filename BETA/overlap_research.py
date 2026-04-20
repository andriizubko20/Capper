import warnings; warnings.filterwarnings('ignore')
import pandas as pd
import numpy as np
import re, sys
sys.path.insert(0,'.')
from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = '/app/BETA/overlap_research.xlsx'
FLAT_STAKE = 50.0
OOS_START  = pd.Timestamp('2025-08-01')

MODELS = {
    'Premier League': ['home[2.5,3.5) xg>=1.0','home[2.2,2.8) xg>=1.2 form>=1.5','home[1.8,2.2) form>=2.2','home[1.7,2.0) xg>=1.8','away[1.55,1.8) xg>=1.2 elo<=-75 form>=2.2'],
    'Bundesliga': ['home[2.2,2.8) xg>=1.2 form>=1.5','home[1.8,2.2) xg>=1.2 elo>=30 mkt>=0.5','home[1.8,2.2) elo>=75 form>=1.5 mkt>=0.5','home[1.55,1.8) elo>=150 form>=1.5','away[2.5,3.5) xg>=1.0 elo<=-30 form>=1.5','away[2.5,3.5) form>=1.5 mkt>=0.35','away[2.5,3.5) elo<=-75 mkt>=0.35','away[2.5,3.5) elo<=-30 form>=1.5 mkt>=0.35','away[2.5,3.5) elo<=-150','away[2.2,2.8) xg>=1.5 mkt>=0.4','away[2.2,2.8) mkt>=0.4','away[2.2,2.8) elo<=-75 mkt>=0.35','away[2.0,2.5) mkt>=0.4','away[1.7,2.0) xg>=1.0 form>=2.2'],
    'Serie A': ['home[2.5,3.5) elo>=30 form>=1.5','home[2.0,2.5) xg>=1.0 mkt>=0.45','home[2.0,2.5) form>=2.2 mkt>=0.45','home[1.8,2.2) xg>=1.5 form>=1.5','home[1.3,1.55) xg>=1.5 elo>=150','away[2.5,3.5) xg>=1.0 form>=1.8 mkt>=0.35','away[2.2,2.8) xg>=1.2 form>=1.8 mkt>=0.35','away[2.2,2.8) xg>=1.0 elo<=-150','away[1.8,2.2) xg>=1.5 form>=2.2','away[1.7,2.0) xg>=1.5 form>=2.2'],
    'La Liga': ['home[2.2,2.8) xg>=1.0 form>=1.5','home[2.0,2.5) xg>=1.0 form>=1.5','home[1.8,2.2) xg>=1.2 form>=1.5','home[1.8,2.2) xg>=1.0 mkt>=0.5','home[1.8,2.2) xg>=1.0 form>=1.5 mkt>=0.5','home[1.8,2.2) xg>=1.0 elo>=30 mkt>=0.5','home[1.55,1.8) xg>=1.8','away[2.5,3.5) xg>=1.8 form>=1.5','away[2.5,3.5) elo<=-75 form>=2.2','away[2.2,2.8) xg>=1.2 mkt>=0.4','away[1.8,2.2) xg>=1.8 elo<=-75 mkt>=0.45'],
    'Ligue 1': ['home[2.5,3.5) xg>=1.5','home[2.2,2.8) xg>=1.2','home[2.0,2.5) elo>=75','home[1.7,2.0) xg>=1.8','home[1.7,2.0) xg>=1.5 mkt>=0.5','home[1.7,2.0) xg>=1.5 elo>=75','home[1.7,2.0) xg>=1.5 elo>=30','home[1.55,1.8) xg>=1.2 elo>=75','away[2.5,3.5) xg>=1.0 form>=2.2','away[2.2,2.8) form>=1.5 mkt>=0.4','away[2.2,2.8) elo<=-30 form>=1.5 mkt>=0.4'],
    'Primeira Liga': ['away[2.2,2.8) xg>=1.2 mkt>=0.4','home[2.0,2.5) form>=1.5 mkt>=0.45','home[1.55,1.8) elo>=150','home[1.8,2.2) form>=2.2'],
    'Serie B': ['home[2.2,2.8) form>=2.2','home[2.0,2.5) elo>=75 form>=2.2','home[1.8,2.2) xg>=1.0 form>=2.2','home[1.7,2.0) xg>=1.0 form>=2.2','home[1.55,1.8) xg>=1.8','home[1.55,1.8) xg>=1.5 elo>=75 form>=2.2','home[1.55,1.8) xg>=1.2 elo>=150','home[1.55,1.8) xg>=1.5','away[2.0,2.5) xg>=1.5 elo<=-30'],
    'Eredivisie': ['home[2.5,3.5) xg>=1.0','home[2.0,2.5) elo>=75','away[2.5,3.5) elo<=-75 mkt>=0.35','away[1.8,2.2) xg>=1.2 form>=2.2','away[1.7,2.0) xg>=1.2 form>=2.2','away[1.55,1.8) elo<=-150'],
    'Jupiler Pro League': ['home[2.5,3.5) form>=2.2','home[2.0,2.5) xg>=1.0 mkt>=0.45','home[1.8,2.2) xg>=1.2 mkt>=0.5','home[1.8,2.2) xg>=1.0 form>=1.5 mkt>=0.5','home[1.7,2.0) xg>=1.0 form>=1.5 mkt>=0.5','home[1.55,1.8) xg>=1.8','home[1.55,1.8) form>=2.2','home[1.3,1.55) xg>=1.5 elo>=150 form>=1.5','away[2.2,2.8) xg>=1.2 form>=2.2 mkt>=0.35','away[2.0,2.5) xg>=1.5 form>=2.2','away[1.8,2.2) xg>=1.5 form>=2.2'],
    'Champions League': ['home[2.2,2.8) xg>=1.8 form>=1.5','home[2.0,2.5) form>=1.5 mkt>=0.45','home[1.8,2.2) mkt>=0.5','home[1.7,2.0) xg>=1.5 mkt>=0.55','home[1.7,2.0) xg>=1.5 elo>=75 mkt>=0.5','home[1.55,1.8) elo>=30 form>=2.2','home[1.55,1.8) elo>=150','home[1.3,1.55) xg>=1.8 form>=1.8','away[2.2,2.8) xg>=1.8 mkt>=0.4','away[2.2,2.8) xg>=1.5 form>=1.8 mkt>=0.4','away[2.0,2.5) xg>=1.5 form>=1.8 mkt>=0.4'],
}

def parse_niche(s):
    m=re.match(r'(home|away)\[(\d+\.?\d*),(\d+\.?\d*)\)',s)
    side=m.group(1); lo=float(m.group(2)); hi=float(m.group(3))
    xg=0.0; elo=0; fm=0.0; mk=0.0
    mx=re.search(r'xg>=([\d.]+)',s);   xg=float(mx.group(1)) if mx else 0.0
    mx=re.search(r'elo>=([-\d]+)',s);  elo=int(mx.group(1)) if mx else elo
    mx=re.search(r'elo<=([-\d]+)',s);  elo=int(mx.group(1)) if mx else elo
    mx=re.search(r'form>=([\d.]+)',s); fm=float(mx.group(1)) if mx else 0.0
    mx=re.search(r'mkt>=([\d.]+)',s);  mk=float(mx.group(1)) if mx else 0.0
    return side,lo,hi,xg,elo,fm,mk

def apply_mask(data,side,lo,hi,xg,elo,fm,mk,league):
    oc='home_odds_val' if side=='home' else 'away_odds_val'
    m=(data['league_name']==league)&(data[oc]>=lo)&(data[oc]<hi)
    if xg>0:  m&=data['xg_ratio_home_5' if side=='home' else 'xg_ratio_away_5'].fillna(0)>=xg
    if side=='home' and elo>0: m&=data['elo_diff'].fillna(0)>=elo
    if side=='away' and elo<0: m&=data['elo_diff'].fillna(0)<=elo
    if fm>0:  m&=data['home_pts_5' if side=='home' else 'away_pts_5'].fillna(0)>=fm
    if mk>0:  m&=data['mkt_home_prob' if side=='home' else 'mkt_away_prob'].fillna(0)>=mk
    return data[m]

def calc_pl(bets, wc, oc):
    if len(bets)==0: return 0,0,0
    pl=(bets[wc]*(bets[oc]-1)-(1-bets[wc])).sum()*FLAT_STAKE
    w=int(bets[wc].sum()); n=len(bets)
    return pl, w, n

print('Loading data...')
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result']=='H').astype(int)
df['away_win']  = (df['result']=='A').astype(int)
df['ym'] = df['date'].dt.to_period('M')
oos_p = OOS_START.to_period('M')

# Styles
fill_hdr  = PatternFill('solid',fgColor='1F3864')
fill_lg   = PatternFill('solid',fgColor='D6E4F0')
fill_ps   = PatternFill('solid',fgColor='70AD47')
fill_p    = PatternFill('solid',fgColor='C6EFCE')
fill_neg  = PatternFill('solid',fgColor='FFCCCC')
fill_warn = PatternFill('solid',fgColor='FFE699')
fill_zero = PatternFill('solid',fgColor='F5F5F5')
thin      = Side(style='thin',color='CCCCCC')
border    = Border(left=thin,right=thin,top=thin,bottom=thin)

wb = Workbook()
wb.remove(wb.active)

all_recommendations = []

for league, niches in MODELS.items():
    print(f'{league}...')

    # Build match_id sets and stats per model
    model_info = []
    for niche_str in niches:
        side,lo,hi,xg,elo,fm,mk = parse_niche(niche_str)
        wc='home_win' if side=='home' else 'away_win'
        oc='home_odds_val' if side=='home' else 'away_odds_val'
        bets = apply_mask(df,side,lo,hi,xg,elo,fm,mk,league)
        ids  = set(bets.index.tolist())

        oos_bets = bets[bets['ym']>=oos_p]
        ins_bets = bets[bets['ym']<oos_p]

        total_pl,total_w,total_n = calc_pl(bets,wc,oc)
        oos_pl,oos_w,oos_n       = calc_pl(oos_bets,wc,oc)
        ins_pl,ins_w,ins_n       = calc_pl(ins_bets,wc,oc)

        model_info.append({
            'niche': niche_str, 'side': side, 'wc': wc, 'oc': oc,
            'bets': bets, 'ids': ids,
            'total_n':total_n,'total_w':total_w,'total_pl':total_pl,
            'roi_pct':total_pl/(total_n*FLAT_STAKE)*100 if total_n else 0,
            'wr_pct':total_w/total_n*100 if total_n else 0,
            'oos_roi':oos_pl/(oos_n*FLAT_STAKE)*100 if oos_n else 0,
            'oos_wr':oos_w/oos_n*100 if oos_n else 0,
            'oos_n':oos_n,'oos_pl':oos_pl,
            'ins_roi':ins_pl/(ins_n*FLAT_STAKE)*100 if ins_n else 0,
        })

    # For each model: split into UNIQUE vs SHARED matches
    all_ids_union = {}
    for mi in model_info:
        for idx in mi['ids']:
            all_ids_union.setdefault(idx, []).append(mi['niche'])

    for mi in model_info:
        wc=mi['wc']; oc=mi['oc']
        # Unique: only this model covers this match
        unique_ids  = {idx for idx in mi['ids'] if len(all_ids_union[idx])==1}
        shared_ids  = mi['ids'] - unique_ids

        unique_bets = mi['bets'].loc[mi['bets'].index.isin(unique_ids)]
        shared_bets = mi['bets'].loc[mi['bets'].index.isin(shared_ids)]

        u_pl,u_w,u_n = calc_pl(unique_bets,wc,oc)
        s_pl,s_w,s_n = calc_pl(shared_bets,wc,oc)

        # OOS split
        u_oos = unique_bets[unique_bets['ym']>=oos_p]
        s_oos = shared_bets[shared_bets['ym']>=oos_p]
        u_oos_pl,u_oos_w,u_oos_n = calc_pl(u_oos,wc,oc)
        s_oos_pl,s_oos_w,s_oos_n = calc_pl(s_oos,wc,oc)

        u_roi = u_pl/(u_n*FLAT_STAKE)*100 if u_n else None
        s_roi = s_pl/(s_n*FLAT_STAKE)*100 if s_n else None
        u_oos_roi = u_oos_pl/(u_oos_n*FLAT_STAKE)*100 if u_oos_n else None
        s_oos_roi = s_oos_pl/(s_oos_n*FLAT_STAKE)*100 if s_oos_n else None

        # Recommendation
        if u_n == 0:
            rec = '⚠️ Тільки дублі — перевір overlap'
        elif u_oos_n is not None and u_oos_n > 0 and u_oos_roi < 0:
            rec = '❌ Унікальні матчі OOS збиткові — ВІДСІЯТИ'
        elif u_oos_n is None or u_oos_n == 0:
            rec = '⚠️ Немає OOS унікальних матчів'
        elif u_oos_roi >= 20:
            rec = '✅ Унікальні матчі прибуткові — ЗАЛИШИТИ'
        elif u_oos_roi >= 0:
            rec = '🟡 Унікальні матчі слабо прибуткові'
        else:
            rec = '❌ Унікальні матчі збиткові — ВІДСІЯТИ'

        mi['unique_n']=u_n; mi['unique_pl']=u_pl; mi['unique_roi']=u_roi
        mi['shared_n']=s_n; mi['shared_pl']=s_pl; mi['shared_roi']=s_roi
        mi['u_oos_n']=u_oos_n; mi['u_oos_roi']=u_oos_roi
        mi['s_oos_n']=s_oos_n; mi['s_oos_roi']=s_oos_roi
        mi['rec']=rec

        all_recommendations.append({'League':league,'Niche':mi['niche'],'Rec':rec,
            'Total n':mi['total_n'],'OOS ROI%':mi['oos_roi'],
            'Unique n':u_n,'Unique OOS n':u_oos_n,'Unique OOS ROI%':u_oos_roi,
            'Shared n':s_n,'Shared OOS n':s_oos_n,'Shared OOS ROI%':s_oos_roi})

    # Write league sheet
    ws = wb.create_sheet(league[:20])
    COLS = ['Niche','Total n','Total ROI%','OOS ROI%','OOS n',
            'Unique n','Unique ROI%','Unique OOS n','Unique OOS ROI%',
            'Shared n','Shared ROI%','Shared OOS n','Shared OOS ROI%',
            'Recommendation']
    for ci,col in enumerate(COLS,1):
        c=ws.cell(1,ci,col); c.fill=fill_hdr
        c.font=Font(color='FFFFFF',bold=True,size=10)
        c.alignment=Alignment(horizontal='center',wrap_text=True)
    ws.row_dimensions[1].height=32

    for ri,mi in enumerate(model_info,2):
        def fmt(v): return round(v,1) if v is not None else '—'
        vals=[mi['niche'],mi['total_n'],fmt(mi['roi_pct']),fmt(mi['oos_roi']),mi['oos_n'],
              mi['unique_n'],fmt(mi['unique_roi']),mi['u_oos_n'] or 0,fmt(mi['u_oos_roi']),
              mi['shared_n'],fmt(mi['shared_roi']),mi['s_oos_n'] or 0,fmt(mi['s_oos_roi']),
              mi['rec']]
        for ci,val in enumerate(vals,1):
            c=ws.cell(ri,ci,val); c.border=border
            c.alignment=Alignment(horizontal='center')
            if ci in (1,14): c.alignment=Alignment(horizontal='left')
            # Color OOS ROI columns
            if ci in (4,8,12) and isinstance(val,(int,float)):
                c.number_format='+0.0;-0.0;0'
                c.fill=fill_ps if val>=20 else (fill_p if val>0 else (fill_neg if val<0 else fill_zero))
            if ci in (3,7,11) and isinstance(val,(int,float)):
                c.number_format='+0.0;-0.0;0'
            # Recommendation coloring
            if ci==14:
                if '✅' in str(val): c.fill=fill_ps; c.font=Font(bold=True,color='155724')
                elif '❌' in str(val): c.fill=fill_neg; c.font=Font(bold=True,color='721c24')
                elif '⚠️' in str(val): c.fill=fill_warn; c.font=Font(bold=True)
                elif '🟡' in str(val): c.fill=fill_warn

    widths=[40,8,10,10,7,9,12,12,14,9,12,12,14,32]
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes='A2'

# Summary sheet
ws_sum = wb.create_sheet('SUMMARY',0)
sum_cols=['League','Niche','Total n','OOS ROI%','Unique n','Unique OOS n',
          'Unique OOS ROI%','Shared n','Shared OOS n','Shared OOS ROI%','Recommendation']
for ci,col in enumerate(sum_cols,1):
    c=ws_sum.cell(1,ci,col); c.fill=fill_hdr
    c.font=Font(color='FFFFFF',bold=True,size=10)
    c.alignment=Alignment(horizontal='center',wrap_text=True)

# Sort: ❌ first, then ⚠️, then ✅
def sort_key(r):
    rec=r['Rec']
    if '❌' in rec: return 0
    if '⚠️' in rec: return 1
    if '🟡' in rec: return 2
    return 3

all_recommendations.sort(key=lambda x:(x['League'],sort_key(x)))
prev_lg=None; row_idx=2
for r in all_recommendations:
    if r['League']!=prev_lg:
        c=ws_sum.cell(row_idx,1,r['League']); c.fill=fill_lg
        c.font=Font(bold=True,size=11)
        ws_sum.merge_cells(start_row=row_idx,start_column=1,end_row=row_idx,end_column=len(sum_cols))
        row_idx+=1; prev_lg=r['League']
    oos_roi=r['OOS ROI%'] or 0; u_oos_roi=r['Unique OOS ROI%']
    vals=[r['League'],r['Niche'],r['Total n'],round(oos_roi,1),
          r['Unique n'],r['Unique OOS n'] or 0,
          round(u_oos_roi,1) if u_oos_roi is not None else '—',
          r['Shared n'],r['Shared OOS n'] or 0,
          round(r['Shared OOS ROI%'],1) if r['Shared OOS ROI%'] is not None else '—',
          r['Rec']]
    for ci,val in enumerate(vals,1):
        c=ws_sum.cell(row_idx,ci,val); c.border=border
        c.alignment=Alignment(horizontal='center')
        if ci in (2,11): c.alignment=Alignment(horizontal='left')
        if ci==4:
            c.number_format='+0.0;-0.0;0'
            c.fill=fill_ps if oos_roi>=20 else (fill_p if oos_roi>0 else fill_neg)
        if ci==7 and isinstance(val,(int,float)):
            c.number_format='+0.0;-0.0;0'
            c.fill=fill_ps if val>=20 else (fill_p if val>0 else fill_neg)
            c.font=Font(bold=True)
        if ci==11:
            rec=r['Rec']
            if '✅' in rec: c.fill=fill_ps; c.font=Font(bold=True)
            elif '❌' in rec: c.fill=fill_neg; c.font=Font(bold=True)
            elif '⚠️' in rec or '🟡' in rec: c.fill=fill_warn; c.font=Font(bold=True)
    row_idx+=1

ws_sum.column_dimensions['A'].width=16; ws_sum.column_dimensions['B'].width=42
ws_sum.column_dimensions[get_column_letter(len(sum_cols))].width=36
for ci in range(3,len(sum_cols)): ws_sum.column_dimensions[get_column_letter(ci)].width=11
ws_sum.freeze_panes='A2'

# Print summary to console
print()
print('='*80)
keep=sum(1 for r in all_recommendations if '✅' in r['Rec'])
drop=sum(1 for r in all_recommendations if '❌' in r['Rec'])
warn=sum(1 for r in all_recommendations if '⚠️' in r['Rec'] or '🟡' in r['Rec'])
print(f'✅ Залишити: {keep}   ❌ Відсіяти: {drop}   ⚠️ Перевірити: {warn}')
print()
print('❌ Кандидати на відсів:')
for r in all_recommendations:
    if '❌' in r['Rec']:
        print(f"  [{r['League']}] {r['Niche']} — unique OOS n={r['Unique OOS n']}, unique OOS ROI={r['Unique OOS ROI%']}")

wb.save(OUT)
print(f'\nSaved: {OUT}')
print('✅ Done!')
