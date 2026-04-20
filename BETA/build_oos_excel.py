"""
BETA/build_oos_excel.py

OOS validation з розширеним сплітом:
  Discovery: 2023-08 → 2024-07 (12 міс)
  OOS:       2024-08 → 2026-04 (20 міс)

Excel структура: один лист на лігу
Колонки: Filter | Side | DISC_n | DISC_WR% | DISC_ROI% |
         OOS_n | OOS_WR% | OOS_ROI% | OOS_EV% | Δ ROI | Stable?
         + 7 квартальних колонок всередині OOS (ROI% + n)
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import numpy as np
import pandas as pd
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as BdrSide
from openpyxl.utils import get_column_letter

from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix

# ── Load ──────────────────────────────────────────────────────────────────────
print("Loading data...")
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result'] == 'H').astype(int)
df['away_win']  = (df['result'] == 'A').astype(int)

DISC_END  = pd.Timestamp('2024-07-31')
OOS_START = pd.Timestamp('2024-08-01')

disc = df[df['date'] <= DISC_END].copy()
oos  = df[df['date'] >= OOS_START].copy()
print(f"Discovery: {len(disc)}  ({disc['date'].min().date()} – {disc['date'].max().date()})")
print(f"OOS:       {len(oos)}  ({oos['date'].min().date()} – {oos['date'].max().date()})")

LEAGUES = [
    'Premier League','Bundesliga','Serie A','La Liga','Ligue 1',
    'Primeira Liga','Serie B','Eredivisie','Jupiler Pro League','Champions League'
]
LEAGUE_SHORT = {
    'Premier League':'EPL','Bundesliga':'Bundesliga','Serie A':'Serie A',
    'La Liga':'La Liga','Ligue 1':'Ligue 1','Primeira Liga':'Portugal',
    'Serie B':'Serie B','Eredivisie':'Eredivisie',
    'Jupiler Pro League':'Jupiler','Champions League':'UCL'
}

# ── OOS quarterly windows ─────────────────────────────────────────────────────
OOS_WINDOWS = []
cur = OOS_START
end = pd.Timestamp('2026-04-30')
while cur < end:
    w_end = cur + pd.DateOffset(months=3) - pd.Timedelta(days=1)
    if w_end > end: w_end = end
    OOS_WINDOWS.append((cur.strftime('%Y-%m'), cur, w_end))
    cur += pd.DateOffset(months=3)
WIN_LABELS = [w[0] for w in OOS_WINDOWS]
print(f"OOS windows: {len(OOS_WINDOWS)} quarters ({WIN_LABELS[0]} → {WIN_LABELS[-1]})")

# Pre-slice OOS by window
oos_win_slices = {wlbl: oos[(oos['date']>=ws)&(oos['date']<=we)]
                  for wlbl,ws,we in OOS_WINDOWS}

# ── Filter space ──────────────────────────────────────────────────────────────
ODDS_RANGES = [
    (1.30,1.55),(1.55,1.80),(1.70,2.00),
    (1.80,2.20),(2.00,2.50),(2.20,2.80),(2.50,3.50),
]
XG_THRS    = [0.0,1.0,1.2,1.5,1.8]
ELO_H_THRS = [0,30,75,150]
ELO_A_THRS = [0,-30,-75,-150]
FORM_THRS  = [0.0,1.5,1.8,2.2]
MKT_H_THRS = [0.0,0.45,0.50,0.55]
MKT_A_THRS = [0.0,0.35,0.40,0.45]
MIN_N      = 3

def calc(sub, side):
    wc = 'home_win' if side=='home' else 'away_win'
    oc = 'home_odds_val' if side=='home' else 'away_odds_val'
    if len(sub) < MIN_N: return None
    wr    = sub[wc].mean()
    avg_o = sub[oc].mean()
    roi   = (sub[wc]*(sub[oc]-1)-(1-sub[wc])).mean()*100
    ev    = (wr*avg_o-1)*100
    return {'n':len(sub),'wr':round(wr*100,1),'roi':round(roi,1),
            'ev':round(ev,1),'avg_odds':round(avg_o,2)}

# ── Main sweep ────────────────────────────────────────────────────────────────
print("Running sweep...")
# results[lg][label] = {meta, disc:{...}, oos:{...}, wins:{wlbl:{...}}}
results = {lg: {} for lg in LEAGUES}

for side in ('home','away'):
    oc  = 'home_odds_val'   if side=='home' else 'away_odds_val'
    xgc = 'xg_ratio_home_5' if side=='home' else 'xg_ratio_away_5'
    fc  = 'home_pts_5'      if side=='home' else 'away_pts_5'
    mc  = 'mkt_home_prob'   if side=='home' else 'mkt_away_prob'
    mt  = MKT_H_THRS if side=='home' else MKT_A_THRS
    et  = ELO_H_THRS if side=='home' else ELO_A_THRS

    for lo,hi in ODDS_RANGES:
        for xg_t,elo_t,form_t,mkt_t in product(XG_THRS,et,FORM_THRS,mt):
            parts = [f'{side}[{lo},{hi})']
            if xg_t  > 0: parts.append(f'xg>={xg_t}')
            if side=='home' and elo_t>0: parts.append(f'elo>={elo_t}')
            if side=='away' and elo_t<0: parts.append(f'elo<={elo_t}')
            if form_t > 0: parts.append(f'form>={form_t}')
            if mkt_t  > 0: parts.append(f'mkt>={mkt_t}')
            label = ' '.join(parts)

            def apply_mask(d):
                m = (d[oc]>=lo)&(d[oc]<hi)
                if xg_t  > 0: m &= (d[xgc].fillna(0)>=xg_t)
                if side=='home' and elo_t>0: m &= (d['elo_diff'].fillna(0)>=elo_t)
                if side=='away' and elo_t<0: m &= (d['elo_diff'].fillna(0)<=elo_t)
                if form_t > 0: m &= (d[fc].fillna(0)>=form_t)
                if mkt_t  > 0: m &= (d[mc].fillna(0)>=mkt_t)
                return d[m]

            for lg in LEAGUES:
                # Discovery
                sub_d = apply_mask(disc)
                sub_d = sub_d[sub_d['league_name']==lg]
                m_d   = calc(sub_d, side)

                # Full OOS
                sub_o = apply_mask(oos)
                sub_o = sub_o[sub_o['league_name']==lg]
                m_o   = calc(sub_o, side)

                if m_d is None and m_o is None: continue

                if label not in results[lg]:
                    results[lg][label] = {'side':side,'lo':lo,'hi':hi,
                                           'xg_t':xg_t,'elo_t':elo_t,
                                           'form_t':form_t,'mkt_t':mkt_t,
                                           'disc':None,'oos':None,'wins':{}}
                results[lg][label]['disc'] = m_d
                results[lg][label]['oos']  = m_o

                # OOS quarters
                for wlbl,_,_ in OOS_WINDOWS:
                    sub_w = apply_mask(oos_win_slices[wlbl])
                    sub_w = sub_w[sub_w['league_name']==lg]
                    m_w   = calc(sub_w, side)
                    if m_w:
                        results[lg][label]['wins'][wlbl] = m_w

print("Sweep done. Building Excel...")

# ── Styles ────────────────────────────────────────────────────────────────────
GREEN_LIGHT = PatternFill('solid', start_color='D5F5E3')
RED_LIGHT   = PatternFill('solid', start_color='FADBD8')
GREY_FILL   = PatternFill('solid', start_color='EAECEE')
HEADER_FILL = PatternFill('solid', start_color='1B2631')
DISC_FILL   = PatternFill('solid', start_color='2E4057')
OOS_FILL    = PatternFill('solid', start_color='1A5276')
WIN_FILL    = PatternFill('solid', start_color='154360')
STAB_FILL   = {0:'D5F5E3', 1:'FEF9E7', 2:'FADBD8'}
GRP_FILL    = {'✅':'1A5276','⚠️':'7D6608','❌':'78281F'}

HDR_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=9)
BODY_FONT = Font(name='Arial', size=9)
CENTER    = Alignment(horizontal='center', vertical='center', wrap_text=False)
LEFT      = Alignment(horizontal='left',   vertical='center')

def tborder():
    s = BdrSide(style='thin', color='CCCCCC')
    return Border(left=s,right=s,top=s,bottom=s)

def hcell(c, v, fill=HEADER_FILL):
    c.value=v; c.font=HDR_FONT; c.fill=fill
    c.alignment=CENTER; c.border=tborder()

def bcell(c, v, align=CENTER, fill=None):
    c.value=v; c.font=BODY_FONT
    c.alignment=align; c.border=tborder()
    if fill: c.fill=fill

def roi_cell(c, roi_val, n_val=None, row_fill=None):
    if roi_val is None:
        bcell(c, '—'); c.fill=GREY_FILL; return
    c.value = roi_val/100
    c.font  = Font(name='Arial', size=9, bold=(abs(roi_val)>20))
    c.number_format = '+0.0%;-0.0%;0.0%'
    c.alignment = CENTER; c.border = tborder()
    c.fill = GREEN_LIGHT if roi_val>0 else RED_LIGHT

# ── Build workbook ─────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

# ── SUMMARY sheet ─────────────────────────────────────────────────────────────
ws_s = wb.create_sheet('SUMMARY')
SUM_H = ['Ліга','Side','Найкращий фільтр OOS',
          'DISC n','DISC WR%','DISC ROI%',
          'OOS n','OOS WR%','OOS ROI%','OOS EV%',
          'Δ ROI (OOS-DISC)','Stable?']
for ci,h in enumerate(SUM_H,1): hcell(ws_s.cell(1,ci), h)

sr = 2
for lg in LEAGUES:
    lname = LEAGUE_SHORT[lg]
    for side in ('home','away'):
        best = None
        for label,data in results[lg].items():
            if data['side']!=side: continue
            mo = data.get('oos')
            if not mo or mo['n']<10: continue
            if best is None or mo['roi']>best['roi']:
                best = {**mo,'label':label,
                        'disc':data.get('disc'),
                        'wins':data.get('wins',{})}
        if not best: continue
        d = best.get('disc')
        delta = round(best['roi']-(d['roi'] if d else 0),1) if d else None
        n_pos = sum(1 for w in best['wins'].values() if w['roi']>0)
        n_tot_w = len(best['wins'])
        stable = ('✅' if best['roi']>0 and n_tot_w>0 and n_pos/n_tot_w>=0.6
                  else '⚠️' if best['roi']>0 else '❌')
        row = [lname, side, best['label'],
               d['n'] if d else '—',
               (d['wr']/100) if d else '—',
               (d['roi']/100) if d else '—',
               best['n'], best['wr']/100, best['roi']/100,
               best['ev']/100,
               (delta/100) if delta is not None else '—',
               stable]
        fmts = [None,None,None,None,'0.0%','+0.0%;-0.0%',
                None,'0.0%','+0.0%;-0.0%','+0.0%;-0.0%','+0.0%;-0.0%',None]
        for ci,(val,fmt) in enumerate(zip(row,fmts),1):
            c = ws_s.cell(sr,ci)
            bcell(c, val, LEFT if ci in (1,2,3,12) else CENTER)
            if fmt: c.number_format = fmt
            if ci==9:
                c.fill = GREEN_LIGHT if best['roi']>0 else RED_LIGHT
            if ci==11 and delta is not None:
                c.fill = GREEN_LIGHT if delta>0 else RED_LIGHT
        sr += 1

ws_s.column_dimensions['A'].width = 12
ws_s.column_dimensions['B'].width = 6
ws_s.column_dimensions['C'].width = 55
for ci in range(4,13):
    ws_s.column_dimensions[get_column_letter(ci)].width = 12
ws_s.freeze_panes = 'A2'
ws_s.auto_filter.ref = f'A1:{get_column_letter(len(SUM_H))}1'

# ── Per-league sheets ──────────────────────────────────────────────────────────
for lg in LEAGUES:
    lname = LEAGUE_SHORT[lg]
    print(f"  Sheet: {lname}...")
    ws = wb.create_sheet(lname)

    # Static cols
    STATIC = ['Filter','Side',
              'DISC n','DISC WR%','DISC ROI%',
              'OOS n','OOS WR%','OOS ROI%','OOS EV%',
              'Δ ROI','Win%','W+/Wtot','Stable']
    NS = len(STATIC)
    NW = len(OOS_WINDOWS)
    total_cols = NS + NW*2

    # Row 1: static headers + window group headers
    for ci,h in enumerate(STATIC,1):
        hcell(ws.cell(1,ci), h)
    for wi,(wlbl,_,_) in enumerate(OOS_WINDOWS):
        cs = NS+wi*2+1
        ws.merge_cells(start_row=1,start_column=cs,end_row=1,end_column=cs+1)
        hcell(ws.cell(1,cs), wlbl, fill=WIN_FILL)

    # Row 2: sub-headers
    for ci in range(1,NS+1):
        hcell(ws.cell(2,ci), '', fill=PatternFill('solid',start_color='2C3E50'))
    for wi in range(NW):
        cs = NS+wi*2+1
        hcell(ws.cell(2,cs),   'ROI%', fill=PatternFill('solid',start_color='1F618D'))
        hcell(ws.cell(2,cs+1), 'n',    fill=PatternFill('solid',start_color='1F618D'))

    # Collect & sort all rows
    all_rows = []
    for label,data in results[lg].items():
        mo = data.get('oos')
        md = data.get('disc')
        if not mo and not md: continue
        wins = data.get('wins',{})
        n_pos = sum(1 for w in wins.values() if w['roi']>0)
        n_w   = len(wins)
        win_pct = round(n_pos/n_w*100) if n_w>0 else 0
        oos_roi = mo['roi'] if mo else None
        stable = ('✅' if oos_roi and oos_roi>0 and n_w>0 and n_pos/n_w>=0.6
                  else '⚠️' if oos_roi and oos_roi>0
                  else '❌')
        delta = None
        if mo and md:
            delta = round(mo['roi']-md['roi'],1)
        all_rows.append({'label':label,'data':data,'md':md,'mo':mo,
                         'wins':wins,'win_pct':win_pct,
                         'n_pos':n_pos,'n_w':n_w,
                         'stable':stable,'delta':delta,
                         'oos_roi':oos_roi or -999})

    def sort_key(r):
        grp = (0 if r['stable']=='✅' else 1 if r['stable']=='⚠️' else 2)
        return (grp, -(r['oos_roi'] or -999))
    all_rows.sort(key=sort_key)

    er = 3
    prev_grp = None
    grp_titles = {
        '✅': '▶ СТАБІЛЬНО ПРИБУТКОВІ OOS  (roi>0, win%≥60%)',
        '⚠️': '▶ ПОМІРНО ПРИБУТКОВІ OOS  (roi>0, win%<60%)',
        '❌': '▶ ЗБИТКОВІ OOS  (roi≤0)',
    }
    grp_colors = {'✅':'1A5276','⚠️':'784212','❌':'641E16'}

    for r in all_rows:
        grp = r['stable']
        if grp != prev_grp:
            ws.merge_cells(start_row=er,start_column=1,end_row=er,end_column=total_cols)
            c = ws.cell(er,1)
            c.value = grp_titles[grp]
            c.font  = Font(name='Arial',bold=True,color='FFFFFF',size=10)
            c.fill  = PatternFill('solid',start_color=grp_colors[grp])
            c.alignment = LEFT
            er += 1
            prev_grp = grp

        mo = r['mo']
        md = r['md']
        rfill = PatternFill('solid', start_color=STAB_FILL[
            0 if grp=='✅' else 1 if grp=='⚠️' else 2])

        static_vals = [
            r['label'],
            r['data']['side'],
            md['n']   if md else '—',
            (md['wr']/100)  if md else '—',
            (md['roi']/100) if md else '—',
            mo['n']   if mo else '—',
            (mo['wr']/100)  if mo else '—',
            (mo['roi']/100) if mo else '—',
            (mo['ev']/100)  if mo else '—',
            (r['delta']/100) if r['delta'] is not None else '—',
            (r['win_pct']/100) if r['n_w']>0 else '—',
            f"{r['n_pos']}/{r['n_w']}" if r['n_w']>0 else '—',
            r['stable'],
        ]
        fmts = [None,None,None,'0.0%','+0.0%;-0.0%',
                None,'0.0%','+0.0%;-0.0%','+0.0%;-0.0%',
                '+0.0%;-0.0%','0%',None,None]

        for ci,(val,fmt) in enumerate(zip(static_vals,fmts),1):
            c = ws.cell(er,ci)
            bcell(c, val, LEFT if ci in (1,2,13) else CENTER, rfill)
            if fmt and isinstance(val,(int,float)): c.number_format = fmt
            # Colour ROI cells
            if ci==5 and md:
                c.fill = GREEN_LIGHT if md['roi']>0 else RED_LIGHT
            if ci==8 and mo:
                c.fill = GREEN_LIGHT if mo['roi']>0 else RED_LIGHT
                c.font = Font(name='Arial',size=9,bold=(abs(mo['roi'])>20))
            if ci==10 and r['delta'] is not None:
                c.fill = GREEN_LIGHT if r['delta']>0 else RED_LIGHT

        # Window cells
        for wi,(wlbl,_,_) in enumerate(OOS_WINDOWS):
            col_roi = NS+wi*2+1
            col_n   = NS+wi*2+2
            w = r['wins'].get(wlbl)
            if w:
                c_r = ws.cell(er,col_roi)
                c_r.value = w['roi']/100
                c_r.font  = Font(name='Arial',size=9,bold=(abs(w['roi'])>20))
                c_r.number_format = '+0.0%;-0.0%;0.0%'
                c_r.alignment = CENTER; c_r.border = tborder()
                c_r.fill = GREEN_LIGHT if w['roi']>0 else RED_LIGHT
                c_n = ws.cell(er,col_n)
                bcell(c_n, w['n'], fill=rfill)
            else:
                for col in (col_roi,col_n):
                    c = ws.cell(er,col)
                    bcell(c,'—'); c.fill=GREY_FILL
        er += 1

    # Column widths
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 6
    for ci in range(3,NS+1):
        ws.column_dimensions[get_column_letter(ci)].width = 9
    for wi in range(NW):
        ws.column_dimensions[get_column_letter(NS+wi*2+1)].width = 8
        ws.column_dimensions[get_column_letter(NS+wi*2+2)].width = 4

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f'A2:{get_column_letter(NS)}2'
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 16

outpath = os.path.join(os.path.dirname(__file__), 'oos_extended_results.xlsx')
wb.save(outpath)
print(f"\nSaved: {outpath}  ({os.path.getsize(outpath)//1024}KB)")
print("Done!")
