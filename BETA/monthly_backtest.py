import warnings; warnings.filterwarnings('ignore')
import pandas as pd
import numpy as np
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
sys.path.insert(0, '.')
from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix

OUT        = '/app/BETA/monthly_backtest.xlsx'
FLAT_STAKE = 50.0

LEAGUES = [
    'Premier League','Bundesliga','Serie A','La Liga','Ligue 1',
    'Primeira Liga','Serie B','Eredivisie','Jupiler Pro League','Champions League'
]
ODDS_RANGES = [(1.30,1.55),(1.55,1.80),(1.70,2.00),(1.80,2.20),(2.00,2.50),(2.20,2.80),(2.50,3.50)]
XG  = [0.0, 1.0, 1.2, 1.5, 1.8]
EH  = [0,  30,  75, 150]
EA  = [0, -30, -75, -150]
FT  = [0.0, 1.5, 1.8, 2.2]
MH  = [0.0, 0.45, 0.50, 0.55]
MA  = [0.0, 0.35, 0.40, 0.45]

OOS_START = pd.Timestamp('2025-08-01')

print('Loading data...')
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result'] == 'H').astype(int)
df['away_win']  = (df['result'] == 'A').astype(int)
df['ym'] = df['date'].dt.to_period('M')

# All months in range
all_months = pd.period_range(df['date'].min().to_period('M'),
                              df['date'].max().to_period('M'), freq='M')
print(f'Data: {df["date"].min().date()} → {df["date"].max().date()}  ({len(all_months)} months)')
print(f'Matches: {len(df)}')

def niche_label(side, lo, hi, xg, elo, fm, mk):
    s = f'{side}[{lo},{hi})'
    if xg  > 0: s += f' xg>={xg}'
    if elo != 0: s += f' elo>={elo}' if elo > 0 else f' elo<={elo}'
    if fm  > 0: s += f' form>={fm}'
    if mk  > 0: s += f' mkt>={mk}'
    return s

def apply_mask(data, side, lo, hi, xg, elo, fm, mk, league):
    oc = 'home_odds_val' if side == 'home' else 'away_odds_val'
    m = (data['league_name'] == league) & (data[oc] >= lo) & (data[oc] < hi)
    if xg > 0:
        m &= data['xg_ratio_home_5' if side=='home' else 'xg_ratio_away_5'].fillna(0) >= xg
    if side == 'home' and elo > 0:
        m &= data['elo_diff'].fillna(0) >= elo
    if side == 'away' and elo < 0:
        m &= data['elo_diff'].fillna(0) <= elo
    if fm > 0:
        m &= data['home_pts_5' if side=='home' else 'away_pts_5'].fillna(0) >= fm
    if mk > 0:
        m &= data['mkt_home_prob' if side=='home' else 'mkt_away_prob'].fillna(0) >= mk
    return data[m]

print('Running all combinations...')
results = []
total = 0

for side in ('home', 'away'):
    et = EH if side == 'home' else EA
    mt = MH if side == 'home' else MA
    wc = 'home_win' if side == 'home' else 'away_win'
    oc = 'home_odds_val' if side == 'home' else 'away_odds_val'

    for lo, hi in ODDS_RANGES:
        for xg, elo, fm, mk in product(XG, et, FT, mt):
            for lg in LEAGUES:
                total += 1
                bets = apply_mask(df, side, lo, hi, xg, elo, fm, mk, lg)
                if len(bets) < 5:
                    continue

                # Monthly stats
                monthly = {}
                for ym, grp in bets.groupby('ym'):
                    pl = (grp[wc] * (grp[oc]-1) - (1-grp[wc])).sum() * FLAT_STAKE
                    monthly[ym] = {'n': len(grp), 'pl': round(pl, 1)}

                total_pl = sum(v['pl'] for v in monthly.values())
                total_n  = sum(v['n']  for v in monthly.values())
                if total_pl <= 0:
                    continue  # remove negatives

                n_filters = niche_label(side,lo,hi,xg,elo,fm,mk).count('>=') + \
                            niche_label(side,lo,hi,xg,elo,fm,mk).count('<=')

                results.append({
                    'league':     lg,
                    'side':       side,
                    'lo':         lo,
                    'hi':         hi,
                    'xg':         xg,
                    'elo':        elo,
                    'fm':         fm,
                    'mk':         mk,
                    'niche':      niche_label(side,lo,hi,xg,elo,fm,mk),
                    'n_filters':  n_filters,
                    'total_n':    total_n,
                    'total_pl':   total_pl,
                    'roi_pct':    total_pl / (total_n * FLAT_STAKE) * 100,
                    'monthly':    monthly,
                })

print(f'Checked: {total}  |  Profitable: {len(results)}')

# ── Dedup: remove subset niches where broader niche is also profitable ──────
# Sort: fewer filters first (broader niches)
results.sort(key=lambda x: (x['league'], x['side'], x['lo'], x['hi'], x['n_filters']))

kept = []
profitable_parents = set()  # (league, side, lo, hi) that are already covered

for r in results:
    key = (r['league'], r['side'], r['lo'], r['hi'])
    if key in profitable_parents and r['n_filters'] > 0:
        # Parent (broader) niche already kept — skip this subset
        continue
    kept.append(r)
    if r['n_filters'] == 0:
        profitable_parents.add(key)

print(f'After dedup: {len(kept)}')

# Sort final output: by league, then total ROI desc
kept.sort(key=lambda x: (x['league'], -x['roi_pct']))

# ── Excel ─────────────────────────────────────────────────────────────────────
print('Writing Excel...')
wb = Workbook()
ws = wb.active
ws.title = 'Monthly Backtest'

fill_hdr    = PatternFill('solid', fgColor='1F3864')
fill_oos    = PatternFill('solid', fgColor='2E4057')   # OOS header
fill_lg     = PatternFill('solid', fgColor='D6E4F0')
fill_pos_s  = PatternFill('solid', fgColor='70AD47')   # strong positive
fill_pos    = PatternFill('solid', fgColor='C6EFCE')   # positive
fill_neg    = PatternFill('solid', fgColor='FFCCCC')   # negative
fill_zero   = PatternFill('solid', fgColor='F5F5F5')   # no bets
thin        = Side(style='thin', color='CCCCCC')
border      = Border(left=thin, right=thin, top=thin, bottom=thin)

META_COLS = ['League','Niche','Side','Total Bets','Total P&L ($)','ROI%']
n_meta = len(META_COLS)

# Row 1: meta headers + month headers
for ci, col in enumerate(META_COLS, 1):
    c = ws.cell(1, ci, col)
    c.fill = fill_hdr
    c.font = Font(color='FFFFFF', bold=True, size=10)
    c.alignment = Alignment(horizontal='center', wrap_text=True)

for mi, m in enumerate(all_months):
    col_n  = n_meta + 1 + mi * 2
    col_pl = col_n + 1
    label  = m.strftime('%b-%y')
    is_oos = m >= OOS_START.to_period('M')
    hdr_fill = fill_oos if is_oos else fill_hdr

    c = ws.cell(1, col_n, f'{label}\nn')
    c.fill = hdr_fill
    c.font = Font(color='FFFFFF', bold=True, size=9)
    c.alignment = Alignment(horizontal='center', wrap_text=True)

    c = ws.cell(1, col_pl, f'{label}\nP&L')
    c.fill = hdr_fill
    c.font = Font(color='FFFFFF', bold=True, size=9)
    c.alignment = Alignment(horizontal='center', wrap_text=True)

ws.row_dimensions[1].height = 30

row_idx = 2
prev_lg = None

for r in kept:
    if r['league'] != prev_lg:
        total_cols = n_meta + len(all_months) * 2
        c = ws.cell(row_idx, 1, r['league'])
        c.fill = fill_lg
        c.font = Font(bold=True, size=11)
        ws.merge_cells(start_row=row_idx, start_column=1,
                       end_row=row_idx, end_column=total_cols)
        row_idx += 1
        prev_lg = r['league']

    # Meta columns
    meta_vals = [r['league'], r['niche'], r['side'],
                 r['total_n'], round(r['total_pl'],1), round(r['roi_pct'],1)]
    for ci, val in enumerate(meta_vals, 1):
        c = ws.cell(row_idx, ci, val)
        c.border = border
        c.alignment = Alignment(horizontal='center')
        if ci == 2: c.alignment = Alignment(horizontal='left')
        if ci == 6:
            c.number_format = '+0.0;-0.0;0'
            c.fill = fill_pos_s if val >= 15 else fill_pos

    # Monthly columns
    for mi, m in enumerate(all_months):
        col_n  = n_meta + 1 + mi * 2
        col_pl = col_n + 1
        md = r['monthly'].get(m, None)

        if md is None:
            for ci in (col_n, col_pl):
                c = ws.cell(row_idx, ci, '')
                c.fill = fill_zero
                c.border = border
        else:
            c_n = ws.cell(row_idx, col_n, md['n'])
            c_n.border = border
            c_n.alignment = Alignment(horizontal='center')
            c_n.font = Font(size=9)

            c_pl = ws.cell(row_idx, col_pl, md['pl'])
            c_pl.border = border
            c_pl.alignment = Alignment(horizontal='center')
            c_pl.number_format = '+0;-0;0'
            c_pl.font = Font(size=9)
            if md['pl'] > 0:
                c_pl.fill = fill_pos_s if md['pl'] >= 100 else fill_pos
            elif md['pl'] < 0:
                c_pl.fill = fill_neg

    row_idx += 1

# Column widths
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 6
ws.column_dimensions['D'].width = 9
ws.column_dimensions['E'].width = 11
ws.column_dimensions['F'].width = 7
for mi in range(len(all_months)):
    for offset in (0, 1):
        col_letter = get_column_letter(n_meta + 1 + mi * 2 + offset)
        ws.column_dimensions[col_letter].width = 5

ws.freeze_panes = f'{get_column_letter(n_meta+1)}2'
ws.auto_filter.ref = f'A1:{get_column_letter(n_meta)}1'

wb.save(OUT)
print(f'\nSaved: {OUT}')
print(f'Rows: {len(kept)}')
print('✅ Done!')
