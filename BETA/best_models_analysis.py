import warnings; warnings.filterwarnings('ignore')
import pandas as pd
import numpy as np
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
sys.path.insert(0, '.')
from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix

OUT       = '/app/BETA/best_models_analysis.xlsx'
FLAT_STAKE = 50.0
OOS_START  = pd.Timestamp('2025-08-01')

# ── Parse best models ─────────────────────────────────────────────────────────
MODELS = {
    'Premier League': [
        'home[2.5,3.5) xg>=1.0',
        'home[2.2,2.8) xg>=1.2 form>=1.5',
        'home[1.8,2.2) form>=2.2',
        'home[1.7,2.0) xg>=1.8',
        'away[1.55,1.8) xg>=1.2 elo<=-75 form>=2.2',
    ],
    'Bundesliga': [
        'home[2.2,2.8) xg>=1.2 form>=1.5',
        'home[1.8,2.2) xg>=1.2 elo>=30 mkt>=0.5',
        'home[1.8,2.2) elo>=75 form>=1.5 mkt>=0.5',
        'home[1.55,1.8) elo>=150 form>=1.5',
        'away[2.5,3.5) xg>=1.0 elo<=-30 form>=1.5',
        'away[2.5,3.5) form>=1.5 mkt>=0.35',
        'away[2.5,3.5) elo<=-75 mkt>=0.35',
        'away[2.5,3.5) elo<=-30 form>=1.5 mkt>=0.35',
        'away[2.5,3.5) elo<=-150',
        'away[2.2,2.8) xg>=1.5 mkt>=0.4',
        'away[2.2,2.8) mkt>=0.4',
        'away[2.2,2.8) elo<=-75 mkt>=0.35',
        'away[2.0,2.5) mkt>=0.4',
        'away[1.7,2.0) xg>=1.0 form>=2.2',
    ],
    'Serie A': [
        'home[2.5,3.5) elo>=30 form>=1.5',
        'home[2.0,2.5) xg>=1.0 mkt>=0.45',
        'home[2.0,2.5) form>=2.2 mkt>=0.45',
        'home[1.8,2.2) xg>=1.5 form>=1.5',
        'home[1.3,1.55) xg>=1.5 elo>=150',
        'away[2.5,3.5) xg>=1.0 form>=1.8 mkt>=0.35',
        'away[2.2,2.8) xg>=1.2 form>=1.8 mkt>=0.35',
        'away[2.2,2.8) xg>=1.0 elo<=-150',
        'away[1.8,2.2) xg>=1.5 form>=2.2',
        'away[1.7,2.0) xg>=1.5 form>=2.2',
    ],
    'La Liga': [
        'home[2.2,2.8) xg>=1.0 form>=1.5',
        'home[2.0,2.5) xg>=1.0 form>=1.5',
        'home[1.8,2.2) xg>=1.2 form>=1.5',
        'home[1.8,2.2) xg>=1.0 mkt>=0.5',
        'home[1.8,2.2) xg>=1.0 form>=1.5 mkt>=0.5',
        'home[1.8,2.2) xg>=1.0 elo>=30 mkt>=0.5',
        'home[1.55,1.8) xg>=1.8',
        'away[2.5,3.5) xg>=1.8 form>=1.5',
        'away[2.5,3.5) elo<=-75 form>=2.2',
        'away[2.2,2.8) xg>=1.2 mkt>=0.4',
        'away[1.8,2.2) xg>=1.8 elo<=-75 mkt>=0.45',
    ],
    'Ligue 1': [
        'home[2.5,3.5) xg>=1.5',
        'home[2.2,2.8) xg>=1.2',
        'home[2.0,2.5) elo>=75',
        'home[1.7,2.0) xg>=1.8',
        'home[1.7,2.0) xg>=1.5 mkt>=0.5',
        'home[1.7,2.0) xg>=1.5 elo>=75',
        'home[1.7,2.0) xg>=1.5 elo>=30',
        'home[1.55,1.8) xg>=1.2 elo>=75',
        'away[2.5,3.5) xg>=1.0 form>=2.2',
        'away[2.2,2.8) form>=1.5 mkt>=0.4',
        'away[2.2,2.8) elo<=-30 form>=1.5 mkt>=0.4',
    ],
    'Primeira Liga': [
        'away[2.2,2.8) xg>=1.2 mkt>=0.4',
        'home[2.0,2.5) form>=1.5 mkt>=0.45',
        'home[1.55,1.8) elo>=150',
        'home[1.8,2.2) form>=2.2',
    ],
    'Serie B': [
        'home[2.2,2.8) form>=2.2',
        'home[2.0,2.5) elo>=75 form>=2.2',
        'home[1.8,2.2) xg>=1.0 form>=2.2',
        'home[1.7,2.0) xg>=1.0 form>=2.2',
        'home[1.55,1.8) xg>=1.8',
        'home[1.55,1.8) xg>=1.5 elo>=75 form>=2.2',
        'home[1.55,1.8) xg>=1.2 elo>=150',
        'home[1.55,1.8) xg>=1.5',
        'away[2.0,2.5) xg>=1.5 elo<=-30',
    ],
    'Eredivisie': [
        'home[2.5,3.5) xg>=1.0',
        'home[2.0,2.5) elo>=75',
        'away[2.5,3.5) elo<=-75 mkt>=0.35',
        'away[1.8,2.2) xg>=1.2 form>=2.2',
        'away[1.7,2.0) xg>=1.2 form>=2.2',
        'away[1.55,1.8) elo<=-150',
    ],
    'Jupiler Pro League': [
        'home[2.5,3.5) form>=2.2',
        'home[2.0,2.5) xg>=1.0 mkt>=0.45',
        'home[1.8,2.2) xg>=1.2 mkt>=0.5',
        'home[1.8,2.2) xg>=1.0 form>=1.5 mkt>=0.5',
        'home[1.7,2.0) xg>=1.0 form>=1.5 mkt>=0.5',
        'home[1.55,1.8) xg>=1.8',
        'home[1.55,1.8) form>=2.2',
        'home[1.3,1.55) xg>=1.5 elo>=150 form>=1.5',
        'away[2.2,2.8) xg>=1.2 form>=2.2 mkt>=0.35',
        'away[2.0,2.5) xg>=1.5 form>=2.2',
        'away[1.8,2.2) xg>=1.5 form>=2.2',
    ],
    'Champions League': [
        'home[2.2,2.8) xg>=1.8 form>=1.5',
        'home[2.0,2.5) form>=1.5 mkt>=0.45',
        'home[1.8,2.2) mkt>=0.5',
        'home[1.7,2.0) xg>=1.5 mkt>=0.55',
        'home[1.7,2.0) xg>=1.5 elo>=75 mkt>=0.5',
        'home[1.55,1.8) elo>=30 form>=2.2',
        'home[1.55,1.8) elo>=150',
        'home[1.3,1.55) xg>=1.8 form>=1.8',
        'away[2.2,2.8) xg>=1.8 mkt>=0.4',
        'away[2.2,2.8) xg>=1.5 form>=1.8 mkt>=0.4',
        'away[2.0,2.5) xg>=1.5 form>=1.8 mkt>=0.4',
    ],
}

def parse_niche(s):
    m = re.match(r'(home|away)\[(\d+\.?\d*),(\d+\.?\d*)\)', s)
    side=m.group(1); lo=float(m.group(2)); hi=float(m.group(3))
    xg=0.0; elo=0; fm=0.0; mk=0.0
    mx=re.search(r'xg>=([\d.]+)',s);   xg=float(mx.group(1)) if mx else 0.0
    mx=re.search(r'elo>=([-\d]+)',s);  elo=int(mx.group(1)) if mx else elo
    mx=re.search(r'elo<=([-\d]+)',s);  elo=int(mx.group(1)) if mx else elo
    mx=re.search(r'form>=([\d.]+)',s); fm=float(mx.group(1)) if mx else 0.0
    mx=re.search(r'mkt>=([\d.]+)',s);  mk=float(mx.group(1)) if mx else 0.0
    return side,lo,hi,xg,elo,fm,mk

def apply_mask(data,side,lo,hi,xg,elo,fm,mk,league):
    oc='home_odds_val' if side=='home' else 'away_odds_val'
    m=(data['league_name']==league)&(data[oc]>=lo)&(data[oc]<hi)
    if xg>0:  m&=data['xg_ratio_home_5' if side=='home' else 'xg_ratio_away_5'].fillna(0)>=xg
    if side=='home' and elo>0: m&=data['elo_diff'].fillna(0)>=elo
    if side=='away' and elo<0: m&=data['elo_diff'].fillna(0)<=elo
    if fm>0:  m&=data['home_pts_5' if side=='home' else 'away_pts_5'].fillna(0)>=fm
    if mk>0:  m&=data['mkt_home_prob' if side=='home' else 'mkt_away_prob'].fillna(0)>=mk
    return data[m]

# ── Load data ─────────────────────────────────────────────────────────────────
print('Loading data...')
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result']=='H').astype(int)
df['away_win']  = (df['result']=='A').astype(int)
df['ym'] = df['date'].dt.to_period('M')
df['_id'] = np.arange(len(df))
oos_p = OOS_START.to_period('M')

all_months = pd.period_range(df['date'].min().to_period('M'),
                              df['date'].max().to_period('M'), freq='M')

# ── Styles ─────────────────────────────────────────────────────────────────────
fill_hdr   = PatternFill('solid',fgColor='1F3864')
fill_oos_h = PatternFill('solid',fgColor='2E4057')
fill_lg    = PatternFill('solid',fgColor='D6E4F0')
fill_ps    = PatternFill('solid',fgColor='70AD47')
fill_p     = PatternFill('solid',fgColor='C6EFCE')
fill_neg   = PatternFill('solid',fgColor='FFCCCC')
fill_zero  = PatternFill('solid',fgColor='F5F5F5')
fill_ovlp  = PatternFill('solid',fgColor='FFE699')
thin       = Side(style='thin',color='CCCCCC')
border     = Border(left=thin,right=thin,top=thin,bottom=thin)

wb = Workbook()
wb.remove(wb.active)

all_match_rows = []  # for global match list

for league, niches in MODELS.items():
    print(f'\n{league} ({len(niches)} niches)...')
    lg_key = 'Champions League' if league == 'Champions League' else league

    # Compute stats + match IDs per niche
    niche_data = []
    for niche_str in niches:
        side,lo,hi,xg,elo,fm,mk = parse_niche(niche_str)
        wc = 'home_win' if side=='home' else 'away_win'
        oc = 'home_odds_val' if side=='home' else 'away_odds_val'
        bets = apply_mask(df,side,lo,hi,xg,elo,fm,mk,lg_key)
        if len(bets)==0:
            niche_data.append({'niche':niche_str,'bets':pd.DataFrame(),'ids':set()})
            continue

        monthly={}
        for ym,grp in bets.groupby('ym'):
            pl=(grp[wc]*(grp[oc]-1)-(1-grp[wc])).sum()*FLAT_STAKE
            n=len(grp); w=int(grp[wc].sum())
            monthly[ym]={'n':n,'w':w,'pl':round(pl,1),'roi':round(pl/(n*FLAT_STAKE)*100,1)}

        total_n=sum(v['n'] for v in monthly.values())
        total_w=sum(v['w'] for v in monthly.values())
        total_pl=sum(v['pl'] for v in monthly.values())
        ins_items=[(m,v) for m,v in monthly.items() if m<oos_p]
        oos_items=[(m,v) for m,v in monthly.items() if m>=oos_p]
        ins_pl=sum(v['pl'] for _,v in ins_items); ins_n=sum(v['n'] for _,v in ins_items); ins_w=sum(v['w'] for _,v in ins_items)
        oos_pl=sum(v['pl'] for _,v in oos_items); oos_n=sum(v['n'] for _,v in oos_items); oos_w=sum(v['w'] for _,v in oos_items)

        niche_data.append({
            'niche': niche_str, 'side': side, 'bets': bets, 'monthly': monthly,
            'ids': set(bets['_id'].tolist()),
            'total_n':total_n,'total_w':total_w,'total_pl':total_pl,
            'roi_pct':total_pl/(total_n*FLAT_STAKE)*100 if total_n else 0,
            'wr_pct':total_w/total_n*100 if total_n else 0,
            'ins_roi':ins_pl/(ins_n*FLAT_STAKE)*100 if ins_n else 0,
            'ins_wr':ins_w/ins_n*100 if ins_n else 0,
            'oos_roi':oos_pl/(oos_n*FLAT_STAKE)*100 if oos_n else 0,
            'oos_wr':oos_w/oos_n*100 if oos_n else 0,
            'oos_n':oos_n,'ins_n':ins_n,
            'oc': oc, 'wc': wc,
        })
        print(f'  {niche_str}: n={total_n} roi={total_pl/(total_n*FLAT_STAKE)*100:.1f}% oos={oos_pl/(oos_n*FLAT_STAKE)*100 if oos_n else 0:.1f}%')

    # ── League sheet: Stats + Monthly ─────────────────────────────────────────
    META=['Niche','Side','Total n','Total P&L','ROI%','WR%','IS ROI%','IS WR%','OOS ROI%','OOS WR%','OOS n']
    N_META=len(META)
    ws=wb.create_sheet(league[:20])
    for ci,col in enumerate(META,1):
        c=ws.cell(1,ci,col); c.fill=fill_hdr
        c.font=Font(color='FFFFFF',bold=True,size=10)
        c.alignment=Alignment(horizontal='center',wrap_text=True)
    for mi,m in enumerate(all_months):
        base=N_META+1+mi*3; lbl=m.strftime('%b-%y')
        hf=fill_oos_h if m>=oos_p else fill_hdr
        for off,sub in enumerate(['n','P&L','ROI%']):
            c=ws.cell(1,base+off,f'{lbl}\n{sub}')
            c.fill=hf; c.font=Font(color='FFFFFF',bold=True,size=8)
            c.alignment=Alignment(horizontal='center',wrap_text=True)
    ws.row_dimensions[1].height=28

    row_idx=2
    for r in niche_data:
        if not r.get('total_n'): continue
        roi=r['roi_pct']; oos_roi=r['oos_roi']; ins_roi=r['ins_roi']
        mv=[r['niche'],r['side'],r['total_n'],round(r['total_pl'],0),
            round(roi,1),round(r['wr_pct'],1),
            round(ins_roi,1),round(r['ins_wr'],1),
            round(oos_roi,1),round(r['oos_wr'],1),int(r['oos_n'])]
        for ci,val in enumerate(mv,1):
            c=ws.cell(row_idx,ci,val); c.border=border
            c.alignment=Alignment(horizontal='center')
            if ci==1: c.alignment=Alignment(horizontal='left')
            if ci in (5,6): c.number_format='+0.0;-0.0;0'
            if ci==7:
                c.number_format='+0.0;-0.0;0'
                c.fill=fill_ps if ins_roi>=15 else (fill_p if ins_roi>0 else fill_neg)
            if ci==8: c.number_format='0.0'
            if ci==9:
                c.number_format='+0.0;-0.0;0'
                c.fill=fill_ps if oos_roi>=15 else (fill_p if oos_roi>0 else fill_neg)
                c.font=Font(bold=True)
            if ci==10: c.number_format='0.0'; c.font=Font(bold=True)
        for mi,m in enumerate(all_months):
            base=N_META+1+mi*3; md=r['monthly'].get(m)
            if md is None:
                for off in range(3):
                    c=ws.cell(row_idx,base+off,''); c.fill=fill_zero; c.border=border
            else:
                c=ws.cell(row_idx,base,md['n']); c.border=border
                c.alignment=Alignment(horizontal='center'); c.font=Font(size=8)
                c=ws.cell(row_idx,base+1,md['pl']); c.border=border
                c.number_format='+0;-0;0'; c.font=Font(size=8); c.alignment=Alignment(horizontal='center')
                c.fill=fill_ps if md['pl']>=100 else (fill_p if md['pl']>0 else (fill_neg if md['pl']<0 else fill_zero))
                c=ws.cell(row_idx,base+2,md['roi']); c.border=border
                c.number_format='+0;-0;0'; c.font=Font(size=8); c.alignment=Alignment(horizontal='center')
                c.fill=fill_ps if md['roi']>=20 else (fill_p if md['roi']>0 else (fill_neg if md['roi']<0 else fill_zero))
        row_idx+=1

    # Blank row, then overlap matrix
    row_idx+=1
    valid = [r for r in niche_data if r.get('total_n')]
    if len(valid)>1:
        c=ws.cell(row_idx,1,'OVERLAP MATRIX (% of row niche matched by column niche)')
        c.font=Font(bold=True,size=11); c.fill=fill_ovlp
        ws.merge_cells(start_row=row_idx,start_column=1,end_row=row_idx,end_column=min(len(valid)+1,20))
        row_idx+=1
        # Header row
        ws.cell(row_idx,1,'Niche \\ Niche').font=Font(bold=True,size=8)
        for ci,r in enumerate(valid,2):
            c=ws.cell(row_idx,ci,r['niche'][:25])
            c.font=Font(bold=True,size=7); c.alignment=Alignment(horizontal='center',wrap_text=True)
            c.fill=fill_hdr; c.font=Font(color='FFFFFF',bold=True,size=7)
        ws.row_dimensions[row_idx].height=40
        row_idx+=1
        for ri,ra in enumerate(valid):
            ws.cell(row_idx,1,ra['niche'][:30]).font=Font(bold=True,size=8)
            for ci,rb in enumerate(valid,2):
                if ra is rb:
                    c=ws.cell(row_idx,ci,'—'); c.fill=fill_zero
                else:
                    if len(ra['ids'])>0:
                        ovlp=len(ra['ids']&rb['ids'])/len(ra['ids'])*100
                    else:
                        ovlp=0
                    c=ws.cell(row_idx,ci,round(ovlp,0))
                    c.number_format='0"%"'
                    c.alignment=Alignment(horizontal='center')
                    c.font=Font(size=8)
                    if ovlp>=70: c.fill=fill_neg
                    elif ovlp>=40: c.fill=fill_ovlp
                    else: c.fill=fill_p
                c.border=border
            row_idx+=1

    ws.column_dimensions['A'].width=42; ws.column_dimensions['B'].width=6
    for ci in range(3,N_META+1): ws.column_dimensions[get_column_letter(ci)].width=9
    for mi in range(len(all_months)):
        for off in range(3): ws.column_dimensions[get_column_letter(N_META+1+mi*3+off)].width=4.5
    ws.freeze_panes=f'{get_column_letter(N_META+1)}2'

    # Collect match rows for global list
    for r in niche_data:
        if not r.get('total_n'): continue
        bets = r['bets']
        wc=r['wc']; oc=r['oc']
        for _,bet in bets.iterrows():
            won=bool(bet[wc]); odds=float(bet[oc])
            pl=FLAT_STAKE*(odds-1) if won else -FLAT_STAKE
            all_match_rows.append({
                'League':league,
                'Model':r['niche'],
                'Date':bet['date'].date() if hasattr(bet['date'],'date') else bet['date'],
                'Home':bet.get('home_team',''),
                'Away':bet.get('away_team',''),
                'Bet Side':r['side'],
                'Odds':round(odds,2),
                'Result':bet.get('result',''),
                'Won':'✓' if won else '✗',
                'P&L':round(pl,1),
                'OOS':'OOS' if bet['ym']>=oos_p else 'IS',
            })

# ── Global Match List sheet ───────────────────────────────────────────────────
print('\nBuilding match list...')
match_df = pd.DataFrame(all_match_rows).sort_values(['League','Model','Date'])

ws_ml = wb.create_sheet('Match List', 0)
ml_cols=['League','Model','Date','Home','Away','Bet Side','Odds','Result','Won','P&L','OOS']
for ci,col in enumerate(ml_cols,1):
    c=ws_ml.cell(1,ci,col); c.fill=fill_hdr
    c.font=Font(color='FFFFFF',bold=True); c.alignment=Alignment(horizontal='center')

for ri,(_,row) in enumerate(match_df.iterrows(),2):
    is_oos = row['OOS']=='OOS'
    for ci,col in enumerate(ml_cols,1):
        val=row[col]; c=ws_ml.cell(ri,ci,val)
        c.border=border; c.alignment=Alignment(horizontal='center')
        if ci==2: c.alignment=Alignment(horizontal='left')
        if ci==9:
            c.fill=fill_ps if val=='✓' else fill_neg
            c.font=Font(bold=True)
        if ci==10:
            c.number_format='+0.0;-0.0;0'
            c.fill=fill_p if row['P&L']>0 else fill_neg
        if ci==11:
            c.fill=fill_oos_h if is_oos else PatternFill()
            if is_oos: c.font=Font(color='FFFFFF',bold=True)

ml_widths=[16,42,12,20,20,9,7,8,5,8,5]
for i,w in enumerate(ml_widths,1):
    ws_ml.column_dimensions[get_column_letter(i)].width=w
ws_ml.freeze_panes='A2'
ws_ml.auto_filter.ref=f'A1:{get_column_letter(len(ml_cols))}1'

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws_sum = wb.create_sheet('SUMMARY', 0)
sum_cols=['League','Niche','Total n','ROI%','WR%','IS ROI%','IS WR%','OOS ROI%','OOS WR%','OOS n']
for ci,col in enumerate(sum_cols,1):
    c=ws_sum.cell(1,ci,col); c.fill=fill_hdr
    c.font=Font(color='FFFFFF',bold=True); c.alignment=Alignment(horizontal='center',wrap_text=True)

row_idx=2; prev_lg=None
all_stats=[]
for league,niches in MODELS.items():
    for niche_str in niches:
        side,lo,hi,xg,elo,fm,mk=parse_niche(niche_str)
        wc='home_win' if side=='home' else 'away_win'
        oc='home_odds_val' if side=='home' else 'away_odds_val'
        lg_key='Champions League' if league=='Champions League' else league
        bets=apply_mask(df,side,lo,hi,xg,elo,fm,mk,lg_key)
        if len(bets)==0: continue
        total_n=len(bets); total_w=int(bets[wc].sum())
        total_pl=(bets[wc]*(bets[oc]-1)-(1-bets[wc])).sum()*FLAT_STAKE
        ins=bets[bets['ym']<oos_p]; oos=bets[bets['ym']>=oos_p]
        ins_pl=(ins[wc]*(ins[oc]-1)-(1-ins[wc])).sum()*FLAT_STAKE if len(ins) else 0
        oos_pl=(oos[wc]*(oos[oc]-1)-(1-oos[wc])).sum()*FLAT_STAKE if len(oos) else 0
        all_stats.append({
            'League':league,'Niche':niche_str,'Total n':total_n,
            'ROI%':total_pl/(total_n*FLAT_STAKE)*100,
            'WR%':total_w/total_n*100,
            'IS ROI%':ins_pl/(len(ins)*FLAT_STAKE)*100 if len(ins) else 0,
            'IS WR%':ins[wc].mean()*100 if len(ins) else 0,
            'OOS ROI%':oos_pl/(len(oos)*FLAT_STAKE)*100 if len(oos) else 0,
            'OOS WR%':oos[wc].mean()*100 if len(oos) else 0,
            'OOS n':len(oos),
        })

all_stats.sort(key=lambda x:(x['League'],-x['OOS ROI%']))
prev_lg=None
for r in all_stats:
    if r['League']!=prev_lg:
        c=ws_sum.cell(row_idx,1,r['League']); c.fill=fill_lg
        c.font=Font(bold=True,size=11)
        ws_sum.merge_cells(start_row=row_idx,start_column=1,end_row=row_idx,end_column=len(sum_cols))
        row_idx+=1; prev_lg=r['League']
    oos_roi=r['OOS ROI%']; ins_roi=r['IS ROI%']
    vals=[r['League'],r['Niche'],r['Total n'],round(r['ROI%'],1),round(r['WR%'],1),
          round(ins_roi,1),round(r['IS WR%'],1),round(oos_roi,1),round(r['OOS WR%'],1),int(r['OOS n'])]
    for ci,val in enumerate(vals,1):
        c=ws_sum.cell(row_idx,ci,val); c.border=border
        c.alignment=Alignment(horizontal='center')
        if ci==2: c.alignment=Alignment(horizontal='left')
        if ci in (4,5): c.number_format='+0.0;-0.0;0'
        if ci==6:
            c.number_format='+0.0;-0.0;0'
            c.fill=fill_ps if ins_roi>=15 else (fill_p if ins_roi>0 else fill_neg)
        if ci==7: c.number_format='0.0'
        if ci==8:
            c.number_format='+0.0;-0.0;0'
            c.fill=fill_ps if oos_roi>=15 else (fill_p if oos_roi>0 else fill_neg)
            c.font=Font(bold=True)
        if ci==9: c.number_format='0.0'; c.font=Font(bold=True)
    row_idx+=1

ws_sum.column_dimensions['A'].width=16; ws_sum.column_dimensions['B'].width=44
for ci in range(3,len(sum_cols)+1): ws_sum.column_dimensions[get_column_letter(ci)].width=10
ws_sum.freeze_panes='A2'; ws_sum.auto_filter.ref=f'A1:{get_column_letter(len(sum_cols))}1'

wb.save(OUT)
print(f'\nTotal match events: {len(all_match_rows)}')
print(f'Saved: {OUT}')
print('✅ Done!')
