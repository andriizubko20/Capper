"""
BETA/expanding_wf.py

Expanding window cross-validation — повний sweep 45,000 фільтрів.

Логіка:
  Фолд 1: Train 2023-08→2024-04  |  Test 2024-05→2024-07
  Фолд 2: Train 2023-08→2024-07  |  Test 2024-08→2024-10
  Фолд 3: Train 2023-08→2024-10  |  Test 2024-11→2025-01
  Фолд 4: Train 2023-08→2025-01  |  Test 2025-02→2025-04
  Фолд 5: Train 2023-08→2025-04  |  Test 2025-05→2025-07
  Фолд 6: Train 2023-08→2025-07  |  Test 2025-08→2025-10
  ─────────────────────────────────────────────────────
  Final OOS: 2025-11 → 2026-04  (ніколи не торкається під час тренування)

Відбір ніші на кожному фолді:
  - ROI на тренувальних даних >= MIN_TRAIN_ROI
  - n >= MIN_TRAIN_N матчів у тренуванні

Стабільність:
  - Ніша вважається стабільною якщо вона прибуткова в >= MIN_STABLE_FOLDS з 6

Фінальний OOS тест: тільки стабільні ніші, n >= MIN_OOS_N.
"""
import sys, os, warnings
warnings.filterwarnings('ignore')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import numpy as np
import pandas as pd
from itertools import product
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as BdrSide
from openpyxl.utils import get_column_letter

from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix

# ── Config ────────────────────────────────────────────────────────────────────
MIN_TRAIN_ROI   = 10.0   # мінімальний ROI на тренувальних даних
MIN_TRAIN_N     = 10     # мінімум матчів у тренуванні
MIN_STABLE_FOLDS = 3     # мінімум прибуткових тестових фолдів з 6
MIN_OOS_N       = 10     # мінімум матчів у фінальному OOS

# ── Folds ─────────────────────────────────────────────────────────────────────
FOLDS = [
    ('Fold1', pd.Timestamp('2023-08-01'), pd.Timestamp('2024-04-30'),
              pd.Timestamp('2024-05-01'), pd.Timestamp('2024-07-31')),
    ('Fold2', pd.Timestamp('2023-08-01'), pd.Timestamp('2024-07-31'),
              pd.Timestamp('2024-08-01'), pd.Timestamp('2024-10-31')),
    ('Fold3', pd.Timestamp('2023-08-01'), pd.Timestamp('2024-10-31'),
              pd.Timestamp('2024-11-01'), pd.Timestamp('2025-01-31')),
    ('Fold4', pd.Timestamp('2023-08-01'), pd.Timestamp('2025-01-31'),
              pd.Timestamp('2025-02-01'), pd.Timestamp('2025-04-30')),
    ('Fold5', pd.Timestamp('2023-08-01'), pd.Timestamp('2025-04-30'),
              pd.Timestamp('2025-05-01'), pd.Timestamp('2025-07-31')),
    ('Fold6', pd.Timestamp('2023-08-01'), pd.Timestamp('2025-07-31'),
              pd.Timestamp('2025-08-01'), pd.Timestamp('2025-10-31')),
]
FOLD_LABELS = [f[0] for f in FOLDS]

FINAL_OOS_START = pd.Timestamp('2025-11-01')

# ── Load ──────────────────────────────────────────────────────────────────────
print("Loading data...")
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result'] == 'H').astype(int)
df['away_win']  = (df['result'] == 'A').astype(int)
print(f"Total: {len(df)} matches ({df['date'].min().date()} – {df['date'].max().date()})")

oos_df = df[df['date'] >= FINAL_OOS_START].copy()
print(f"Final OOS: {len(oos_df)} matches ({oos_df['date'].min().date()} – {oos_df['date'].max().date()})")

LEAGUES = [
    'Premier League','Bundesliga','Serie A','La Liga','Ligue 1',
    'Primeira Liga','Serie B','Eredivisie','Jupiler Pro League','Champions League'
]
LSHORT = {
    'Premier League':'EPL','Bundesliga':'Bundesliga','Serie A':'Serie A',
    'La Liga':'La Liga','Ligue 1':'Ligue 1','Primeira Liga':'Portugal',
    'Serie B':'Serie B','Eredivisie':'Eredivisie',
    'Jupiler Pro League':'Jupiler','Champions League':'UCL'
}

# Pre-slice folds
fold_slices = {}
for fname, tr_s, tr_e, te_s, te_e in FOLDS:
    fold_slices[fname] = {
        'train': df[(df['date'] >= tr_s) & (df['date'] <= tr_e)].copy(),
        'test':  df[(df['date'] >= te_s) & (df['date'] <= te_e)].copy(),
        'te_s': te_s, 'te_e': te_e,
    }

# ── Filter space ──────────────────────────────────────────────────────────────
ODDS_RANGES = [(1.30,1.55),(1.55,1.80),(1.70,2.00),(1.80,2.20),(2.00,2.50),(2.20,2.80),(2.50,3.50)]
XG_THRS    = [0.0,1.0,1.2,1.5,1.8]
ELO_H_THRS = [0,30,75,150]
ELO_A_THRS = [0,-30,-75,-150]
FORM_THRS  = [0.0,1.5,1.8,2.2]
MKT_H_THRS = [0.0,0.45,0.50,0.55]
MKT_A_THRS = [0.0,0.35,0.40,0.45]

def apply_mask(d, side, lo, hi, xg_t, elo_t, form_t, mkt_t):
    oc  = 'home_odds_val'   if side == 'home' else 'away_odds_val'
    xgc = 'xg_ratio_home_5' if side == 'home' else 'xg_ratio_away_5'
    fc  = 'home_pts_5'      if side == 'home' else 'away_pts_5'
    mc  = 'mkt_home_prob'   if side == 'home' else 'mkt_away_prob'
    m = (d[oc] >= lo) & (d[oc] < hi)
    if xg_t > 0:                        m &= (d[xgc].fillna(0) >= xg_t)
    if side == 'home' and elo_t > 0:    m &= (d['elo_diff'].fillna(0) >= elo_t)
    if side == 'away' and elo_t < 0:    m &= (d['elo_diff'].fillna(0) <= elo_t)
    if form_t > 0:                      m &= (d[fc].fillna(0) >= form_t)
    if mkt_t > 0:                       m &= (d[mc].fillna(0) >= mkt_t)
    return d[m]

# ── SWEEP ─────────────────────────────────────────────────────────────────────
# results[(lg, label)] = {
#   'meta': {...},
#   folds: {FoldN: {'train_roi', 'train_n', 'test_roi', 'test_n', 'test_wr', 'selected'}}
# }
print("\nRunning sweep across 6 folds...")
results = defaultdict(lambda: {'folds': {}})

total_combos = len(ODDS_RANGES) * 2 * len(XG_THRS) * len(ELO_H_THRS) * len(FORM_THRS) * len(MKT_H_THRS)
print(f"Filter combinations per side: {len(ODDS_RANGES)*len(XG_THRS)*len(ELO_H_THRS)*len(FORM_THRS)*len(MKT_H_THRS)}")

for side in ('home', 'away'):
    et = ELO_H_THRS if side == 'home' else ELO_A_THRS
    mt = MKT_H_THRS if side == 'home' else MKT_A_THRS
    wc = 'home_win'      if side == 'home' else 'away_win'
    oc = 'home_odds_val' if side == 'home' else 'away_odds_val'
    print(f"  Side: {side}")

    for lo, hi in ODDS_RANGES:
        # Pre-filter OOS by odds only (optimization)
        oos_pre = apply_mask(oos_df, side, lo, hi, 0, 0, 0, 0)

        for xg_t, elo_t, form_t, mkt_t in product(XG_THRS, et, FORM_THRS, mt):
            parts = [f'{side}[{lo},{hi})']
            if xg_t > 0:                     parts.append(f'xg>={xg_t}')
            if side == 'home' and elo_t > 0: parts.append(f'elo>={elo_t}')
            if side == 'away' and elo_t < 0: parts.append(f'elo<={elo_t}')
            if form_t > 0:                   parts.append(f'form>={form_t}')
            if mkt_t > 0:                    parts.append(f'mkt>={mkt_t}')
            label = ' '.join(parts)

            for lg in LEAGUES:
                key = (lg, label)

                for fname, _, _, _, _ in FOLDS:
                    fs = fold_slices[fname]
                    # Training
                    tr = apply_mask(fs['train'], side, lo, hi, xg_t, elo_t, form_t, mkt_t)
                    tr = tr[tr['league_name'] == lg]
                    if len(tr) < MIN_TRAIN_N: continue
                    train_roi = (tr[wc]*(tr[oc]-1)-(1-tr[wc])).mean()*100
                    if train_roi < MIN_TRAIN_ROI: continue

                    # Test
                    te = apply_mask(fs['test'], side, lo, hi, xg_t, elo_t, form_t, mkt_t)
                    te = te[te['league_name'] == lg]
                    if len(te) < 3: continue
                    test_roi = (te[wc]*(te[oc]-1)-(1-te[wc])).mean()*100
                    test_wr  = te[wc].mean()*100

                    if key not in results or 'meta' not in results[key]:
                        results[key]['meta'] = {
                            'label': label, 'side': side, 'lo': lo, 'hi': hi,
                            'xg_t': xg_t, 'elo_t': elo_t, 'form_t': form_t, 'mkt_t': mkt_t,
                            'lg': lg, 'lshort': LSHORT[lg],
                        }
                    results[key]['folds'][fname] = {
                        'train_roi': round(train_roi, 1),
                        'train_n':   len(tr),
                        'test_roi':  round(test_roi, 1),
                        'test_n':    len(te),
                        'test_wr':   round(test_wr, 1),
                    }

                # OOS (always compute if any fold was selected)
                if key in results and results[key]['folds']:
                    sl_oos = apply_mask(oos_pre, side, lo, hi, xg_t, elo_t, form_t, mkt_t)
                    sl_oos = sl_oos[sl_oos['league_name'] == lg]
                    if len(sl_oos) >= MIN_OOS_N:
                        oos_roi = (sl_oos[wc]*(sl_oos[oc]-1)-(1-sl_oos[wc])).mean()*100
                        results[key]['oos'] = {
                            'n': len(sl_oos),
                            'roi': round(oos_roi, 1),
                            'wr':  round(sl_oos[wc].mean()*100, 1),
                        }

print(f"Total filter×league combinations found: {len(results)}")

# ── Stability analysis ────────────────────────────────────────────────────────
stable, unstable = [], []

for key, data in results.items():
    if 'meta' not in data: continue
    folds = data['folds']
    if not folds: continue

    n_selected   = len(folds)
    n_profitable = sum(1 for f in folds.values() if f['test_roi'] > 0)
    avg_test_roi = np.mean([f['test_roi'] for f in folds.values()])
    avg_train_roi= np.mean([f['train_roi'] for f in folds.values()])
    avg_test_wr  = np.mean([f['test_wr'] for f in folds.values()])

    entry = {
        **data['meta'],
        'n_selected':   n_selected,
        'n_profitable': n_profitable,
        'stability':    round(n_profitable / n_selected * 100, 0),
        'avg_train_roi':round(avg_train_roi, 1),
        'avg_test_roi': round(avg_test_roi, 1),
        'avg_test_wr':  round(avg_test_wr, 1),
        'folds':        folds,
        'oos':          data.get('oos'),
    }

    if n_profitable >= MIN_STABLE_FOLDS:
        stable.append(entry)
    else:
        unstable.append(entry)

stable.sort(key=lambda x: (-x['n_profitable'], -x['avg_test_roi']))
print(f"\nStable niches (profitable in >={MIN_STABLE_FOLDS}/6 folds): {len(stable)}")
print(f"Unstable: {len(unstable)}")

# Per-league summary
by_lg = defaultdict(list)
for s in stable: by_lg[s['lshort']].append(s)

print("\nPer-league (stable):")
for lg, items in sorted(by_lg.items(), key=lambda x: -len(x[1])):
    oos_items = [i for i in items if i['oos'] and i['oos']['roi'] > 0]
    print(f"  {lg:<12} {len(items):>3} stable  |  "
          f"avg test ROI={np.mean([i['avg_test_roi'] for i in items]):+.1f}%  |  "
          f"OOS confirmed: {len(oos_items)}/{len([i for i in items if i['oos']])}")

# ── Dedup ─────────────────────────────────────────────────────────────────────
def is_subset(a, b, side):
    if b['xg_t'] < a['xg_t'] or b['form_t'] < a['form_t'] or b['mkt_t'] < a['mkt_t']:
        return False
    return b['elo_t'] >= a['elo_t'] if side == 'home' else b['elo_t'] <= a['elo_t']

def dedup(filters):
    filters = sorted(filters, key=lambda x: (-x['n_profitable'], -x['avg_test_roi']))
    kept = []
    for f in filters:
        if not any(is_subset(k, f, f['side']) and k['n_profitable'] >= f['n_profitable'] for k in kept):
            kept.append(f)
    return kept

catalog = {}
for lg in LEAGUES:
    lshort = LSHORT[lg]
    lg_items = [s for s in stable if s['lg'] == lg]
    groups = defaultdict(list)
    for f in lg_items:
        groups[(f['side'], f['lo'], f['hi'])].append(f)
    catalog[lshort] = []
    for group in groups.values():
        catalog[lshort].extend(dedup(group))

total_catalog = sum(len(v) for v in catalog.values())
print(f"\nAfter dedup: {total_catalog} niches")
for lshort, items in sorted(catalog.items(), key=lambda x: -len(x[1])):
    if items:
        oos_pos = [i for i in items if i['oos'] and i['oos']['roi'] > 0]
        print(f"  {lshort:<12} {len(items):>2} ніш  |  OOS+: {len(oos_pos)}")

# ── Excel ─────────────────────────────────────────────────────────────────────
print("\nBuilding Excel...")

GREEN_L = PatternFill('solid', start_color='D5F5E3')
RED_L   = PatternFill('solid', start_color='FADBD8')
GREY    = PatternFill('solid', start_color='F2F3F4')
DARK    = PatternFill('solid', start_color='1B2631')
OOS_BG  = PatternFill('solid', start_color='154360')
SUBHDR  = PatternFill('solid', start_color='5D6D7E')
YELLOW  = PatternFill('solid', start_color='FEF9E7')

HDR_F  = Font(name='Arial', bold=True, color='FFFFFF', size=9)
BODY_F = Font(name='Arial', size=9)
C = Alignment(horizontal='center', vertical='center')
L = Alignment(horizontal='left',   vertical='center')

def tb():
    s = BdrSide(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def hc(c, v, fill=DARK):
    c.value=v; c.font=HDR_F; c.fill=fill; c.alignment=C; c.border=tb()

def bc(c, v, align=C, fill=None, bold=False):
    c.value=v; c.font=Font(name='Arial', size=9, bold=bold)
    c.alignment=align; c.border=tb()
    if fill: c.fill=fill

wb = Workbook(); wb.remove(wb.active)

# ── Sheet: SUMMARY ────────────────────────────────────────────────────────────
ws_s = wb.create_sheet('SUMMARY')
SHDR = ['Ліга','Ніш (dedup)','Avg Train ROI%','Avg Test ROI%','Stability%',
        'OOS підтверджено','OOS n (avg)','Найкраща ніша']
for ci, h in enumerate(SHDR, 1): hc(ws_s.cell(1, ci), h)
sr = 2
for lshort, items in sorted(catalog.items(), key=lambda x: -len(x[1])):
    if not items: continue
    oos_pos = [i for i in items if i['oos'] and i['oos']['roi'] > 0]
    oos_all = [i for i in items if i['oos']]
    best = max(items, key=lambda x: x['avg_test_roi'])
    row = [
        lshort, len(items),
        round(np.mean([i['avg_train_roi'] for i in items]), 1),
        round(np.mean([i['avg_test_roi']  for i in items]), 1),
        round(np.mean([i['stability']     for i in items]), 0),
        f"{len(oos_pos)}/{len(oos_all)}",
        round(np.mean([i['oos']['n'] for i in oos_all]), 0) if oos_all else '—',
        best['label'],
    ]
    fmts=[None,None,'+0.0%','+0.0%','0%',None,None,None]
    for ci,(v,fmt) in enumerate(zip(row,fmts),1):
        c=ws_s.cell(sr,ci)
        bc(c, v/100 if fmt and isinstance(v,(int,float)) else v, L if ci in(1,8) else C)
        if fmt and isinstance(v,(int,float)): c.number_format=fmt
        if ci==3: c.fill=GREEN_L if row[2]>0 else RED_L
        if ci==4: c.fill=GREEN_L if row[3]>0 else RED_L
    sr+=1

ws_s.column_dimensions['A'].width=12
ws_s.column_dimensions['B'].width=12
for ci in range(3,8): ws_s.column_dimensions[get_column_letter(ci)].width=14
ws_s.column_dimensions['H'].width=55
ws_s.freeze_panes='A2'

# ── Per-league sheets ─────────────────────────────────────────────────────────
# Columns:
# Ніша | Side | Stability | Profitable/Selected | Avg Train ROI | Avg Test ROI | Avg Test WR
# | OOS n | OOS WR% | OOS ROI%
# | Fold1 train | Fold1 test | ... × 6

STATIC = ['Ніша','Side','Stability%','Prof/Sel',
          'Avg Train ROI%','Avg Test ROI%','Avg Test WR%',
          'OOS n','OOS WR%','OOS ROI%']
N_S = len(STATIC)

for lshort, items in catalog.items():
    if not items: continue
    ws = wb.create_sheet(lshort)
    total_c = N_S + len(FOLDS)*2

    # Row 1: static + fold group headers
    for ci, h in enumerate(STATIC, 1): hc(ws.cell(1, ci), h)
    for fi, (fname,_,_,ts,te) in enumerate(FOLDS):
        cs = N_S + fi*2 + 1
        ws.merge_cells(start_row=1,start_column=cs,end_row=1,end_column=cs+1)
        hc(ws.cell(1,cs), f'{fname}\n{ts.strftime("%b%y")}→{te.strftime("%b%y")}')

    # Row 2: sub-headers
    for ci in range(1, N_S+1): hc(ws.cell(2,ci),'',fill=SUBHDR)
    for fi in range(len(FOLDS)):
        cs=N_S+fi*2+1
        hc(ws.cell(2,cs),  'Tr ROI%',fill=SUBHDR)
        hc(ws.cell(2,cs+1),'Te ROI%',fill=SUBHDR)

    items_s = sorted(items, key=lambda x:(-x['n_profitable'],-x['avg_test_roi']))
    for er, niche in enumerate(items_s, 3):
        oos = niche.get('oos')
        row_fill = (GREEN_L if niche['avg_test_roi']>0 and niche['n_profitable']>=4 else
                    YELLOW  if niche['avg_test_roi']>0 else RED_L)

        static_vals = [
            niche['label'], niche['side'],
            niche['stability']/100,
            f"{niche['n_profitable']}/{niche['n_selected']}",
            niche['avg_train_roi']/100,
            niche['avg_test_roi']/100,
            niche['avg_test_wr']/100,
            oos['n']   if oos else '—',
            oos['wr']/100  if oos else '—',
            oos['roi']/100 if oos else '—',
        ]
        sfmts=[None,None,'0%',None,'+0.0%;-0.0%','+0.0%;-0.0%','0.0%',
               None,'0.0%','+0.0%;-0.0%']
        for ci,(v,fmt) in enumerate(zip(static_vals,sfmts),1):
            c=ws.cell(er,ci)
            bc(c,v,L if ci in(1,2,4) else C, row_fill)
            if fmt and isinstance(v,(int,float)): c.number_format=fmt
            if ci==3: c.fill=GREEN_L if niche['stability']>=50 else RED_L
            if ci==6:
                c.fill=GREEN_L if niche['avg_test_roi']>0 else RED_L
                c.font=Font(name='Arial',size=9,bold=abs(niche['avg_test_roi'])>10)
            if ci==10 and oos:
                c.fill=GREEN_L if oos['roi']>0 else RED_L
                c.font=Font(name='Arial',size=9,bold=True)

        # Fold cells
        for fi, (fname,_,_,_,_) in enumerate(FOLDS):
            cs = N_S+fi*2+1
            fd = niche['folds'].get(fname)
            if fd:
                c_tr = ws.cell(er,cs)
                c_tr.value=fd['train_roi']/100; c_tr.number_format='+0.0%;-0.0%'
                c_tr.font=BODY_F; c_tr.alignment=C; c_tr.border=tb()
                c_tr.fill=GREEN_L if fd['train_roi']>0 else RED_L

                c_te=ws.cell(er,cs+1)
                c_te.value=fd['test_roi']/100; c_te.number_format='+0.0%;-0.0%'
                c_te.font=Font(name='Arial',size=9,bold=abs(fd['test_roi'])>15)
                c_te.alignment=C; c_te.border=tb()
                c_te.fill=GREEN_L if fd['test_roi']>0 else RED_L
            else:
                bc(ws.cell(er,cs),  '—',fill=GREY)
                bc(ws.cell(er,cs+1),'—',fill=GREY)

    ws.column_dimensions['A'].width=52
    ws.column_dimensions['B'].width=6
    for ci in range(3,N_S+1): ws.column_dimensions[get_column_letter(ci)].width=11
    for fi in range(len(FOLDS)):
        cs=N_S+fi*2+1
        ws.column_dimensions[get_column_letter(cs)].width=8
        ws.column_dimensions[get_column_letter(cs+1)].width=8
    ws.freeze_panes='A3'
    ws.row_dimensions[1].height=28; ws.row_dimensions[2].height=18

outpath = os.path.join(os.path.dirname(__file__), 'expanding_wf_results.xlsx')
wb.save(outpath)
print(f"\nSaved: {outpath}  ({os.path.getsize(outpath)//1024}KB)")
print("\n✅ Done!")
