"""
BETA/build_niche_catalog.py

Будує нішевий каталог — армія міні-моделей per ліга.

Алгоритм:
  1. Збираємо всі стабільні фільтри з walk-forward
     (win% >= 60%, avg_roi > 0, n_tot >= 10, n_windows >= 4)
  2. Групуємо по: ліга × сторона × odds range
  3. В кожній групі видаляємо підмножини — якщо фільтр B ⊂ A,
     залишаємо тільки кращий по avg_roi
  4. Незалежні фільтри (різні ознаки) — лишаємо обидва
  5. Підтверджуємо кожну нішу по оригінальному OOS (2025-11→2026-04)
  6. Пакуємо в Excel: каталог + portfolio summary
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import numpy as np
import pandas as pd
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as BdrSide
from openpyxl.utils import get_column_letter

from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix

# ── Load ──────────────────────────────────────────────────────────────────────
print("Loading data...")
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result'] == 'H').astype(int)
df['away_win']  = (df['result'] == 'A').astype(int)

DISC_END  = pd.Timestamp('2025-10-31')
OOS_START = pd.Timestamp('2025-11-01')

disc = df[df['date'] <= DISC_END].copy()
oos  = df[df['date'] >= OOS_START].copy()
print(f"Discovery: {len(disc)}  OOS: {len(oos)}")

LEAGUES = [
    'Premier League','Bundesliga','Serie A','La Liga','Ligue 1',
    'Primeira Liga','Serie B','Eredivisie','Jupiler Pro League','Champions League'
]
LEAGUE_SHORT = {
    'Premier League':'EPL','Bundesliga':'Bundesliga','Serie A':'Serie A',
    'La Liga':'La Liga','Ligue 1':'Ligue 1','Primeira Liga':'Portugal',
    'Serie B':'Serie B','Eredivisie':'Eredivisie',
    'Jupiler Pro League':'Jupiler','Champions League':'UCL'
}

# ── Walk-forward windows (повний датасет) ─────────────────────────────────────
WF_WINDOWS = []
cur = pd.Timestamp('2023-08-01')
end = pd.Timestamp('2026-04-30')
while cur < end:
    w_end = cur + pd.DateOffset(months=3) - pd.Timedelta(days=1)
    if w_end > end: w_end = end
    WF_WINDOWS.append((cur.strftime('%Y-%m'), cur, w_end))
    cur += pd.DateOffset(months=3)
WIN_LABELS = [w[0] for w in WF_WINDOWS]

wf_slices = {wlbl: df[(df['date']>=ws)&(df['date']<=we)]
             for wlbl,ws,we in WF_WINDOWS}

# ── Filter space ──────────────────────────────────────────────────────────────
ODDS_RANGES = [
    (1.30,1.55),(1.55,1.80),(1.70,2.00),
    (1.80,2.20),(2.00,2.50),(2.20,2.80),(2.50,3.50),
]
XG_THRS    = [0.0,1.0,1.2,1.5,1.8]
ELO_H_THRS = [0,30,75,150]
ELO_A_THRS = [0,-30,-75,-150]
FORM_THRS  = [0.0,1.5,1.8,2.2]
MKT_H_THRS = [0.0,0.45,0.50,0.55]
MKT_A_THRS = [0.0,0.35,0.40,0.45]
MIN_N      = 3

def calc(sub, side):
    wc = 'home_win' if side=='home' else 'away_win'
    oc = 'home_odds_val' if side=='home' else 'away_odds_val'
    if len(sub) < MIN_N: return None
    wr    = sub[wc].mean()
    avg_o = sub[oc].mean()
    roi   = (sub[wc]*(sub[oc]-1)-(1-sub[wc])).mean()*100
    ev    = (wr*avg_o-1)*100
    return {'n':len(sub),'wr':round(wr*100,1),'roi':round(roi,1),
            'ev':round(ev,1),'avg_odds':round(avg_o,2)}

def apply_mask(d, side, lo, hi, xg_t, elo_t, form_t, mkt_t):
    oc  = 'home_odds_val'   if side=='home' else 'away_odds_val'
    xgc = 'xg_ratio_home_5' if side=='home' else 'xg_ratio_away_5'
    fc  = 'home_pts_5'      if side=='home' else 'away_pts_5'
    mc  = 'mkt_home_prob'   if side=='home' else 'mkt_away_prob'
    m = (d[oc]>=lo)&(d[oc]<hi)
    if xg_t  > 0: m &= (d[xgc].fillna(0)>=xg_t)
    if side=='home' and elo_t>0: m &= (d['elo_diff'].fillna(0)>=elo_t)
    if side=='away' and elo_t<0: m &= (d['elo_diff'].fillna(0)<=elo_t)
    if form_t > 0: m &= (d[fc].fillna(0)>=form_t)
    if mkt_t  > 0: m &= (d[mc].fillna(0)>=mkt_t)
    return d[m]

# ── SWEEP ─────────────────────────────────────────────────────────────────────
print("Running sweep...")
# wf_results[lg][label] = {meta, windows:{wlbl:metrics}}
wf_results = {lg: {} for lg in LEAGUES}

for side in ('home','away'):
    et = ELO_H_THRS if side=='home' else ELO_A_THRS
    mt = MKT_H_THRS if side=='home' else MKT_A_THRS
    for lo,hi in ODDS_RANGES:
        for xg_t,elo_t,form_t,mkt_t in product(XG_THRS,et,FORM_THRS,mt):
            parts = [f'{side}[{lo},{hi})']
            if xg_t  > 0: parts.append(f'xg>={xg_t}')
            if side=='home' and elo_t>0: parts.append(f'elo>={elo_t}')
            if side=='away' and elo_t<0: parts.append(f'elo<={elo_t}')
            if form_t > 0: parts.append(f'form>={form_t}')
            if mkt_t  > 0: parts.append(f'mkt>={mkt_t}')
            label = ' '.join(parts)

            for wlbl,_,_ in WF_WINDOWS:
                sub = apply_mask(wf_slices[wlbl], side, lo, hi, xg_t, elo_t, form_t, mkt_t)
                for lg in LEAGUES:
                    sub_l = sub[sub['league_name']==lg]
                    m = calc(sub_l, side)
                    if not m: continue
                    if label not in wf_results[lg]:
                        wf_results[lg][label] = {
                            'side':side,'lo':lo,'hi':hi,
                            'xg_t':xg_t,'elo_t':elo_t,
                            'form_t':form_t,'mkt_t':mkt_t,
                            'windows':{}
                        }
                    wf_results[lg][label]['windows'][wlbl] = m

print("Sweep done.")

# ── STEP 1: Collect stable filters ───────────────────────────────────────────
print("Collecting stable filters...")
stable = {lg: [] for lg in LEAGUES}

for lg in LEAGUES:
    for label, data in wf_results[lg].items():
        wins = data['windows']
        if not wins: continue
        n_tot    = sum(w['n']   for w in wins.values())
        avg_roi  = np.mean([w['roi'] for w in wins.values()])
        avg_wr   = np.mean([w['wr']  for w in wins.values()])
        avg_ev   = np.mean([w['ev']  for w in wins.values()])
        n_pos    = sum(1 for w in wins.values() if w['roi']>0)
        n_w      = len(wins)
        win_pct  = n_pos/n_w*100

        if n_tot < 10 or n_w < 4: continue
        if avg_roi < 10.0 or win_pct < 60: continue

        stable[lg].append({
            'label':label, 'side':data['side'],
            'lo':data['lo'], 'hi':data['hi'],
            'xg_t':data['xg_t'], 'elo_t':data['elo_t'],
            'form_t':data['form_t'], 'mkt_t':data['mkt_t'],
            'n_tot':n_tot, 'avg_roi':round(avg_roi,1),
            'avg_wr':round(avg_wr,1), 'avg_ev':round(avg_ev,1),
            'win_pct':round(win_pct,1),
            'n_pos':n_pos, 'n_windows':n_w,
        })

total_stable = sum(len(v) for v in stable.values())
print(f"Stable filters before dedup: {total_stable}")

# ── STEP 2: Subset detection & deduplication ─────────────────────────────────
def is_subset(a, b, side):
    """Returns True if b is a strict subset of a (b selects fewer matches)."""
    if a['xg_t']   > b['xg_t']:   return False
    if a['form_t'] > b['form_t']: return False
    if a['mkt_t']  > b['mkt_t']:  return False
    if side == 'home':
        if a['elo_t'] > b['elo_t']: return False
    else:
        if a['elo_t'] < b['elo_t']: return False  # away elo is negative
    return True  # b is at least as strict as a on all dims

def dedup_group(filters):
    """
    Given a list of filters in same (side, odds_range),
    remove those that are subsets of a better-performing filter.
    Keep all truly independent filters.
    """
    # Sort by avg_roi desc
    filters = sorted(filters, key=lambda x: x['avg_roi'], reverse=True)
    kept = []
    for f in filters:
        dominated = False
        for k in kept:
            # f is subset of k AND k is better → skip f
            if is_subset(k, f, f['side']) and k['avg_roi'] >= f['avg_roi']:
                dominated = True
                break
        if not dominated:
            kept.append(f)
    return kept

print("Deduplicating...")
catalog = {lg: [] for lg in LEAGUES}

for lg in LEAGUES:
    # Group by (side, lo, hi)
    groups = {}
    for f in stable[lg]:
        key = (f['side'], f['lo'], f['hi'])
        groups.setdefault(key, []).append(f)

    for key, group in groups.items():
        deduped = dedup_group(group)
        catalog[lg].extend(deduped)

    # Sort final catalog by: side, lo, hi, avg_roi desc
    catalog[lg].sort(key=lambda x: (x['side'], x['lo'], x['hi'], -x['avg_roi']))

total_catalog = sum(len(v) for v in catalog.values())
print(f"After dedup: {total_catalog} niches in catalog")

# ── STEP 3: Confirm each niche on original OOS (2025-11 → 2026-04) ───────────
print("Confirming on original OOS...")
for lg in LEAGUES:
    sub_lg_oos = oos[oos['league_name']==lg]
    for f in catalog[lg]:
        sub_o = apply_mask(sub_lg_oos, f['side'], f['lo'], f['hi'],
                           f['xg_t'], f['elo_t'], f['form_t'], f['mkt_t'])
        m = calc(sub_o, f['side'])
        f['oos'] = m  # None if < MIN_N

        # Discovery stats
        sub_lg_disc = disc[disc['league_name']==lg]
        sub_d = apply_mask(sub_lg_disc, f['side'], f['lo'], f['hi'],
                           f['xg_t'], f['elo_t'], f['form_t'], f['mkt_t'])
        md = calc(sub_d, f['side'])
        f['disc'] = md

        # OOS confirmation status
        if m is None:
            f['oos_status'] = '❓ замало даних'
        elif m['roi'] > 0:
            f['oos_status'] = '✅ підтверджено'
        else:
            f['oos_status'] = '❌ не підтверджено'

# ── STEP 4: Build Excel ───────────────────────────────────────────────────────
print("Building Excel...")

GREEN      = PatternFill('solid', start_color='D5F5E3')
RED        = PatternFill('solid', start_color='FADBD8')
YELLOW     = PatternFill('solid', start_color='FEF9E7')
GREY       = PatternFill('solid', start_color='EAECEE')
DARK_HDR   = PatternFill('solid', start_color='1B2631')
GRP_HOME   = PatternFill('solid', start_color='154360')
GRP_AWAY   = PatternFill('solid', start_color='1A3A2A')
ODDS_HDR   = PatternFill('solid', start_color='2C3E50')

HDR_FONT   = Font(name='Arial', bold=True, color='FFFFFF', size=9)
BODY_FONT  = Font(name='Arial', size=9)
BOLD_FONT  = Font(name='Arial', bold=True, size=9)
CENTER     = Alignment(horizontal='center', vertical='center')
LEFT       = Alignment(horizontal='left',   vertical='center')

def tb():
    s = BdrSide(style='thin', color='CCCCCC')
    return Border(left=s,right=s,top=s,bottom=s)

def hc(c, v, fill=DARK_HDR, align=CENTER):
    c.value=v; c.font=HDR_FONT; c.fill=fill
    c.alignment=align; c.border=tb()

def bc(c, v, align=CENTER, fill=None, bold=False, fmt=None):
    c.value=v
    c.font = Font(name='Arial', size=9, bold=bold)
    c.alignment=align; c.border=tb()
    if fill: c.fill=fill
    if fmt:  c.number_format=fmt

def roi_c(c, roi_val, bold=False):
    if roi_val is None:
        bc(c,'—'); c.fill=GREY; return
    c.value = roi_val/100
    c.font  = Font(name='Arial', size=9, bold=(bold or abs(roi_val)>20))
    c.number_format = '+0.0%;-0.0%;0.0%'
    c.alignment = CENTER; c.border = tb()
    c.fill = GREEN if roi_val>0 else RED

wb = Workbook()
wb.remove(wb.active)

# ════════════════════════════════════════════════════════════
# Sheet 1: PORTFOLIO SUMMARY
# ════════════════════════════════════════════════════════════
ws_p = wb.create_sheet('PORTFOLIO SUMMARY')

P_COLS = ['Ліга','Всього ніш','Home ніш','Away ніш',
          'Підтверджено OOS','Не підтверджено','Замало даних',
          'Avg ROI% (WF)','Best нішa','Best ROI%']
for ci,h in enumerate(P_COLS,1): hc(ws_p.cell(1,ci), h)

pr = 2
for lg in LEAGUES:
    lname = LEAGUE_SHORT[lg]
    niches = catalog[lg]
    if not niches:
        continue
    n_home   = sum(1 for f in niches if f['side']=='home')
    n_away   = sum(1 for f in niches if f['side']=='away')
    n_conf   = sum(1 for f in niches if f['oos_status']=='✅ підтверджено')
    n_fail   = sum(1 for f in niches if f['oos_status']=='❌ не підтверджено')
    n_unk    = sum(1 for f in niches if f['oos_status']=='❓ замало даних')
    avg_roi  = np.mean([f['avg_roi'] for f in niches])
    best     = max(niches, key=lambda x: x['avg_roi'])

    row = [lname, len(niches), n_home, n_away,
           n_conf, n_fail, n_unk,
           round(avg_roi,1)/100,
           best['label'], best['avg_roi']/100]
    fmts = [None]*7 + ['+0.0%;-0.0%',None,'+0.0%;-0.0%']
    fills = [None]*7 + [GREEN if avg_roi>0 else RED, None,
                        GREEN if best['avg_roi']>0 else RED]
    for ci,(val,fmt,fill) in enumerate(zip(row,fmts,fills),1):
        c = ws_p.cell(pr,ci)
        bc(c, val, LEFT if ci in (1,9) else CENTER, fill, fmt=fmt)
    pr += 1

ws_p.column_dimensions['A'].width = 12
ws_p.column_dimensions['I'].width = 55
ws_p.column_dimensions['J'].width = 12
for ci in range(2,9):
    ws_p.column_dimensions[get_column_letter(ci)].width = 13
ws_p.freeze_panes = 'A2'

# ════════════════════════════════════════════════════════════
# Sheet 2: NICHE CATALOG (всі ніші)
# ════════════════════════════════════════════════════════════
ws_c = wb.create_sheet('NICHE CATALOG')

CAT_COLS = [
    'Ліга','Side','Odds Range','Filter',
    'WF n','WF Avg_ROI%','WF Avg_WR%','WF Avg_EV%','WF Win%','WF W+/Wtot',
    'DISC n','DISC ROI%',
    'OOS n','OOS WR%','OOS ROI%','OOS EV%',
    'OOS статус'
]
for ci,h in enumerate(CAT_COLS,1): hc(ws_c.cell(1,ci), h)

cr = 2
prev_lg = None
for lg in LEAGUES:
    lname = LEAGUE_SHORT[lg]
    niches = catalog[lg]
    if not niches: continue

    # League separator
    ws_c.merge_cells(start_row=cr, start_column=1,
                     end_row=cr,   end_column=len(CAT_COLS))
    c = ws_c.cell(cr,1)
    c.value = f'▶ {lname} ({lg})  —  {len(niches)} ніш'
    c.font  = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    c.fill  = DARK_HDR; c.alignment=LEFT; c.border=tb()
    cr += 1

    prev_key = None
    for f in niches:
        key = (f['side'], f['lo'], f['hi'])

        # Odds range sub-header
        if key != prev_key:
            odds_label = f"{f['side'].upper()}  [{f['lo']},{f['hi']})"
            fill = GRP_HOME if f['side']=='home' else GRP_AWAY
            ws_c.merge_cells(start_row=cr, start_column=1,
                             end_row=cr,   end_column=len(CAT_COLS))
            c = ws_c.cell(cr,1)
            c.value = odds_label
            c.font  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            c.fill  = fill; c.alignment=LEFT; c.border=tb()
            cr += 1
            prev_key = key

        mo = f.get('oos')
        md = f.get('disc')
        oos_stat = f['oos_status']
        row_fill = (GREEN  if oos_stat=='✅ підтверджено' else
                    YELLOW if oos_stat=='❓ замало даних'  else RED)

        vals = [
            lname, f['side'],
            f"[{f['lo']},{f['hi']})",
            f['label'],
            f['n_tot'], f['avg_roi']/100, f['avg_wr']/100,
            f['avg_ev']/100, f['win_pct']/100,
            f"{f['n_pos']}/{f['n_windows']}",
            md['n']   if md else '—',
            (md['roi']/100) if md else '—',
            mo['n']   if mo else '—',
            (mo['wr']/100)  if mo else '—',
            (mo['roi']/100) if mo else '—',
            (mo['ev']/100)  if mo else '—',
            oos_stat,
        ]
        fmts = [None,None,None,None,
                None,'+0.0%;-0.0%','0.0%','+0.0%;-0.0%','0%',None,
                None,'+0.0%;-0.0%',
                None,'0.0%','+0.0%;-0.0%','+0.0%;-0.0%',
                None]

        for ci,(val,fmt) in enumerate(zip(vals,fmts),1):
            c = ws_c.cell(cr,ci)
            bc(c, val, LEFT if ci in (1,2,3,4,17) else CENTER, row_fill)
            if fmt and isinstance(val,(int,float)): c.number_format=fmt
            # Colour specific cells
            if ci==6:  c.fill = GREEN if f['avg_roi']>0 else RED
            if ci==12 and md: c.fill = GREEN if md['roi']>0 else RED
            if ci==15 and mo: c.fill = GREEN if mo['roi']>0 else RED
        cr += 1

# Column widths catalog
ws_c.column_dimensions['A'].width = 12
ws_c.column_dimensions['B'].width = 6
ws_c.column_dimensions['C'].width = 12
ws_c.column_dimensions['D'].width = 55
for ci in range(5, len(CAT_COLS)+1):
    ws_c.column_dimensions[get_column_letter(ci)].width = 11
ws_c.column_dimensions[get_column_letter(17)].width = 18
ws_c.freeze_panes = 'A2'
ws_c.auto_filter.ref = f'A1:{get_column_letter(len(CAT_COLS))}1'

# ════════════════════════════════════════════════════════════
# Sheets 3-12: Per-league detail
# ════════════════════════════════════════════════════════════
for lg in LEAGUES:
    lname = LEAGUE_SHORT[lg]
    niches = catalog[lg]
    if not niches: continue
    print(f"  Sheet: {lname} ({len(niches)} ніш)...")

    ws = wb.create_sheet(lname)

    # Header
    ws.merge_cells('A1:Q1')
    c = ws.cell(1,1)
    _nc = sum(1 for f in niches if f['oos_status']=='✅ підтверджено')
    _nc = sum(1 for f in niches if f['oos_status']=='✅ підтверджено')
    _nf = sum(1 for f in niches if f['oos_status']=='❌ не підтверджено')
    _nu = sum(1 for f in niches if f['oos_status']=='❓ замало даних')
    c.value = f'{lname} ({lg})  —  {len(niches)} ніш  |  ✅ {_nc}  ❌ {_nf}  ❓ {_nu}'
    c.fill=DARK_HDR; c.alignment=LEFT

    LCOLS = ['Side','Odds Range','Filter',
             'WF n','WF Avg_ROI%','WF Avg_WR%','WF Avg_EV%','WF Win%','WF W+/Wtot',
             'DISC n','DISC ROI%',
             'OOS n','OOS WR%','OOS ROI%','OOS EV%',
             'OOS статус']
    for ci,h in enumerate(LCOLS,1): hc(ws.cell(2,ci), h)

    lr = 3
    prev_key = None
    for f in niches:
        key = (f['side'], f['lo'], f['hi'])
        if key != prev_key:
            odds_label = f"{f['side'].upper()}  [{f['lo']},{f['hi']})"
            fill = GRP_HOME if f['side']=='home' else GRP_AWAY
            ws.merge_cells(start_row=lr, start_column=1,
                           end_row=lr,   end_column=len(LCOLS))
            c = ws.cell(lr,1)
            c.value=odds_label
            c.font=Font(name='Arial',bold=True,color='FFFFFF',size=10)
            c.fill=fill; c.alignment=LEFT; c.border=tb()
            lr += 1
            prev_key = key

        mo = f.get('oos'); md = f.get('disc')
        oos_stat = f['oos_status']
        row_fill = (GREEN  if oos_stat=='✅ підтверджено' else
                    YELLOW if oos_stat=='❓ замало даних'  else RED)

        vals = [
            f['side'], f"[{f['lo']},{f['hi']})", f['label'],
            f['n_tot'], f['avg_roi']/100, f['avg_wr']/100,
            f['avg_ev']/100, f['win_pct']/100,
            f"{f['n_pos']}/{f['n_windows']}",
            md['n']   if md else '—',
            (md['roi']/100) if md else '—',
            mo['n']   if mo else '—',
            (mo['wr']/100)  if mo else '—',
            (mo['roi']/100) if mo else '—',
            (mo['ev']/100)  if mo else '—',
            oos_stat,
        ]
        fmts=[None,None,None,
              None,'+0.0%;-0.0%','0.0%','+0.0%;-0.0%','0%',None,
              None,'+0.0%;-0.0%',
              None,'0.0%','+0.0%;-0.0%','+0.0%;-0.0%',None]

        for ci,(val,fmt) in enumerate(zip(vals,fmts),1):
            c = ws.cell(lr,ci)
            bc(c, val, LEFT if ci in (1,2,3,16) else CENTER, row_fill)
            if fmt and isinstance(val,(int,float)): c.number_format=fmt
            if ci==5:  c.fill = GREEN if f['avg_roi']>0 else RED
            if ci==11 and md: c.fill = GREEN if md['roi']>0 else RED
            if ci==14 and mo: c.fill = GREEN if mo['roi']>0 else RED
        lr += 1

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 55
    for ci in range(4, len(LCOLS)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 11
    ws.column_dimensions[get_column_letter(16)].width = 18
    ws.freeze_panes = 'A3'
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

outpath = os.path.join(os.path.dirname(__file__), 'niche_catalog.xlsx')
wb.save(outpath)
print(f"\nSaved: {outpath}  ({os.path.getsize(outpath)//1024}KB)")
print("\nNiches per league:")
for lg in LEAGUES:
    n = len(catalog[lg])
    conf = sum(1 for f in catalog[lg] if f['oos_status']=='✅ підтверджено')
    if n: print(f"  {LEAGUE_SHORT[lg]:<12} {n:>3} ніш  ✅ {conf} підтверджено OOS")
print("\nDone!")
