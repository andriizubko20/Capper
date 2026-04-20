import warnings; warnings.filterwarnings('ignore')
import pandas as pd
import numpy as np
import re, sys
sys.path.insert(0,'.')
from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix
from sqlalchemy import create_engine
from config.settings import settings
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATABASE_URL = settings.database_url
OUT          = '/app/BETA/kelly_simulation.xlsx'
OOS_START    = pd.Timestamp('2025-08-01')
FLAT_STAKE   = 50.0
KELLY_FRAC   = 0.25
START_BANK   = 1000.0

MODELS = {
    'Premier League': [
        'home[2.5,3.5) xg>=1.0','home[2.2,2.8) xg>=1.2 form>=1.5',
        'home[1.8,2.2) form>=2.2','home[1.7,2.0) xg>=1.8',
        'away[1.55,1.8) xg>=1.2 elo<=-75 form>=2.2',
    ],
    'Bundesliga': [
        'home[2.2,2.8) xg>=1.2 form>=1.5','home[1.8,2.2) xg>=1.2 elo>=30 mkt>=0.5',
        'home[1.8,2.2) elo>=75 form>=1.5 mkt>=0.5','home[1.55,1.8) elo>=150 form>=1.5',
        'away[2.5,3.5) xg>=1.0 elo<=-30 form>=1.5','away[2.5,3.5) form>=1.5 mkt>=0.35',
        'away[2.5,3.5) elo<=-75 mkt>=0.35','away[2.5,3.5) elo<=-30 form>=1.5 mkt>=0.35',
        'away[2.2,2.8) xg>=1.5 mkt>=0.4','away[2.2,2.8) mkt>=0.4',
        'away[2.2,2.8) elo<=-75 mkt>=0.35','away[2.0,2.5) mkt>=0.4',
        'away[1.7,2.0) xg>=1.0 form>=2.2',
    ],
    'Serie A': [
        'home[2.5,3.5) elo>=30 form>=1.5','home[2.0,2.5) xg>=1.0 mkt>=0.45',
        'home[2.0,2.5) form>=2.2 mkt>=0.45','home[1.8,2.2) xg>=1.5 form>=1.5',
        'home[1.3,1.55) xg>=1.5 elo>=150','away[2.5,3.5) xg>=1.0 form>=1.8 mkt>=0.35',
        'away[2.2,2.8) xg>=1.2 form>=1.8 mkt>=0.35','away[2.2,2.8) xg>=1.0 elo<=-150',
        'away[1.8,2.2) xg>=1.5 form>=2.2',
    ],
    'La Liga': [
        'home[2.2,2.8) xg>=1.0 form>=1.5','home[2.0,2.5) xg>=1.0 form>=1.5',
        'home[1.8,2.2) xg>=1.2 form>=1.5','home[1.8,2.2) xg>=1.0 form>=1.5 mkt>=0.5',
        'home[1.8,2.2) xg>=1.0 elo>=30 mkt>=0.5','home[1.55,1.8) xg>=1.8',
        'away[2.5,3.5) xg>=1.8 form>=1.5','away[2.5,3.5) elo<=-75 form>=2.2',
        'away[2.2,2.8) xg>=1.2 mkt>=0.4','away[1.8,2.2) xg>=1.8 elo<=-75 mkt>=0.45',
    ],
    'Ligue 1': [
        'home[2.5,3.5) xg>=1.5','home[2.2,2.8) xg>=1.2','home[2.0,2.5) elo>=75',
        'home[1.7,2.0) xg>=1.8','home[1.7,2.0) xg>=1.5 mkt>=0.5',
        'home[1.7,2.0) xg>=1.5 elo>=75','home[1.7,2.0) xg>=1.5 elo>=30',
        'home[1.55,1.8) xg>=1.2 elo>=75','away[2.5,3.5) xg>=1.0 form>=2.2',
        'away[2.2,2.8) form>=1.5 mkt>=0.4','away[2.2,2.8) elo<=-30 form>=1.5 mkt>=0.4',
    ],
    'Primeira Liga': [
        'away[2.2,2.8) xg>=1.2 mkt>=0.4','home[2.0,2.5) form>=1.5 mkt>=0.45',
        'home[1.55,1.8) elo>=150','home[1.8,2.2) form>=2.2',
    ],
    'Serie B': [
        'home[2.2,2.8) form>=2.2','home[2.0,2.5) elo>=75 form>=2.2',
        'home[1.7,2.0) xg>=1.0 form>=2.2','home[1.55,1.8) xg>=1.8',
        'home[1.55,1.8) xg>=1.5 elo>=75 form>=2.2','home[1.55,1.8) xg>=1.2 elo>=150',
        'home[1.55,1.8) xg>=1.5','away[2.0,2.5) xg>=1.5 elo<=-30',
    ],
    'Eredivisie': [
        'home[2.5,3.5) xg>=1.0','home[2.0,2.5) elo>=75',
        'away[2.5,3.5) elo<=-75 mkt>=0.35','away[1.8,2.2) xg>=1.2 form>=2.2',
        'away[1.7,2.0) xg>=1.2 form>=2.2','away[1.55,1.8) elo<=-150',
    ],
    'Jupiler Pro League': [
        'home[2.5,3.5) form>=2.2','home[2.0,2.5) xg>=1.0 mkt>=0.45',
        'home[1.8,2.2) xg>=1.2 mkt>=0.5','home[1.8,2.2) xg>=1.0 form>=1.5 mkt>=0.5',
        'home[1.7,2.0) xg>=1.0 form>=1.5 mkt>=0.5','home[1.55,1.8) xg>=1.8',
        'home[1.55,1.8) form>=2.2','home[1.3,1.55) xg>=1.5 elo>=150 form>=1.5',
        'away[2.2,2.8) xg>=1.2 form>=2.2 mkt>=0.35','away[2.0,2.5) xg>=1.5 form>=2.2',
    ],
    'Champions League': [
        'home[2.2,2.8) xg>=1.8 form>=1.5','home[2.0,2.5) form>=1.5 mkt>=0.45',
        'home[1.8,2.2) mkt>=0.5','home[1.7,2.0) xg>=1.5 mkt>=0.55',
        'home[1.7,2.0) xg>=1.5 elo>=75 mkt>=0.5','home[1.55,1.8) elo>=30 form>=2.2',
        'home[1.55,1.8) elo>=150','home[1.3,1.55) xg>=1.8 form>=1.8',
        'away[2.2,2.8) xg>=1.8 mkt>=0.4','away[2.2,2.8) xg>=1.5 form>=1.8 mkt>=0.4',
    ],
}

def parse_niche(s):
    m=re.match(r'(home|away)\[(\d+\.?\d*),(\d+\.?\d*)\)',s)
    side=m.group(1); lo=float(m.group(2)); hi=float(m.group(3))
    xg=0.0; elo=0; fm=0.0; mk=0.0
    mx=re.search(r'xg>=([\d.]+)',s);   xg=float(mx.group(1)) if mx else 0.0
    mx=re.search(r'elo>=([-\d]+)',s);  elo=int(mx.group(1)) if mx else elo
    mx=re.search(r'elo<=([-\d]+)',s);  elo=int(mx.group(1)) if mx else elo
    mx=re.search(r'form>=([\d.]+)',s); fm=float(mx.group(1)) if mx else 0.0
    mx=re.search(r'mkt>=([\d.]+)',s);  mk=float(mx.group(1)) if mx else 0.0
    return side,lo,hi,xg,elo,fm,mk

def apply_mask(data,side,lo,hi,xg,elo,fm,mk,league):
    oc='home_odds_val' if side=='home' else 'away_odds_val'
    m=(data['league_name']==league)&(data[oc]>=lo)&(data[oc]<hi)
    if xg>0:  m&=data['xg_ratio_home_5' if side=='home' else 'xg_ratio_away_5'].fillna(0)>=xg
    if side=='home' and elo>0: m&=data['elo_diff'].fillna(0)>=elo
    if side=='away' and elo<0: m&=data['elo_diff'].fillna(0)<=elo
    if fm>0:  m&=data['home_pts_5' if side=='home' else 'away_pts_5'].fillna(0)>=fm
    if mk>0:  m&=data['mkt_home_prob' if side=='home' else 'mkt_away_prob'].fillna(0)>=mk
    return data[m]

# ── Load ────────────────────────────────────────────────────────────────────────
print('Loading data...')
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result']=='H').astype(int)
df['away_win']  = (df['result']=='A').astype(int)
df['ym'] = df['date'].dt.to_period('M')
df['_id'] = np.arange(len(df))
oos_p = OOS_START.to_period('M')

print('Loading team names...')
engine = create_engine(DATABASE_URL)
with engine.connect() as conn:
    teams_df = pd.read_sql('SELECT id, name FROM teams', conn)
    matches_raw = pd.read_sql('SELECT id as match_id, home_team_id, away_team_id FROM matches', conn)
teams_map = matches_raw.merge(teams_df.rename(columns={'id':'home_team_id','name':'home_name'}), on='home_team_id', how='left')
teams_map = teams_map.merge(teams_df.rename(columns={'id':'away_team_id','name':'away_name'}), on='away_team_id', how='left')
name_map = teams_map.set_index('match_id')[['home_name','away_name']]
if 'match_id' in df.columns:
    df = df.join(name_map, on='match_id')

# ── Step 1: Collect all bets with model metadata ────────────────────────────────
print('Collecting bets...')
all_bets = []
model_oos_roi = {}  # (league, niche) → oos_roi for dedup priority

for league, niches in MODELS.items():
    for niche_str in niches:
        side,lo,hi,xg,elo,fm,mk = parse_niche(niche_str)
        wc = 'home_win' if side=='home' else 'away_win'
        oc = 'home_odds_val' if side=='home' else 'away_odds_val'
        bets = apply_mask(df,side,lo,hi,xg,elo,fm,mk,league)
        if len(bets)==0: continue

        # Full-period win rate → Kelly probability
        full_wr = bets[wc].mean()

        # OOS ROI for dedup priority
        oos = bets[bets['ym']>=oos_p]
        oos_n = len(oos)
        if oos_n > 0:
            oos_pl = (oos[wc]*(oos[oc]-1)-(1-oos[wc])).sum()*FLAT_STAKE
            oos_roi = oos_pl/(oos_n*FLAT_STAKE)*100
        else:
            oos_roi = 0.0
        model_oos_roi[(league, niche_str)] = oos_roi

        for idx, row in bets.iterrows():
            won = bool(row[wc])
            odds = float(row[oc])
            hname = row.get('home_name','') or ''
            aname = row.get('away_name','') or ''
            all_bets.append({
                'date':     row['date'],
                'ym':       row['ym'],
                'league':   league,
                'side':     side,
                'match_id': idx,
                'niche':    niche_str,
                'wc':       wc,
                'oc':       oc,
                'won':      won,
                'odds':     odds,
                'oos_roi':  oos_roi,
                'full_wr':  full_wr,
                'match':    f"{hname} vs {aname}",
                'period':   'OOS' if row['ym']>=oos_p else 'IS',
            })

bets_df = pd.DataFrame(all_bets)
print(f'  Raw bets (with dups): {len(bets_df)}')

# ── Step 2: Dedup — one bet per (match_id, side) — pick highest OOS ROI ─────────
bets_df = bets_df.sort_values(['date','match_id','side','oos_roi'], ascending=[True,True,True,False])
bets_df = bets_df.drop_duplicates(subset=['match_id','side'], keep='first').reset_index(drop=True)
print(f'  After dedup: {len(bets_df)} unique bets')

# ── Step 3: Kelly simulation (sequential, sorted by date) ──────────────────────
bets_df = bets_df.sort_values('date').reset_index(drop=True)
bankroll = START_BANK
kelly_rows = []

for _, row in bets_df.iterrows():
    p = row['full_wr']
    b = row['odds'] - 1.0
    if b <= 0 or p <= 0:
        continue
    f_star = (p * b - (1 - p)) / b
    if f_star <= 0:
        # Negative EV by Kelly → skip
        continue
    stake = KELLY_FRAC * f_star * bankroll
    stake = min(stake, 0.02 * bankroll)  # cap: max 2% of bankroll per bet
    stake = max(stake, 0.5)              # min $0.5

    if row['won']:
        pl = stake * b
    else:
        pl = -stake

    bankroll_before = bankroll
    bankroll += pl

    kelly_rows.append({
        'date':     row['date'],
        'ym':       row['ym'],
        'league':   row['league'],
        'side':     row['side'],
        'match':    row['match'],
        'niche':    row['niche'],
        'odds':     round(row['odds'],2),
        'p_est':    round(p,3),
        'f_star':   round(f_star,4),
        'stake':    round(stake,2),
        'won':      row['won'],
        'pl':       round(pl,2),
        'bankroll': round(bankroll,2),
        'period':   row['period'],
    })

kelly_df = pd.DataFrame(kelly_rows)
print(f'  Kelly bets placed: {len(kelly_df)}  (skipped {len(bets_df)-len(kelly_df)} negative EV)')

# ── Step 4: Monthly summary ─────────────────────────────────────────────────────
all_months = pd.period_range(bets_df['ym'].min(), bets_df['ym'].max(), freq='M')

flat_monthly = {}
kelly_monthly = {}
for ym, grp in bets_df.groupby('ym'):
    n = len(grp); w = int(grp['won'].sum())
    pl = grp.apply(lambda r: FLAT_STAKE*(r['odds']-1) if r['won'] else -FLAT_STAKE, axis=1).sum()
    flat_monthly[ym] = {'n':n,'w':w,'pl':round(pl,1),'roi':round(pl/(n*FLAT_STAKE)*100,1)}

for ym, grp in kelly_df.groupby('ym'):
    n = len(grp); w = int(grp['won'].sum())
    pl = grp['pl'].sum()
    stake_total = grp['stake'].sum()
    flat_monthly_roi = flat_monthly.get(ym,{}).get('roi',0)
    kelly_monthly[ym] = {'n':n,'w':w,'pl':round(pl,1),'roi':round(pl/stake_total*100,1) if stake_total>0 else 0}

# Cumulative bankroll per month
cum_bank = START_BANK
cum_flat = 0.0
monthly_bank = {}
for m in all_months:
    km = kelly_monthly.get(m,{})
    fm = flat_monthly.get(m,{})
    cum_bank_start = cum_bank
    cum_bank += km.get('pl',0)
    cum_flat  += fm.get('pl',0)
    monthly_bank[m] = {
        'kelly_pl': km.get('pl',0),
        'kelly_bank': round(cum_bank,2),
        'kelly_n': km.get('n',0),
        'kelly_w': km.get('w',0),
        'kelly_roi': km.get('roi',0),
        'flat_pl': fm.get('pl',0),
        'flat_n': fm.get('n',0),
        'flat_roi': fm.get('roi',0),
        'cum_flat': round(cum_flat,1),
        'is_oos': m >= oos_p,
    }

# ── Styles ─────────────────────────────────────────────────────────────────────
fill_hdr   = PatternFill('solid',fgColor='1F3864')
fill_oos_h = PatternFill('solid',fgColor='2E4057')
fill_lg    = PatternFill('solid',fgColor='D6E4F0')
fill_ps    = PatternFill('solid',fgColor='70AD47')
fill_p     = PatternFill('solid',fgColor='C6EFCE')
fill_neg   = PatternFill('solid',fgColor='FFCCCC')
fill_zero  = PatternFill('solid',fgColor='F5F5F5')
fill_warn  = PatternFill('solid',fgColor='FFE699')
thin       = Side(style='thin',color='CCCCCC')
border     = Border(left=thin,right=thin,top=thin,bottom=thin)

wb = Workbook()
wb.remove(wb.active)

# ── Sheet 1: Monthly Overview ───────────────────────────────────────────────────
ws_mo = wb.create_sheet('Monthly Overview')
mo_cols = ['Month','Period',
           'Flat n','Flat P&L','Flat ROI%','Flat Cumul P&L',
           'Kelly n','Kelly P&L','Kelly ROI%','Kelly Bankroll','Kelly vs Start %']
for ci,col in enumerate(mo_cols,1):
    c=ws_mo.cell(1,ci,col); c.fill=fill_hdr
    c.font=Font(color='FFFFFF',bold=True); c.alignment=Alignment(horizontal='center',wrap_text=True)
ws_mo.row_dimensions[1].height=28

flat_cum=0.0
for ri,m in enumerate(all_months,2):
    mb = monthly_bank[m]
    is_oos = mb['is_oos']
    flat_cum += mb['flat_pl']
    kelly_bank = mb['kelly_bank']
    kelly_pct = round((kelly_bank-START_BANK)/START_BANK*100,1)
    vals=[m.strftime('%b-%Y'),'OOS' if is_oos else 'IS',
          mb['flat_n'],round(mb['flat_pl'],0),mb['flat_roi'],round(flat_cum,0),
          mb['kelly_n'],round(mb['kelly_pl'],0),mb['kelly_roi'],round(kelly_bank,0),kelly_pct]
    for ci,val in enumerate(vals,1):
        c=ws_mo.cell(ri,ci,val); c.border=border; c.alignment=Alignment(horizontal='center')
        hf = fill_oos_h if is_oos else PatternFill()
        if ci in (1,2):
            if is_oos: c.fill=fill_oos_h; c.font=Font(color='FFFFFF',bold=True)
        if ci==4:
            c.number_format='+0;-0;0'
            c.fill=fill_ps if val>=200 else (fill_p if val>0 else fill_neg)
        if ci==5:
            c.number_format='+0.0;-0.0;0'
            c.fill=fill_ps if val>=15 else (fill_p if val>0 else fill_neg)
        if ci==6:
            c.number_format='+$#,##0;-$#,##0;$0'
            c.font=Font(bold=True)
        if ci==8:
            c.number_format='+0;-0;0'
            c.fill=fill_ps if val>=200 else (fill_p if val>0 else fill_neg)
        if ci==9:
            c.number_format='+0.0;-0.0;0'
            c.fill=fill_ps if val>=15 else (fill_p if val>0 else fill_neg)
        if ci==10:
            c.number_format='$#,##0'
            c.font=Font(bold=True,color='155724' if kelly_bank>=START_BANK else '721c24')
        if ci==11:
            c.number_format='+0.0;-0.0;0"%"'
            c.fill=fill_ps if val>=20 else (fill_p if val>=0 else fill_neg)
            c.font=Font(bold=True)

# Totals
ri=len(all_months)+2
is_bets = bets_df[bets_df['period']=='IS']
oos_bets = bets_df[bets_df['period']=='OOS']
total_flat_pl = bets_df.apply(lambda r: FLAT_STAKE*(r['odds']-1) if r['won'] else -FLAT_STAKE, axis=1).sum()
oos_flat_pl   = oos_bets.apply(lambda r: FLAT_STAKE*(r['odds']-1) if r['won'] else -FLAT_STAKE, axis=1).sum()
total_kelly_pl = kelly_df['pl'].sum()
oos_kelly_df  = kelly_df[kelly_df['period']=='OOS']
oos_kelly_pl  = oos_kelly_df['pl'].sum()
final_bank = kelly_df['bankroll'].iloc[-1] if len(kelly_df) else START_BANK

summary_vals=[
    'TOTAL','ALL',
    len(bets_df), round(total_flat_pl,0),
    round(total_flat_pl/(len(bets_df)*FLAT_STAKE)*100,1) if len(bets_df) else 0,
    round(total_flat_pl,0),
    len(kelly_df), round(total_kelly_pl,0),
    round(total_kelly_pl/kelly_df['stake'].sum()*100,1) if len(kelly_df) else 0,
    round(final_bank,0),
    round((final_bank-START_BANK)/START_BANK*100,1),
]
for ci,val in enumerate(summary_vals,1):
    c=ws_mo.cell(ri,ci,val); c.font=Font(bold=True,size=11); c.fill=fill_lg
    c.border=border; c.alignment=Alignment(horizontal='center')
    if ci==6: c.number_format='+$#,##0;-$#,##0;$0'
    if ci==10: c.number_format='$#,##0'

# OOS total row
ri+=1
oos_kelly_stake = oos_kelly_df['stake'].sum()
oos_vals=[
    'OOS TOTAL','OOS',
    len(oos_bets), round(oos_flat_pl,0),
    round(oos_flat_pl/(len(oos_bets)*FLAT_STAKE)*100,1) if len(oos_bets) else 0,
    '—',
    len(oos_kelly_df), round(oos_kelly_pl,0),
    round(oos_kelly_pl/oos_kelly_stake*100,1) if oos_kelly_stake>0 else 0,
    '—','—',
]
for ci,val in enumerate(oos_vals,1):
    c=ws_mo.cell(ri,ci,val); c.font=Font(bold=True); c.fill=fill_oos_h
    c.font=Font(bold=True,color='FFFFFF'); c.border=border; c.alignment=Alignment(horizontal='center')

mo_widths=[12,7,8,10,10,14,8,10,10,14,13]
for i,w in enumerate(mo_widths,1): ws_mo.column_dimensions[get_column_letter(i)].width=w
ws_mo.freeze_panes='A2'

# ── Sheet 2: Kelly Bet List ─────────────────────────────────────────────────────
ws_kl = wb.create_sheet('Kelly Bets')
kl_cols=['Date','Period','League','Match','Side','Niche','Odds','p_est','f*','Stake $','Won','P&L $','Bankroll $']
for ci,col in enumerate(kl_cols,1):
    c=ws_kl.cell(1,ci,col); c.fill=fill_hdr
    c.font=Font(color='FFFFFF',bold=True); c.alignment=Alignment(horizontal='center',wrap_text=True)

for ri,(_,row) in enumerate(kelly_df.iterrows(),2):
    is_oos = row['period']=='OOS'
    won_str = '✓' if row['won'] else '✗'
    vals=[str(row['date'])[:10], row['period'], row['league'], row['match'],
          row['side'], row['niche'], row['odds'],
          row['p_est'], row['f_star'], row['stake'],
          won_str, row['pl'], row['bankroll']]
    for ci,val in enumerate(vals,1):
        c=ws_kl.cell(ri,ci,val); c.border=border; c.alignment=Alignment(horizontal='center')
        if ci in (3,4,6): c.alignment=Alignment(horizontal='left')
        if ci==2 and is_oos: c.fill=fill_oos_h; c.font=Font(color='FFFFFF',bold=True)
        if ci==11:
            c.fill=fill_ps if row['won'] else fill_neg; c.font=Font(bold=True)
        if ci==12:
            c.number_format='+0.00;-0.00;0'
            c.fill=fill_p if row['pl']>0 else fill_neg
        if ci==13:
            c.number_format='$#,##0.00'
            c.font=Font(bold=True,color='155724' if row['bankroll']>=START_BANK else '721c24')

kl_widths=[12,7,16,32,6,42,7,7,7,9,5,10,12]
for i,w in enumerate(kl_widths,1): ws_kl.column_dimensions[get_column_letter(i)].width=w
ws_kl.freeze_panes='A2'
ws_kl.auto_filter.ref=f'A1:{get_column_letter(len(kl_cols))}1'

# ── Sheet 3: League breakdown ───────────────────────────────────────────────────
ws_lg = wb.create_sheet('By League')
lg_cols=['League','Flat n','Flat P&L','Flat ROI%','Kelly n','Kelly P&L','Kelly ROI%','Kelly Avg Stake']
for ci,col in enumerate(lg_cols,1):
    c=ws_lg.cell(1,ci,col); c.fill=fill_hdr
    c.font=Font(color='FFFFFF',bold=True); c.alignment=Alignment(horizontal='center',wrap_text=True)

ri=2
for league in MODELS.keys():
    fb = bets_df[bets_df['league']==league]
    kb = kelly_df[kelly_df['league']==league]
    if len(fb)==0: continue
    flat_pl = fb.apply(lambda r: FLAT_STAKE*(r['odds']-1) if r['won'] else -FLAT_STAKE, axis=1).sum()
    kelly_pl = kb['pl'].sum() if len(kb) else 0
    kelly_stake = kb['stake'].sum() if len(kb) else 0
    kelly_roi = kelly_pl/kelly_stake*100 if kelly_stake>0 else 0
    flat_roi = flat_pl/(len(fb)*FLAT_STAKE)*100
    avg_stake = kb['stake'].mean() if len(kb) else 0
    vals=[league, len(fb), round(flat_pl,0), round(flat_roi,1),
          len(kb), round(kelly_pl,0), round(kelly_roi,1), round(avg_stake,1)]
    for ci,val in enumerate(vals,1):
        c=ws_lg.cell(ri,ci,val); c.border=border; c.alignment=Alignment(horizontal='center')
        if ci==1: c.alignment=Alignment(horizontal='left')
        if ci==3:
            c.number_format='$#,##0'; c.fill=fill_ps if val>=500 else (fill_p if val>0 else fill_neg)
        if ci==4:
            c.number_format='+0.0;-0.0;0'; c.fill=fill_ps if val>=15 else (fill_p if val>0 else fill_neg)
            c.font=Font(bold=True)
        if ci==6:
            c.number_format='$#,##0'; c.fill=fill_ps if val>=200 else (fill_p if val>0 else fill_neg)
        if ci==7:
            c.number_format='+0.0;-0.0;0'; c.fill=fill_ps if val>=15 else (fill_p if val>0 else fill_neg)
            c.font=Font(bold=True)
    ri+=1

# Total
flat_total_pl=bets_df.apply(lambda r: FLAT_STAKE*(r['odds']-1) if r['won'] else -FLAT_STAKE, axis=1).sum()
kelly_total_stake=kelly_df['stake'].sum()
vals=['TOTAL',len(bets_df),round(flat_total_pl,0),
      round(flat_total_pl/(len(bets_df)*FLAT_STAKE)*100,1),
      len(kelly_df),round(total_kelly_pl,0),
      round(total_kelly_pl/kelly_total_stake*100,1) if kelly_total_stake else 0,
      round(kelly_df['stake'].mean(),1) if len(kelly_df) else 0]
for ci,val in enumerate(vals,1):
    c=ws_lg.cell(ri,ci,val); c.border=border; c.font=Font(bold=True,size=11)
    c.fill=fill_lg; c.alignment=Alignment(horizontal='center')
    if ci==3: c.number_format='$#,##0'
    if ci==6: c.number_format='$#,##0'

lg_widths=[20,9,11,11,9,11,11,12]
for i,w in enumerate(lg_widths,1): ws_lg.column_dimensions[get_column_letter(i)].width=w
ws_lg.freeze_panes='A2'

wb.save(OUT)

# ── Console summary ─────────────────────────────────────────────────────────────
print(f'\n{"="*60}')
print(f'  Total unique bets (deduped): {len(bets_df)}')
print(f'  Kelly bets placed:           {len(kelly_df)}')
print()
print(f'  FLAT STAKE ${FLAT_STAKE:.0f}:')
print(f'    Total P&L:  ${total_flat_pl:>+,.0f}')
print(f'    OOS P&L:    ${oos_flat_pl:>+,.0f}')
print(f'    Total ROI:  {total_flat_pl/(len(bets_df)*FLAT_STAKE)*100:>+.1f}%')
print()
print(f'  KELLY 25% (start ${START_BANK:.0f}):')
print(f'    Final bank: ${final_bank:>,.0f}  ({(final_bank-START_BANK)/START_BANK*100:>+.1f}%)')
print(f'    Total P&L:  ${total_kelly_pl:>+,.0f}')
print(f'    OOS P&L:    ${oos_kelly_pl:>+,.0f}')
print(f'    Avg stake:  ${kelly_df["stake"].mean():.1f}')
print(f'    Min stake:  ${kelly_df["stake"].min():.2f}')
print(f'    Max stake:  ${kelly_df["stake"].max():.2f}')
print()

# IS vs OOS breakdown
is_kelly = kelly_df[kelly_df['period']=='IS']
print(f'  IS  Kelly P&L: ${is_kelly["pl"].sum():>+,.0f}  (n={len(is_kelly)})')
print(f'  OOS Kelly P&L: ${oos_kelly_pl:>+,.0f}  (n={len(oos_kelly_df)})')
print(f'\nSaved: {OUT}')
print('✅ Done!')
