import warnings; warnings.filterwarnings('ignore')
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = '/app/BETA/expanding_wf_v3_results.xlsx'
OUT = '/app/BETA/oos_confirmed_niches.xlsx'

MIN_OOS_N   = 3
MIN_OOS_ROI = 0.0   # > 0

xl = pd.ExcelFile(SRC)
leagues = [s for s in xl.sheet_names if s != 'SUMMARY']

all_rows = []

for lg in leagues:
    raw = pd.read_excel(xl, sheet_name=lg, header=None)
    current_vol = None
    for i in range(2, len(raw)):
        row = raw.iloc[i]
        niche = str(row[0])
        if 'HIGH VOLUME' in niche:
            current_vol = 'High'
            continue
        if 'LOW VOLUME' in niche:
            current_vol = 'Low'
            continue
        if niche == 'nan' or current_vol is None:
            continue

        try:
            oos_n   = float(row[8])
            oos_wr  = float(row[9])
            oos_roi = float(row[10])
        except:
            continue

        if oos_n < MIN_OOS_N or oos_roi <= MIN_OOS_ROI:
            continue

        prof_folds = str(row[3])
        stability  = float(row[4]) if pd.notna(row[4]) else None
        avg_te_n   = float(row[5]) if pd.notna(row[5]) else None
        avg_te_roi = float(row[6]) if pd.notna(row[6]) else None
        side       = str(row[1])

        fold_data = []
        for fc, tc in [(11,12),(13,14),(15,16),(17,18)]:
            try:
                fold_data.append((float(row[fc]), float(row[tc])))
            except (ValueError, TypeError):
                fold_data.append((None, None))

        all_rows.append({
            'League':      lg,
            'Volume':      current_vol,
            'Niche':       niche,
            'Side':        side,
            'Prof/Folds':  prof_folds,
            'Stability':   stability,
            'Avg Te n':    avg_te_n,
            'Avg Te ROI':  avg_te_roi,
            'OOS n':       oos_n,
            'OOS WR':      oos_wr,
            'OOS ROI':     oos_roi,
            '_fold_data':  fold_data,
        })

df = pd.DataFrame(all_rows)
df = df.sort_values(['League', 'Volume', 'OOS ROI'], ascending=[True, True, False])

print(f'Total OOS-confirmed niches: {len(df)}')
print(f"  High volume: {(df['Volume']=='High').sum()}")
print(f"  Low volume:  {(df['Volume']=='Low').sum()}")
print()
print('Per league:')
for lg, g in df.groupby('League'):
    h = (g['Volume']=='High').sum()
    l = (g['Volume']=='Low').sum()
    print(f"  {lg:<25s} High={h:3d}  Low={l:3d}  Total={len(g):3d}")

# ── Excel output ──────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = 'OOS Confirmed'

fill_header  = PatternFill('solid', fgColor='1F3864')
fill_high    = PatternFill('solid', fgColor='E2EFDA')
fill_low     = PatternFill('solid', fgColor='FFF2CC')
fill_league  = PatternFill('solid', fgColor='D6E4F0')
fill_pos     = PatternFill('solid', fgColor='C6EFCE')
fill_strong  = PatternFill('solid', fgColor='70AD47')

COLS = ['League','Volume','Niche','Side','Prof/Folds','Stability','Avg Te n',
        'Avg Te ROI%','OOS n','OOS WR%','OOS ROI%',
        'F3 Tr','F3 Te','F4 Tr','F4 Te','F5 Tr','F5 Te','F6 Tr','F6 Te']

for ci, col in enumerate(COLS, 1):
    c = ws.cell(row=1, column=ci, value=col)
    c.fill = fill_header
    c.font = Font(color='FFFFFF', bold=True, size=10)
    c.alignment = Alignment(horizontal='center', wrap_text=True)

thin   = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

row_idx    = 2
prev_league = None

for _, r in df.iterrows():
    if r['League'] != prev_league:
        c = ws.cell(row=row_idx, column=1, value=r['League'])
        c.fill = fill_league
        c.font = Font(bold=True, size=11)
        ws.merge_cells(start_row=row_idx, start_column=1,
                       end_row=row_idx, end_column=len(COLS))
        row_idx    += 1
        prev_league = r['League']

    fill    = fill_high if r['Volume'] == 'High' else fill_low
    oos_roi = r['OOS ROI']
    oos_fill = fill_strong if oos_roi >= 0.30 else (fill_pos if oos_roi >= 0.10 else None)

    fd = r['_fold_data']
    values = [
        r['League'], r['Volume'], r['Niche'], r['Side'], r['Prof/Folds'],
        r['Stability'], r['Avg Te n'], r['Avg Te ROI'],
        int(r['OOS n']), r['OOS WR'], oos_roi,
    ]
    for tr, te in fd:
        values.extend([tr, te])
    while len(values) < len(COLS):
        values.append(None)

    for ci, val in enumerate(values, 1):
        c = ws.cell(row=row_idx, column=ci, value=val)
        c.fill = fill
        c.border = border
        c.alignment = Alignment(horizontal='center')
        if ci == 3:
            c.alignment = Alignment(horizontal='left')
        if ci == 11 and oos_fill:
            c.fill = oos_fill
            c.font = Font(bold=True)
        if isinstance(val, float) and ci >= 8:
            c.number_format = '+0.0%;-0.0%;0%'

    row_idx += 1

widths = [14, 7, 38, 6, 10, 9, 8, 10, 6, 8, 10, 7, 7, 7, 7, 7, 7, 7, 7]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}1'

# ── Summary sheet ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Summary')
headers = ['League','High OOS+','Low OOS+','Total','Best OOS ROI','Avg OOS ROI']
for ci, h in enumerate(headers, 1):
    c = ws2.cell(1, ci, h)
    c.fill = fill_header
    c.font = Font(color='FFFFFF', bold=True)
    c.alignment = Alignment(horizontal='center')

for ri, (lg, g) in enumerate(df.groupby('League'), 2):
    h    = (g['Volume']=='High').sum()
    l    = (g['Volume']=='Low').sum()
    best = g['OOS ROI'].max()
    avg  = g['OOS ROI'].mean()
    ws2.cell(ri,1,lg)
    ws2.cell(ri,2,h)
    ws2.cell(ri,3,l)
    ws2.cell(ri,4,len(g))
    for ci, val in [(5,best),(6,avg)]:
        c = ws2.cell(ri,ci,val)
        c.number_format = '+0.0%;-0.0%;0%'
        c.fill = fill_strong if val>=0.3 else (fill_pos if val>=0.1 else PatternFill())

for col in range(1,7):
    ws2.column_dimensions[get_column_letter(col)].width = 20

wb.save(OUT)
print(f'\nSaved: {OUT}')
print('✅ Done!')
