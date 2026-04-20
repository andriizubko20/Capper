"""
BETA/model_v7.py

Model v7 — портфоліо з ніш (армія міні-моделей).
Базується на нішевому каталозі: avg_roi>=10%, win%>=60%, n>=10, windows>=4.

Симуляція на OOS: 2025-11 → 2026-04
Початковий банкрол: $1000
Kelly: fractional 25%, cap per niche = 4%
1 bet per match: якщо кілька ніш спрацьовують на один матч — обирається найкраща (вищий avg_roi).
"""
import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import numpy as np
import pandas as pd
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as BdrSide
from openpyxl.utils import get_column_letter

from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix

# ── Config ────────────────────────────────────────────────────────────────────
KELLY_FRAC   = 0.25
NICHE_CAP    = 0.04   # max stake per niche = 4% of bankroll
MIN_ROI      = 20.0   # мінімум avg_roi% для включення ніші
WIN_PCT_MIN  = 65.0
MIN_N        = 10
MIN_WINDOWS  = 6
START_BANK   = 1000.0

DISC_END  = pd.Timestamp('2025-10-31')
OOS_START = pd.Timestamp('2025-11-01')

# ── Load ──────────────────────────────────────────────────────────────────────
print("Loading data...")
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result'] == 'H').astype(int)
df['away_win']  = (df['result'] == 'A').astype(int)

disc = df[df['date'] <= DISC_END].copy()
oos  = df[df['date'] >= OOS_START].copy().sort_values('date').reset_index(drop=True)
print(f"Discovery: {len(disc)} | OOS: {len(oos)} ({oos['date'].min().date()} – {oos['date'].max().date()})")

LEAGUES = [
    'Premier League','Bundesliga','Serie A','La Liga','Ligue 1',
    'Primeira Liga','Serie B','Eredivisie','Jupiler Pro League','Champions League'
]
LEAGUE_SHORT = {
    'Premier League':'EPL','Bundesliga':'Bundesliga','Serie A':'Serie A',
    'La Liga':'La Liga','Ligue 1':'Ligue 1','Primeira Liga':'Portugal',
    'Serie B':'Serie B','Eredivisie':'Eredivisie',
    'Jupiler Pro League':'Jupiler','Champions League':'UCL'
}

# ── Walk-forward windows ──────────────────────────────────────────────────────
WF_WINDOWS = []
cur = pd.Timestamp('2023-08-01')
end_wf = pd.Timestamp('2025-10-31')  # строго до OOS (2025-11+)
while cur < end_wf:
    w_end = cur + pd.DateOffset(months=3) - pd.Timedelta(days=1)
    if w_end > end_wf: w_end = end_wf
    WF_WINDOWS.append((cur.strftime('%Y-%m'), cur, w_end))
    cur += pd.DateOffset(months=3)
WIN_LABELS = [w[0] for w in WF_WINDOWS]
wf_slices  = {wlbl: df[(df['date']>=ws)&(df['date']<=we)]
              for wlbl,ws,we in WF_WINDOWS}

# ── Filter space ──────────────────────────────────────────────────────────────
ODDS_RANGES = [
    (1.30,1.55),(1.55,1.80),(1.70,2.00),
    (1.80,2.20),(2.00,2.50),(2.20,2.80),(2.50,3.50),
]
XG_THRS    = [0.0,1.0,1.2,1.5,1.8]
ELO_H_THRS = [0,30,75,150]
ELO_A_THRS = [0,-30,-75,-150]
FORM_THRS  = [0.0,1.5,1.8,2.2]
MKT_H_THRS = [0.0,0.45,0.50,0.55]
MKT_A_THRS = [0.0,0.35,0.40,0.45]
MIN_N_WIN  = 3

def calc(sub, side):
    wc = 'home_win' if side=='home' else 'away_win'
    oc = 'home_odds_val' if side=='home' else 'away_odds_val'
    if len(sub) < MIN_N_WIN: return None
    wr  = sub[wc].mean()
    ao  = sub[oc].mean()
    roi = (sub[wc]*(sub[oc]-1)-(1-sub[wc])).mean()*100
    ev  = (wr*ao-1)*100
    return {'n':len(sub),'wr':round(wr,4),'roi':round(roi,2),
            'ev':round(ev,2),'avg_odds':round(ao,4)}

def apply_mask(d, side, lo, hi, xg_t, elo_t, form_t, mkt_t):
    oc  = 'home_odds_val'   if side=='home' else 'away_odds_val'
    xgc = 'xg_ratio_home_5' if side=='home' else 'xg_ratio_away_5'
    fc  = 'home_pts_5'      if side=='home' else 'away_pts_5'
    mc  = 'mkt_home_prob'   if side=='home' else 'mkt_away_prob'
    m = (d[oc]>=lo)&(d[oc]<hi)
    if xg_t  > 0: m &= (d[xgc].fillna(0)>=xg_t)
    if side=='home' and elo_t>0: m &= (d['elo_diff'].fillna(0)>=elo_t)
    if side=='away' and elo_t<0: m &= (d['elo_diff'].fillna(0)<=elo_t)
    if form_t > 0: m &= (d[fc].fillna(0)>=form_t)
    if mkt_t  > 0: m &= (d[mc].fillna(0)>=mkt_t)
    return d[m]

# ── STEP 1: Build niche catalog from walk-forward ─────────────────────────────
print("Building niche catalog...")
raw_stable = {lg: [] for lg in LEAGUES}

for side in ('home','away'):
    et = ELO_H_THRS if side=='home' else ELO_A_THRS
    mt = MKT_H_THRS if side=='home' else MKT_A_THRS
    for lo,hi in ODDS_RANGES:
        for xg_t,elo_t,form_t,mkt_t in product(XG_THRS,et,FORM_THRS,mt):
            parts = [f'{side}[{lo},{hi})']
            if xg_t  > 0: parts.append(f'xg>={xg_t}')
            if side=='home' and elo_t>0: parts.append(f'elo>={elo_t}')
            if side=='away' and elo_t<0: parts.append(f'elo<={elo_t}')
            if form_t > 0: parts.append(f'form>={form_t}')
            if mkt_t  > 0: parts.append(f'mkt>={mkt_t}')
            label = ' '.join(parts)

            win_data = {}
            for wlbl,_,_ in WF_WINDOWS:
                sub = apply_mask(wf_slices[wlbl], side, lo, hi, xg_t, elo_t, form_t, mkt_t)
                for lg in LEAGUES:
                    sub_l = sub[sub['league_name']==lg]
                    m = calc(sub_l, side)
                    if m:
                        win_data.setdefault(lg, {})[wlbl] = m

            for lg, wins in win_data.items():
                if not wins: continue
                n_tot   = sum(w['n']   for w in wins.values())
                avg_roi = np.mean([w['roi'] for w in wins.values()])
                avg_wr  = np.mean([w['wr']  for w in wins.values()])
                avg_ao  = np.mean([w['avg_odds'] for w in wins.values()])
                n_pos   = sum(1 for w in wins.values() if w['roi']>0)
                n_w     = len(wins)
                win_pct = n_pos/n_w*100

                if n_tot < MIN_N or n_w < MIN_WINDOWS: continue
                if avg_roi < MIN_ROI or win_pct < WIN_PCT_MIN: continue

                raw_stable[lg].append({
                    'label':label,'side':side,'lo':lo,'hi':hi,
                    'xg_t':xg_t,'elo_t':elo_t,'form_t':form_t,'mkt_t':mkt_t,
                    'n_tot':n_tot,'avg_roi':round(avg_roi,1),
                    'avg_wr':round(avg_wr,4),'avg_odds':round(avg_ao,4),
                    'win_pct':round(win_pct,1),'n_pos':n_pos,'n_windows':n_w,
                })

# Dedup subsets
def is_subset(a, b, side):
    if a['xg_t']>b['xg_t'] or a['form_t']>b['form_t'] or a['mkt_t']>b['mkt_t']:
        return False
    if side=='home': return a['elo_t']<=b['elo_t']
    else:            return a['elo_t']>=b['elo_t']

def dedup(filters):
    filters = sorted(filters, key=lambda x: x['avg_roi'], reverse=True)
    kept = []
    for f in filters:
        if not any(is_subset(k,f,f['side']) and k['avg_roi']>=f['avg_roi'] for k in kept):
            kept.append(f)
    return kept

catalog = {}
for lg in LEAGUES:
    groups = {}
    for f in raw_stable[lg]:
        groups.setdefault((f['side'],f['lo'],f['hi']),[]).append(f)
    catalog[lg] = []
    for group in groups.values():
        catalog[lg].extend(dedup(group))

total_niches = sum(len(v) for v in catalog.values())
print(f"Catalog: {total_niches} niches across {len(LEAGUES)} leagues")
for lg in LEAGUES:
    print(f"  {LEAGUE_SHORT[lg]:<12} {len(catalog[lg]):>3} ніш")

# ── STEP 2: OOS Simulation ────────────────────────────────────────────────────
print("\nRunning OOS simulation...")

bank = START_BANK
equity_curve = []   # (date, bank)
bet_log = []        # all individual bets

for idx, row in oos.iterrows():
    lg      = row['league_name']
    if lg not in catalog: continue
    niches  = catalog[lg]
    date    = row['date']
    result  = row['result']

    candidates = []
    for niche in niches:
        side = niche['side']
        oc   = 'home_odds_val' if side=='home' else 'away_odds_val'
        wc   = 'home_win'      if side=='home' else 'away_win'

        # Check if match passes niche filter
        sub = apply_mask(pd.DataFrame([row]), side,
                         niche['lo'], niche['hi'],
                         niche['xg_t'], niche['elo_t'],
                         niche['form_t'], niche['mkt_t'])
        if len(sub) == 0: continue

        odds = row[oc]
        won  = int(row[wc])

        # Kelly sizing based on niche's avg_wr
        wr = niche['avg_wr']
        b  = odds - 1
        f  = max(0.0, (wr*b - (1-wr)) / b) * KELLY_FRAC
        f  = min(f, NICHE_CAP)
        if f <= 0: continue

        candidates.append({
            'date':   date,
            'league': LEAGUE_SHORT.get(lg, lg),
            'niche':  niche['label'],
            'side':   side,
            'odds':   round(odds, 2),
            'won':    won,
            'kelly_f': round(f*100, 2),
            'niche_wr':  round(wr*100, 1),
            'niche_roi': niche['avg_roi'],
            '_f': f,
            '_b': b,
        })

    # Pick only the best niche per match (highest avg_roi)
    if candidates:
        best = max(candidates, key=lambda x: x['niche_roi'])
        f, b = best.pop('_f'), best.pop('_b')
        stake  = round(bank * f, 2)
        profit = round(stake * b if best['won'] else -stake, 2)
        best['stake']       = stake
        best['profit']      = profit
        best['bank_before'] = round(bank, 2)
        bank += profit
        bank  = max(bank, 0.01)
        best['bank_after'] = round(bank, 2)
        bet_log.append(best)

    equity_curve.append({'date': date, 'bank': round(bank, 2)})

df_bets   = pd.DataFrame(bet_log)
df_equity = pd.DataFrame(equity_curve).drop_duplicates('date')

print(f"\nTotal bets:   {len(df_bets)}")
if len(df_bets) > 0:
    won = df_bets['won'].sum()
    wr  = won/len(df_bets)*100
    total_profit = df_bets['profit'].sum()
    roi = total_profit / df_bets['stake'].sum() * 100
    avg_odds = df_bets['odds'].mean()
    # Max drawdown
    peak = START_BANK
    max_dd = 0
    for b in df_equity['bank']:
        if b > peak: peak = b
        dd = (peak - b) / peak * 100
        if dd > max_dd: max_dd = dd

    print(f"Win rate:     {wr:.1f}%  ({int(won)}/{len(df_bets)})")
    print(f"Avg odds:     {avg_odds:.2f}")
    print(f"Flat ROI:     {roi:+.1f}%")
    print(f"Start bank:   ${START_BANK:,.0f}")
    print(f"Final bank:   ${bank:,.2f}  ({bank/START_BANK:.2f}x)")
    print(f"Max drawdown: {max_dd:.1f}%")
    print(f"Bets/month:   {len(df_bets)/6:.0f}")

# ── STEP 3: Per-league breakdown ──────────────────────────────────────────────
print("\nPer-league breakdown:")
league_stats = []
for lg in LEAGUES:
    lname = LEAGUE_SHORT[lg]
    sub = df_bets[df_bets['league']==lname] if len(df_bets)>0 else pd.DataFrame()
    if len(sub)==0:
        league_stats.append({'league':lname,'n':0,'wr':0,'roi':0,'profit':0})
        continue
    n    = len(sub)
    wr   = sub['won'].mean()*100
    prof = sub['profit'].sum()
    roi  = prof/sub['stake'].sum()*100
    print(f"  {lname:<12} {n:>4} bets  WR={wr:.1f}%  ROI={roi:+.1f}%  P/L=${prof:+.0f}")
    league_stats.append({'league':lname,'n':n,'wr':round(wr,1),
                         'roi':round(roi,1),'profit':round(prof,2)})

# ── STEP 4: Build Excel ───────────────────────────────────────────────────────
print("\nBuilding Excel report...")

GREEN  = PatternFill('solid', start_color='D5F5E3')
RED    = PatternFill('solid', start_color='FADBD8')
GREY   = PatternFill('solid', start_color='EAECEE')
DARK   = PatternFill('solid', start_color='1B2631')
BLUE   = PatternFill('solid', start_color='154360')
YELLOW = PatternFill('solid', start_color='FEF9E7')

HDR  = Font(name='Arial', bold=True, color='FFFFFF', size=9)
BODY = Font(name='Arial', size=9)
BOLD = Font(name='Arial', bold=True, size=9)
C    = Alignment(horizontal='center', vertical='center')
L    = Alignment(horizontal='left',   vertical='center')

def tb():
    s = BdrSide(style='thin', color='CCCCCC')
    return Border(left=s,right=s,top=s,bottom=s)

def hc(c,v,fill=DARK):
    c.value=v; c.font=HDR; c.fill=fill; c.alignment=C; c.border=tb()

def bc(c,v,align=C,fill=None,bold=False,fmt=None):
    c.value=v; c.font=Font(name='Arial',size=9,bold=bold)
    c.alignment=align; c.border=tb()
    if fill: c.fill=fill
    if fmt:  c.number_format=fmt

wb = Workbook()
wb.remove(wb.active)

# ── Sheet 1: SUMMARY ──────────────────────────────────────────────────────────
ws = wb.create_sheet('SUMMARY')

summary_data = [
    ('Початковий банкрол', f'${START_BANK:,.0f}'),
    ('Фінальний банкрол', f'${bank:,.2f}'),
    ('Зростання', f'{bank/START_BANK:.2f}x'),
    ('Загальний P/L', f'${bank-START_BANK:+,.2f}'),
    ('Flat ROI', f'{roi:+.1f}%'),
    ('Кількість ставок', len(df_bets)),
    ('Win Rate', f'{wr:.1f}%'),
    ('Середній коефіцієнт', f'{avg_odds:.2f}'),
    ('Max Drawdown', f'{max_dd:.1f}%'),
    ('Ставок на місяць', f'{len(df_bets)/6:.0f}'),
    ('OOS Period', f'{oos["date"].min().date()} – {oos["date"].max().date()}'),
    ('Ніш в каталозі', total_niches),
    ('Kelly fraction', f'{KELLY_FRAC*100:.0f}%'),
    ('Cap per niche', f'{NICHE_CAP*100:.0f}%'),
    ('Min avg ROI (WF)', f'{MIN_ROI:.0f}%'),
    ('Min Win% (WF)', f'{WIN_PCT_MIN:.0f}%'),
]

ws['A1'] = 'Model v7 — OOS Simulation Results'
ws['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
ws['A1'].fill = DARK
ws['A1'].alignment = L
ws.merge_cells('A1:C1')

for i, (k, v) in enumerate(summary_data, 3):
    c_k = ws.cell(i, 1); c_v = ws.cell(i, 2)
    c_k.value = k; c_k.font=BOLD; c_k.alignment=L; c_k.border=tb()
    c_k.fill = PatternFill('solid', start_color='EBF5FB')
    c_v.value = v; c_v.font=BODY; c_v.alignment=C; c_v.border=tb()
    # Highlight key metrics
    if k == 'Фінальний банкрол':
        c_v.fill = GREEN if bank > START_BANK else RED
        c_v.font = Font(name='Arial', bold=True, size=11)
    elif k == 'Зростання':
        c_v.fill = GREEN if bank > START_BANK else RED
    elif k == 'Max Drawdown':
        c_v.fill = (GREEN if max_dd<20 else YELLOW if max_dd<35 else RED)

# Per-league table
ws.cell(len(summary_data)+5, 1).value = 'Per-league breakdown'
ws.cell(len(summary_data)+5, 1).font = Font(name='Arial',bold=True,size=11,color='FFFFFF')
ws.cell(len(summary_data)+5, 1).fill = BLUE
ws.merge_cells(f'A{len(summary_data)+5}:E{len(summary_data)+5}')

lr = len(summary_data)+6
for h,ci in zip(['Ліга','Ставок','Win Rate','Flat ROI','P/L ($)'],range(1,6)):
    hc(ws.cell(lr,ci), h)
lr += 1
for s in sorted(league_stats, key=lambda x: x['profit'], reverse=True):
    if s['n']==0: continue
    row = [s['league'],s['n'],f"{s['wr']:.1f}%",f"{s['roi']:+.1f}%",f"${s['profit']:+.0f}"]
    for ci, val in enumerate(row, 1):
        c = ws.cell(lr, ci)
        bc(c, val, L if ci==1 else C)
        if ci==4: c.fill = GREEN if s['roi']>0 else RED
        if ci==5: c.fill = GREEN if s['profit']>0 else RED
    lr += 1

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 12

# ── Sheet 2: EQUITY CURVE ─────────────────────────────────────────────────────
ws_eq = wb.create_sheet('EQUITY CURVE')
for ci,h in enumerate(['Date','Bank ($)','Bets that day','Daily P/L','Drawdown%'],1):
    hc(ws_eq.cell(1,ci), h)

# Daily equity
daily = df_bets.groupby('date').agg(
    n_bets=('stake','count'),
    daily_pl=('profit','sum')
).reset_index()
daily = daily.sort_values('date')

er = 2
peak_eq = START_BANK
curr_bank = START_BANK
for _, drow in daily.iterrows():
    curr_bank += drow['daily_pl']
    curr_bank  = max(curr_bank, 0.01)
    if curr_bank > peak_eq: peak_eq = curr_bank
    dd = (peak_eq - curr_bank)/peak_eq*100

    row = [drow['date'].strftime('%Y-%m-%d'),
           round(curr_bank,2), int(drow['n_bets']),
           round(drow['daily_pl'],2), round(dd,1)]
    for ci, val in enumerate(row, 1):
        c = ws_eq.cell(er, ci)
        bc(c, val)
        if ci==2: c.fill = GREEN if curr_bank>=START_BANK else RED
        if ci==4: c.fill = GREEN if drow['daily_pl']>0 else (RED if drow['daily_pl']<0 else GREY)
        if ci==5: c.fill = (GREEN if dd<5 else YELLOW if dd<20 else RED)
    er += 1

for ci,w in zip(range(1,6),[12,12,12,12,12]):
    ws_eq.column_dimensions[get_column_letter(ci)].width = w
ws_eq.freeze_panes = 'A2'

# ── Sheet 3: ALL BETS ─────────────────────────────────────────────────────────
ws_b = wb.create_sheet('ALL BETS')
BET_COLS = ['Date','League','Niche','Side','Odds','Stake($)','Won','Profit($)',
            'Bank Before($)','Bank After($)','Kelly%','Niche WR%','Niche ROI%']
for ci,h in enumerate(BET_COLS,1): hc(ws_b.cell(1,ci), h)

for br, bet in enumerate(bet_log, 2):
    row = [bet['date'].strftime('%Y-%m-%d'), bet['league'],
           bet['niche'], bet['side'],
           bet['odds'], bet['stake'],
           '✅' if bet['won'] else '❌',
           bet['profit'], bet['bank_before'], bet['bank_after'],
           bet['kelly_f'], bet['niche_wr'], bet['niche_roi']]
    fmts=[None]*4+[None,'0.00',None,'+0.00;-0.00','0.00','0.00',
                   '0.00%','0.0%','+0.0%']
    for ci,(val,fmt) in enumerate(zip(row,fmts),1):
        c = ws_b.cell(br, ci)
        bc(c, val, L if ci in (1,2,3,4,7) else C)
        if ci==7:  c.fill = GREEN if bet['won'] else RED
        if ci==8:  c.fill = GREEN if bet['profit']>0 else RED

ws_b.column_dimensions['A'].width = 12
ws_b.column_dimensions['B'].width = 12
ws_b.column_dimensions['C'].width = 52
ws_b.column_dimensions['D'].width = 6
for ci in range(5, len(BET_COLS)+1):
    ws_b.column_dimensions[get_column_letter(ci)].width = 12
ws_b.freeze_panes = 'A2'
ws_b.auto_filter.ref = f'A1:{get_column_letter(len(BET_COLS))}1'

# ── Sheet 4: NICHE PERFORMANCE ───────────────────────────────────────────────
ws_n = wb.create_sheet('NICHE PERFORMANCE')
NP_COLS = ['League','Niche','Side','WF Avg_ROI%','WF WR%',
           'OOS Bets','OOS WR%','OOS ROI%','OOS P/L($)','Status']
for ci,h in enumerate(NP_COLS,1): hc(ws_n.cell(1,ci), h)

niche_perf = []
for lg in LEAGUES:
    lname = LEAGUE_SHORT[lg]
    for niche in catalog[lg]:
        sub = df_bets[df_bets['niche']==niche['label']] if len(df_bets)>0 else pd.DataFrame()
        sub = sub[sub['league']==lname] if len(sub)>0 else sub
        if len(sub) == 0:
            niche_perf.append({
                'league':lname,'label':niche['label'],'side':niche['side'],
                'wf_roi':niche['avg_roi'],'wf_wr':round(niche['avg_wr']*100,1),
                'n':0,'oos_wr':None,'oos_roi':None,'oos_pl':0,'status':'❓ 0 ставок'
            })
        else:
            oos_wr  = sub['won'].mean()*100
            oos_roi = sub['profit'].sum()/sub['stake'].sum()*100
            oos_pl  = sub['profit'].sum()
            status  = '✅' if oos_roi>0 else '❌'
            niche_perf.append({
                'league':lname,'label':niche['label'],'side':niche['side'],
                'wf_roi':niche['avg_roi'],'wf_wr':round(niche['avg_wr']*100,1),
                'n':len(sub),'oos_wr':round(oos_wr,1),
                'oos_roi':round(oos_roi,1),'oos_pl':round(oos_pl,2),
                'status':status
            })

niche_perf.sort(key=lambda x: (x['oos_pl'] or 0), reverse=True)

for nr, np_ in enumerate(niche_perf, 2):
    row = [np_['league'], np_['label'], np_['side'],
           np_['wf_roi']/100, np_['wf_wr']/100,
           np_['n'],
           (np_['oos_wr']/100) if np_['oos_wr'] else '—',
           (np_['oos_roi']/100) if np_['oos_roi'] is not None else '—',
           np_['oos_pl'], np_['status']]
    fmts=[None,None,None,'+0.0%;-0.0%','0.0%',
          None,'0.0%','+0.0%;-0.0%','#,##0.00;[Red]-#,##0.00',None]
    for ci,(val,fmt) in enumerate(zip(row,fmts),1):
        c = ws_n.cell(nr, ci)
        f = GREEN if np_['status']=='✅' else (RED if np_['status']=='❌' else GREY)
        bc(c, val, L if ci in (1,2,3,10) else C, f if ci in (8,9,10) else None)
        if fmt and isinstance(val,(int,float)): c.number_format=fmt

ws_n.column_dimensions['A'].width = 12
ws_n.column_dimensions['B'].width = 52
ws_n.column_dimensions['C'].width = 6
for ci in range(4, len(NP_COLS)+1):
    ws_n.column_dimensions[get_column_letter(ci)].width = 12
ws_n.freeze_panes = 'A2'
ws_n.auto_filter.ref = f'A1:{get_column_letter(len(NP_COLS))}1'

outpath = os.path.join(os.path.dirname(__file__), 'model_v7_results.xlsx')
wb.save(outpath)
print(f"\nSaved: {outpath}  ({os.path.getsize(outpath)//1024}KB)")
print("\n✅ Done!")
