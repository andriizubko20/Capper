import warnings; warnings.filterwarnings('ignore')
import pandas as pd
import numpy as np
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys, os
sys.path.insert(0, '.')
from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix

CONFIRMED_SRC = '/app/BETA/oos_confirmed_niches.xlsx'
OUT           = '/app/BETA/oos_simulation_results.xlsx'
INITIAL_BANK  = 1000.0
FLAT_STAKE    = 50.0   # 5% of initial

OOS_START = pd.Timestamp('2025-11-01')
OOS_END   = pd.Timestamp('2026-04-11')

# ── Load data ─────────────────────────────────────────────────────────────────
print('Loading data...')
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result'] == 'H').astype(int)
df['away_win'] = (df['result'] == 'A').astype(int)

oos = df[(df['date'] >= OOS_START) & (df['date'] <= OOS_END)].copy()
print(f'OOS matches: {len(oos)}  ({OOS_START.date()} – {OOS_END.date()})')

# ── Parse niche string → filter params ───────────────────────────────────────
def parse_niche(niche_str, side):
    params = {}
    # odds range e.g. home[1.3,1.55) or away[2.0,2.5)
    m = re.search(r'\[(\d+\.?\d*),(\d+\.?\d*)\)', niche_str)
    if m:
        params['odds_lo'] = float(m.group(1))
        params['odds_hi'] = float(m.group(2))
    # xg>=X
    m = re.search(r'xg>=(\d+\.?\d*)', niche_str)
    params['xg'] = float(m.group(1)) if m else 0.0
    # elo>=X or elo<=-X
    m = re.search(r'elo>=([-\d]+)', niche_str)
    if m: params['elo'] = float(m.group(1))
    m = re.search(r'elo<=([-\d]+)', niche_str)
    if m: params['elo'] = float(m.group(1))
    if 'elo' not in params: params['elo'] = 0
    # form>=X
    m = re.search(r'form>=(\d+\.?\d*)', niche_str)
    params['form'] = float(m.group(1)) if m else 0.0
    # mkt>=X
    m = re.search(r'mkt>=(\d+\.?\d*)', niche_str)
    params['mkt'] = float(m.group(1)) if m else 0.0
    return params

def apply_mask(data, side, p, league):
    oc  = 'home_odds_val' if side == 'home' else 'away_odds_val'
    m = (
        (data['league_name'] == league) &
        (data[oc] >= p['odds_lo']) &
        (data[oc] <  p['odds_hi'])
    )
    if p['xg'] > 0:
        xg_col = 'xg_ratio_home_5' if side == 'home' else 'xg_ratio_away_5'
        m &= (data[xg_col].fillna(0) >= p['xg'])
    if side == 'home' and p['elo'] > 0:
        m &= (data['elo_diff'].fillna(0) >= p['elo'])
    if side == 'away' and p['elo'] < 0:
        m &= (data['elo_diff'].fillna(0) <= p['elo'])
    if p['form'] > 0:
        fm_col = 'home_pts_5' if side == 'home' else 'away_pts_5'
        m &= (data[fm_col].fillna(0) >= p['form'])
    if p['mkt'] > 0:
        mk_col = 'mkt_home_prob' if side == 'home' else 'mkt_away_prob'
        m &= (data[mk_col].fillna(0) >= p['mkt'])
    return data[m]

# ── Load confirmed niches ─────────────────────────────────────────────────────
print('Loading confirmed niches...')
cn = pd.read_excel(CONFIRMED_SRC, sheet_name='OOS Confirmed')
# Drop league separator rows (Volume is NaN)
cn = cn[cn['Volume'].notna()].copy()
print(f'  {len(cn)} niches to simulate')

# ── Simulate each niche ───────────────────────────────────────────────────────
results = []

for _, row in cn.iterrows():
    league = row['League']
    side   = row['Side']
    niche  = row['Niche']
    vol    = row['Volume']

    try:
        p = parse_niche(niche, side)
    except Exception as e:
        continue

    bets = apply_mask(oos, side, p, league)
    if len(bets) == 0:
        continue

    wc = 'home_win' if side == 'home' else 'away_win'
    oc = 'home_odds_val' if side == 'home' else 'away_odds_val'

    bank      = INITIAL_BANK
    peak_bank = INITIAL_BANK
    max_dd    = 0.0
    profits   = []

    for _, bet in bets.sort_values('date').iterrows():
        won    = bool(bet[wc])
        odds   = float(bet[oc])
        profit = FLAT_STAKE * (odds - 1) if won else -FLAT_STAKE
        bank  += profit
        profits.append(profit)
        peak_bank = max(peak_bank, bank)
        dd = (peak_bank - bank) / peak_bank
        max_dd = max(max_dd, dd)

    n_bets   = len(profits)
    n_wins   = sum(1 for p in profits if p > 0)
    total_pl = bank - INITIAL_BANK
    roi_pct  = total_pl / (n_bets * FLAT_STAKE) * 100
    wr       = n_wins / n_bets

    results.append({
        'League':      league,
        'Volume':      vol,
        'Niche':       niche,
        'Side':        side,
        'N bets':      n_bets,
        'Wins':        n_wins,
        'WR%':         wr,
        'Avg Odds':    bets[oc].mean(),
        'Final Bank':  round(bank, 2),
        'P&L':         round(total_pl, 2),
        'ROI%':        round(roi_pct, 2),
        'Max DD%':     round(max_dd * 100, 2),
    })

res = pd.DataFrame(results)
res = res.sort_values('ROI%', ascending=False)

print(f'\nSimulated: {len(res)} niches')
print(f'  Profitable (ROI>0): {(res["ROI%"]>0).sum()}')
print(f'  Avg ROI: {res["ROI%"].mean():.1f}%')
print(f'  Median ROI: {res["ROI%"].median():.1f}%')
print()
print('Top 10 niches:')
print(res[['League','Volume','Niche','N bets','WR%','ROI%','Final Bank']].head(10).to_string(index=False))
print()
print('Per league (profitable niches):')
for lg, g in res[res['ROI%']>0].groupby('League'):
    print(f"  {lg:<25s} {len(g):3d} profitable  avg ROI={g['ROI%'].mean():.1f}%  best={g['ROI%'].max():.1f}%")

# ── Excel ─────────────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = 'Simulation'

fill_hdr    = PatternFill('solid', fgColor='1F3864')
fill_high   = PatternFill('solid', fgColor='E2EFDA')
fill_low    = PatternFill('solid', fgColor='FFF2CC')
fill_lg     = PatternFill('solid', fgColor='D6E4F0')
fill_pos    = PatternFill('solid', fgColor='C6EFCE')
fill_strong = PatternFill('solid', fgColor='70AD47')
fill_neg    = PatternFill('solid', fgColor='FFCCCC')

COLS = ['League','Volume','Niche','Side','N bets','Wins','WR%','Avg Odds',
        'Final Bank ($)','P&L ($)','ROI%','Max DD%']

for ci, col in enumerate(COLS, 1):
    c = ws.cell(1, ci, col)
    c.fill = fill_hdr
    c.font = Font(color='FFFFFF', bold=True, size=10)
    c.alignment = Alignment(horizontal='center', wrap_text=True)

thin   = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

row_idx = 2
prev_league = None

for _, r in res.iterrows():
    if r['League'] != prev_league:
        c = ws.cell(row_idx, 1, r['League'])
        c.fill = fill_lg
        c.font = Font(bold=True, size=11)
        ws.merge_cells(start_row=row_idx, start_column=1,
                       end_row=row_idx, end_column=len(COLS))
        row_idx += 1
        prev_league = r['League']

    fill = fill_high if r['Volume'] == 'High' else fill_low
    roi  = r['ROI%']

    values = [r['League'], r['Volume'], r['Niche'], r['Side'],
              r['N bets'], r['Wins'], r['WR%'], r['Avg Odds'],
              r['Final Bank'], r['P&L'], roi/100, r['Max DD%']/100]

    for ci, val in enumerate(values, 1):
        c = ws.cell(row_idx, ci, val)
        c.fill = fill
        c.border = border
        c.alignment = Alignment(horizontal='center')
        if ci == 3:
            c.alignment = Alignment(horizontal='left')
        if ci == 7:
            c.number_format = '0.0%'
        if ci == 11:
            c.number_format = '+0.0%;-0.0%;0%'
            if roi >= 20:
                c.fill = fill_strong
                c.font = Font(bold=True)
            elif roi > 0:
                c.fill = fill_pos
            else:
                c.fill = fill_neg
        if ci == 12:
            c.number_format = '0.0%'
        if ci == 8:
            c.number_format = '0.00'
        if ci in (9, 10):
            c.number_format = '$#,##0.00'

    row_idx += 1

widths = [16, 7, 42, 6, 7, 6, 7, 9, 13, 11, 8, 9]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}1'

# ── Summary by league ──────────────────────────────────────────────────────────
ws2 = wb.create_sheet('By League')
s_cols = ['League','Total','Profitable','% Profitable','Avg ROI%','Best ROI%','Worst ROI%']
for ci, col in enumerate(s_cols, 1):
    c = ws2.cell(1, ci, col)
    c.fill = fill_hdr
    c.font = Font(color='FFFFFF', bold=True)
    c.alignment = Alignment(horizontal='center')

for ri, (lg, g) in enumerate(res.groupby('League'), 2):
    total = len(g)
    prof  = (g['ROI%'] > 0).sum()
    ws2.cell(ri,1,lg)
    ws2.cell(ri,2,total)
    ws2.cell(ri,3,prof)
    c = ws2.cell(ri,4, prof/total)
    c.number_format = '0%'
    c = ws2.cell(ri,5, g['ROI%'].mean()/100)
    c.number_format = '+0.0%;-0.0%;0%'
    c = ws2.cell(ri,6, g['ROI%'].max()/100)
    c.number_format = '+0.0%;-0.0%;0%'
    c = ws2.cell(ri,7, g['ROI%'].min()/100)
    c.number_format = '+0.0%;-0.0%;0%'

for col in range(1,8):
    ws2.column_dimensions[get_column_letter(col)].width = 18

wb.save(OUT)
print(f'\nSaved: {OUT}')
print('✅ Done!')
