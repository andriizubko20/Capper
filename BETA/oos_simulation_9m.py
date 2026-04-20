import warnings; warnings.filterwarnings('ignore')
import pandas as pd
import numpy as np
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys
sys.path.insert(0, '.')
from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix

OUT          = '/app/BETA/oos_simulation_9m.xlsx'
INITIAL_BANK = 1000.0
FLAT_STAKE   = 50.0   # 5% of initial

# Folds 3-5 for selection (Fold 6 + OOS = independent test)
FOLDS = [
    ('F3', pd.Timestamp('2023-08-01'), pd.Timestamp('2024-10-31'),
           pd.Timestamp('2024-11-01'), pd.Timestamp('2025-01-31')),
    ('F4', pd.Timestamp('2023-08-01'), pd.Timestamp('2025-01-31'),
           pd.Timestamp('2025-02-01'), pd.Timestamp('2025-04-30')),
    ('F5', pd.Timestamp('2023-08-01'), pd.Timestamp('2025-04-30'),
           pd.Timestamp('2025-05-01'), pd.Timestamp('2025-07-31')),
]
MIN_STABLE_FOLDS = 2   # >=2/3 folds profitable
MIN_TRAIN_N      = 5
MIN_TEST_N       = 1

OOS_START = pd.Timestamp('2025-08-01')   # Fold6 start — never seen during selection
OOS_END   = pd.Timestamp('2026-04-11')

LEAGUES = [
    'Premier League','Bundesliga','Serie A','La Liga','Ligue 1',
    'Primeira Liga','Serie B','Eredivisie','Jupiler Pro League','Champions League'
]
ODDS_RANGES = [(1.30,1.55),(1.55,1.80),(1.70,2.00),(1.80,2.20),(2.00,2.50),(2.20,2.80),(2.50,3.50)]
XG  = [0.0, 1.0, 1.2, 1.5, 1.8]
EH  = [0,  30,  75, 150]
EA  = [0, -30, -75, -150]
FT  = [0.0, 1.5, 1.8, 2.2]
MH  = [0.0, 0.45, 0.50, 0.55]
MA  = [0.0, 0.35, 0.40, 0.45]

print('Loading data...')
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result'] == 'H').astype(int)
df['away_win']  = (df['result'] == 'A').astype(int)

fold_slices = {
    n: {
        'train': df[(df['date'] >= ts) & (df['date'] <= te)],
        'test':  df[(df['date'] >= tes) & (df['date'] <= tee)],
    }
    for n, ts, te, tes, tee in FOLDS
}
oos = df[(df['date'] >= OOS_START) & (df['date'] <= OOS_END)].copy()
print(f'OOS matches: {len(oos)}  ({OOS_START.date()} – {OOS_END.date()})')

def apply_mask(data, side, lo, hi, xg, elo, fm, mk, league):
    oc = 'home_odds_val' if side == 'home' else 'away_odds_val'
    m = (
        (data['league_name'] == league) &
        (data[oc] >= lo) & (data[oc] < hi)
    )
    if xg > 0:
        xg_col = 'xg_ratio_home_5' if side == 'home' else 'xg_ratio_away_5'
        m &= data[xg_col].fillna(0) >= xg
    if side == 'home' and elo > 0:
        m &= data['elo_diff'].fillna(0) >= elo
    if side == 'away' and elo < 0:
        m &= data['elo_diff'].fillna(0) <= elo
    if fm > 0:
        fm_col = 'home_pts_5' if side == 'home' else 'away_pts_5'
        m &= data[fm_col].fillna(0) >= fm
    if mk > 0:
        mk_col = 'mkt_home_prob' if side == 'home' else 'mkt_away_prob'
        m &= data[mk_col].fillna(0) >= mk
    return data[m]

def niche_label(side, lo, hi, xg, elo, fm, mk):
    s = f'{side}[{lo},{hi})'
    if xg  > 0: s += f' xg>={xg}'
    if elo != 0: s += f' elo>={elo}' if elo > 0 else f' elo<={elo}'
    if fm  > 0: s += f' form>={fm}'
    if mk  > 0: s += f' mkt>={mk}'
    return s

def simulate(bets_df, side):
    wc = 'home_win' if side == 'home' else 'away_win'
    oc = 'home_odds_val' if side == 'home' else 'away_odds_val'
    bank = INITIAL_BANK
    peak = INITIAL_BANK
    max_dd = 0.0
    profits = []
    for _, bet in bets_df.sort_values('date').iterrows():
        won    = bool(bet[wc])
        odds   = float(bet[oc])
        profit = FLAT_STAKE * (odds - 1) if won else -FLAT_STAKE
        bank  += profit
        profits.append(profit)
        peak   = max(peak, bank)
        dd     = (peak - bank) / peak if peak > 0 else 0
        max_dd = max(max_dd, dd)
    return bank, profits, max_dd

print('Running niche discovery on Folds 3-5...')
confirmed = []
total_checked = 0

for side in ('home', 'away'):
    et = EH if side == 'home' else EA
    mt = MH if side == 'home' else MA
    wc = 'home_win' if side == 'home' else 'away_win'
    oc = 'home_odds_val' if side == 'home' else 'away_odds_val'

    for lo, hi in ODDS_RANGES:
        for xg, elo, fm, mk in product(XG, et, FT, mt):
            for lg in LEAGUES:
                total_checked += 1
                fold_results = []
                for fn, _, _, _, _ in FOLDS:
                    tr = apply_mask(fold_slices[fn]['train'], side, lo, hi, xg, elo, fm, mk, lg)
                    te = apply_mask(fold_slices[fn]['test'],  side, lo, hi, xg, elo, fm, mk, lg)
                    if len(tr) < MIN_TRAIN_N or len(te) < MIN_TEST_N:
                        continue
                    te_roi = (te[wc] * (te[oc] - 1) - (1 - te[wc])).mean()
                    fold_results.append(te_roi)

                if len(fold_results) < MIN_STABLE_FOLDS:
                    continue
                n_prof = sum(1 for r in fold_results if r > 0)
                if n_prof < MIN_STABLE_FOLDS:
                    continue

                # Simulate on OOS (9 months)
                oos_bets = apply_mask(oos, side, lo, hi, xg, elo, fm, mk, lg)
                if len(oos_bets) == 0:
                    continue

                avg_te_n = np.mean([
                    len(apply_mask(fold_slices[fn]['test'], side, lo, hi, xg, elo, fm, mk, lg))
                    for fn, *_ in FOLDS
                ])
                volume = 'High' if avg_te_n >= 8 else 'Low'

                bank, profits, max_dd = simulate(oos_bets, side)
                n_bets  = len(profits)
                n_wins  = sum(1 for p in profits if p > 0)
                total_pl = bank - INITIAL_BANK
                roi_pct  = total_pl / (n_bets * FLAT_STAKE) * 100

                confirmed.append({
                    'League':     lg,
                    'Volume':     volume,
                    'Niche':      niche_label(side, lo, hi, xg, elo, fm, mk),
                    'Side':       side,
                    'Folds OK':   f'{n_prof}/{len(fold_results)}',
                    'Avg Te n':   round(avg_te_n, 1),
                    'OOS n':      n_bets,
                    'OOS Wins':   n_wins,
                    'OOS WR%':    n_wins / n_bets if n_bets else 0,
                    'OOS ROI%':   round(roi_pct, 2),
                    'Final Bank': round(bank, 2),
                    'P&L':        round(total_pl, 2),
                    'Max DD%':    round(max_dd * 100, 2),
                })

print(f'Checked {total_checked} combinations. Found {len(confirmed)} niches with >=2/3 stable folds.')

res = pd.DataFrame(confirmed)

# Dedup: remove subsets (niche A is subset of B if A has more filters and same/lower OOS n)
# Simple approach: sort by specificity (number of active filters), keep broader if OOS ROI better
res['n_filters'] = res['Niche'].apply(lambda x: x.count('>=') + x.count('<='))
res = res.sort_values(['League','Side','OOS ROI%'], ascending=[True, True, False])

# Keep only OOS profitable for primary output; show all in separate sheet
res_all       = res.copy()
res_profitable = res[res['OOS ROI%'] > 0].copy()

print(f'\nTotal niches: {len(res_all)}')
print(f'OOS profitable: {len(res_profitable)}')
print(f"  High volume: {(res_profitable['Volume']=='High').sum()}")
print(f"  Low volume:  {(res_profitable['Volume']=='Low').sum()}")
print()
print('Per league (OOS profitable):')
for lg, g in res_profitable.groupby('League'):
    h = (g['Volume']=='High').sum()
    l = (g['Volume']=='Low').sum()
    print(f"  {lg:<25s} High={h:3d}  Low={l:3d}  avg ROI={g['OOS ROI%'].mean():.1f}%  best={g['OOS ROI%'].max():.1f}%")

# ── Excel ─────────────────────────────────────────────────────────────────────
wb = Workbook()

fill_hdr    = PatternFill('solid', fgColor='1F3864')
fill_high   = PatternFill('solid', fgColor='E2EFDA')
fill_low    = PatternFill('solid', fgColor='FFF2CC')
fill_lg     = PatternFill('solid', fgColor='D6E4F0')
fill_pos    = PatternFill('solid', fgColor='C6EFCE')
fill_strong = PatternFill('solid', fgColor='70AD47')
fill_neg    = PatternFill('solid', fgColor='FFCCCC')
thin        = Side(style='thin', color='CCCCCC')
border      = Border(left=thin, right=thin, top=thin, bottom=thin)

COLS = ['League','Volume','Niche','Side','Folds OK','Avg Te n',
        'OOS n','OOS Wins','OOS WR%','OOS ROI%','Final Bank ($)','P&L ($)','Max DD%']

def write_sheet(ws, data, title):
    ws.title = title
    for ci, col in enumerate(COLS, 1):
        c = ws.cell(1, ci, col)
        c.fill = fill_hdr
        c.font = Font(color='FFFFFF', bold=True, size=10)
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    row_idx = 2
    prev_lg = None
    for _, r in data.sort_values(['League','Volume','OOS ROI%'], ascending=[True,True,False]).iterrows():
        if r['League'] != prev_lg:
            c = ws.cell(row_idx, 1, r['League'])
            c.fill = fill_lg
            c.font = Font(bold=True, size=11)
            ws.merge_cells(start_row=row_idx, start_column=1,
                           end_row=row_idx, end_column=len(COLS))
            row_idx += 1
            prev_lg = r['League']

        fill = fill_high if r['Volume'] == 'High' else fill_low
        roi  = r['OOS ROI%']

        values = [r['League'], r['Volume'], r['Niche'], r['Side'],
                  r['Folds OK'], r['Avg Te n'],
                  int(r['OOS n']), int(r['OOS Wins']), r['OOS WR%'],
                  roi/100, r['Final Bank'], r['P&L'], r['Max DD%']/100]

        for ci, val in enumerate(values, 1):
            c = ws.cell(row_idx, ci, val)
            c.fill = fill
            c.border = border
            c.alignment = Alignment(horizontal='center')
            if ci == 3: c.alignment = Alignment(horizontal='left')
            if ci == 9:  c.number_format = '0.0%'
            if ci == 10:
                c.number_format = '+0.0%;-0.0%;0%'
                if roi >= 20:   c.fill = fill_strong; c.font = Font(bold=True)
                elif roi > 0:   c.fill = fill_pos
                else:           c.fill = fill_neg
            if ci == 13: c.number_format = '0.0%'
            if ci in (11, 12): c.number_format = '$#,##0'

        row_idx += 1

    widths = [16, 7, 44, 6, 8, 8, 7, 8, 8, 9, 13, 11, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS))}1'

ws1 = wb.active
write_sheet(ws1, res_profitable, 'OOS Profitable')

ws2 = wb.create_sheet('All Niches')
write_sheet(ws2, res_all, 'All Niches')

# Summary
ws3 = wb.create_sheet('By League')
s_cols = ['League','Total','OOS+','% OOS+','High+','Low+','Avg ROI%','Best ROI%','Avg Final Bank']
for ci, col in enumerate(s_cols, 1):
    c = ws3.cell(1, ci, col)
    c.fill = fill_hdr
    c.font = Font(color='FFFFFF', bold=True)
    c.alignment = Alignment(horizontal='center')

for ri, (lg, g) in enumerate(res_all.groupby('League'), 2):
    prof = g[g['OOS ROI%'] > 0]
    ws3.cell(ri, 1, lg)
    ws3.cell(ri, 2, len(g))
    ws3.cell(ri, 3, len(prof))
    c = ws3.cell(ri, 4, len(prof)/len(g) if len(g) else 0)
    c.number_format = '0%'
    ws3.cell(ri, 5, (prof['Volume']=='High').sum())
    ws3.cell(ri, 6, (prof['Volume']=='Low').sum())
    for ci, val in [(7, prof['OOS ROI%'].mean()/100 if len(prof) else 0),
                    (8, g['OOS ROI%'].max()/100),
                    (9, prof['Final Bank'].mean() if len(prof) else INITIAL_BANK)]:
        c = ws3.cell(ri, ci, val)
        c.number_format = '+0.0%;-0.0%;0%' if ci in (7,8) else '$#,##0'
        if ci in (7,8):
            c.fill = fill_strong if val>=0.2 else (fill_pos if val>0 else fill_neg)

for col in range(1,10):
    ws3.column_dimensions[get_column_letter(col)].width = 18

wb.save(OUT)
print(f'\nSaved: {OUT}')
print('✅ Done!')
