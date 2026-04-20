"""
BETA/export_working_niches.py
Експортує підтверджені ніші в Excel:
- WF avg_roi>=10%, win%>=60%, n_windows>=4 (discovery до 2025-10)
- OOS n>=15, ROI>0 (holdout 2025-11 → 2026-04)
- Dedup: тільки незалежні ніші
"""
import sys, os, warnings
warnings.filterwarnings('ignore')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import numpy as np
import pandas as pd
from itertools import product
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix

WF_END    = pd.Timestamp('2025-10-31')
OOS_START = pd.Timestamp('2025-11-01')
MIN_OOS_N = 15

print("Loading data...")
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result'] == 'H').astype(int)
df['away_win']  = (df['result'] == 'A').astype(int)

disc_df = df[df['date'] <= WF_END].copy()
oos_df  = df[df['date'] >= OOS_START].copy()

WINDOWS = []
cur = pd.Timestamp('2023-08-01')
while cur <= WF_END:
    w_end = cur + pd.DateOffset(months=3) - pd.Timedelta(days=1)
    if w_end > WF_END: w_end = WF_END
    WINDOWS.append((cur.strftime('%Y-%m'), cur, w_end))
    cur += pd.DateOffset(months=3)
WIN_LABELS = [w[0] for w in WINDOWS]

LEAGUES = [
    'Premier League','Bundesliga','Serie A','La Liga','Ligue 1',
    'Primeira Liga','Serie B','Eredivisie','Jupiler Pro League','Champions League'
]
LSHORT = {
    'Premier League':'EPL','Bundesliga':'Bundesliga','Serie A':'Serie A',
    'La Liga':'La Liga','Ligue 1':'Ligue 1','Primeira Liga':'Portugal',
    'Serie B':'Serie B','Eredivisie':'Eredivisie',
    'Jupiler Pro League':'Jupiler','Champions League':'UCL'
}

ODDS_RANGES = [(1.30,1.55),(1.55,1.80),(1.70,2.00),(1.80,2.20),(2.00,2.50),(2.20,2.80),(2.50,3.50)]
XG=[0.0,1.0,1.2,1.5,1.8]; EH=[0,30,75,150]; EA=[0,-30,-75,-150]
FT=[0.0,1.5,1.8,2.2]; MH=[0.0,0.45,0.50,0.55]; MA=[0.0,0.35,0.40,0.45]
win_slices = {wlbl: disc_df[(disc_df['date']>=ws)&(disc_df['date']<=we)]
              for wlbl,ws,we in WINDOWS}

def apply_mask(d, side, lo, hi, xg_t, elo_t, form_t, mkt_t):
    oc  = 'home_odds_val'   if side=='home' else 'away_odds_val'
    xgc = 'xg_ratio_home_5' if side=='home' else 'xg_ratio_away_5'
    fc  = 'home_pts_5'      if side=='home' else 'away_pts_5'
    mc  = 'mkt_home_prob'   if side=='home' else 'mkt_away_prob'
    m = (d[oc]>=lo)&(d[oc]<hi)
    if xg_t>0:                       m&=(d[xgc].fillna(0)>=xg_t)
    if side=='home' and elo_t>0:     m&=(d['elo_diff'].fillna(0)>=elo_t)
    if side=='away' and elo_t<0:     m&=(d['elo_diff'].fillna(0)<=elo_t)
    if form_t>0:                     m&=(d[fc].fillna(0)>=form_t)
    if mkt_t>0:                      m&=(d[mc].fillna(0)>=mkt_t)
    return d[m]

def is_subset(a, b, side):
    if b['xg_t']<a['xg_t'] or b['form_t']<a['form_t'] or b['mkt_t']<a['mkt_t']: return False
    return b['elo_t']>=a['elo_t'] if side=='home' else b['elo_t']<=a['elo_t']

def dedup(filters):
    filters = sorted(filters, key=lambda x: x['avg_roi'], reverse=True)
    kept = []
    for f in filters:
        if not any(is_subset(k,f,f['side']) and k['avg_roi']>=f['avg_roi'] for k in kept):
            kept.append(f)
    return kept

print("Sweeping...")
raw = {lg: [] for lg in LEAGUES}

for side in ('home','away'):
    et=EH if side=='home' else EA; mt=MH if side=='home' else MA
    wc='home_win' if side=='home' else 'away_win'
    oc='home_odds_val' if side=='home' else 'away_odds_val'
    for lo,hi in ODDS_RANGES:
        oos_pre = apply_mask(oos_df, side, lo, hi, 0, 0, 0, 0)
        for xg_t,elo_t,form_t,mkt_t in product(XG,et,FT,mt):
            parts=[f'{side}[{lo},{hi})']
            if xg_t>0: parts.append(f'xg>={xg_t}')
            if side=='home' and elo_t>0: parts.append(f'elo>={elo_t}')
            if side=='away' and elo_t<0: parts.append(f'elo<={elo_t}')
            if form_t>0: parts.append(f'form>={form_t}')
            if mkt_t>0: parts.append(f'mkt>={mkt_t}')
            label=' '.join(parts)
            for lg in LEAGUES:
                wf_rois,wf_wrs,wf_aos=[],[],[]
                for wlbl,_,_ in WINDOWS:
                    sl=apply_mask(win_slices[wlbl],side,lo,hi,xg_t,elo_t,form_t,mkt_t)
                    sl=sl[sl['league_name']==lg]
                    if len(sl)<3: continue
                    wf_rois.append((sl[wc]*(sl[oc]-1)-(1-sl[wc])).mean()*100)
                    wf_wrs.append(sl[wc].mean())
                    wf_aos.append(sl[oc].mean())
                if len(wf_rois)<4: continue
                avg_roi=np.mean(wf_rois); win_pct=sum(1 for r in wf_rois if r>0)/len(wf_rois)*100
                if avg_roi<10 or win_pct<60: continue
                sl_oos=apply_mask(oos_pre,side,lo,hi,xg_t,elo_t,form_t,mkt_t)
                sl_oos=sl_oos[sl_oos['league_name']==lg]
                if len(sl_oos)<MIN_OOS_N: continue
                oos_roi=(sl_oos[wc]*(sl_oos[oc]-1)-(1-sl_oos[wc])).mean()*100
                if oos_roi<=0: continue
                # per-window detail
                win_detail={wlbl:None for wlbl in WIN_LABELS}
                for wlbl,_,_ in WINDOWS:
                    sl=apply_mask(win_slices[wlbl],side,lo,hi,xg_t,elo_t,form_t,mkt_t)
                    sl=sl[sl['league_name']==lg]
                    if len(sl)>=3:
                        r=(sl[wc]*(sl[oc]-1)-(1-sl[wc])).mean()*100
                        win_detail[wlbl]={'roi':round(r,1),'n':len(sl),'wr':round(sl[wc].mean()*100,1)}
                raw[lg].append({
                    'label':label,'side':side,'lo':lo,'hi':hi,
                    'xg_t':xg_t,'elo_t':elo_t,'form_t':form_t,'mkt_t':mkt_t,
                    'avg_roi':round(avg_roi,1),
                    'avg_wr':round(np.mean(wf_wrs)*100,1),
                    'avg_odds':round(np.mean(wf_aos),2),
                    'win_pct':round(win_pct,1),
                    'n_windows':len(wf_rois),
                    'oos_n':len(sl_oos),
                    'oos_roi':round(oos_roi,1),
                    'oos_wr':round(sl_oos[wc].mean()*100,1),
                    'win_detail':win_detail,
                })

catalog={lg:[] for lg in LEAGUES}
for lg in LEAGUES:
    groups={}
    for f in raw[lg]: groups.setdefault((f['side'],f['lo'],f['hi']),[]).append(f)
    for group in groups.values(): catalog[lg].extend(dedup(group))

total=sum(len(v) for v in catalog.values())
print(f"After dedup: {total} niches")
for lg in LEAGUES:
    if catalog[lg]: print(f"  {LSHORT[lg]}: {len(catalog[lg])}")

# ── Excel ─────────────────────────────────────────────────────────────────────
GREEN_D = PatternFill('solid', start_color='1E8449')
GREEN_L = PatternFill('solid', start_color='D5F5E3')
RED_L   = PatternFill('solid', start_color='FADBD8')
GREY    = PatternFill('solid', start_color='F2F3F4')
HDR_BG  = PatternFill('solid', start_color='1B2631')
OOS_BG  = PatternFill('solid', start_color='154360')
SUBHDR  = PatternFill('solid', start_color='5D6D7E')
YELLOW  = PatternFill('solid', start_color='FEF9E7')

HDR_F  = Font(name='Arial', bold=True, color='FFFFFF', size=9)
BODY_F = Font(name='Arial', size=9)
BOLD_F = Font(name='Arial', bold=True, size=9)
C = Alignment(horizontal='center', vertical='center')
L = Alignment(horizontal='left',   vertical='center')

def tb():
    s=Side(style='thin',color='CCCCCC')
    return Border(left=s,right=s,top=s,bottom=s)
def hc(c,v,fill=HDR_BG):
    c.value=v;c.font=HDR_F;c.fill=fill;c.alignment=C;c.border=tb()
def bc(c,v,align=C,fill=None,bold=False):
    c.value=v;c.font=Font(name='Arial',size=9,bold=bold)
    c.alignment=align;c.border=tb()
    if fill:c.fill=fill

wb=Workbook(); wb.remove(wb.active)

# ── SUMMARY sheet ──────────────────────────────────────────────────────────────
ws_s=wb.create_sheet('SUMMARY')
SHDR=['Ліга','Ніш','Avg WF ROI%','Avg OOS ROI%','Avg OOS WR%','Найкраща ніша (OOS ROI)','OOS n']
for ci,h in enumerate(SHDR,1): hc(ws_s.cell(1,ci),h)
sr=2
for lg in LEAGUES:
    items=catalog[lg]
    if not items: continue
    lname=LSHORT[lg]
    best=max(items,key=lambda x:x['oos_roi'])
    row=[lname,len(items),
         round(np.mean([i['avg_roi'] for i in items]),1),
         round(np.mean([i['oos_roi'] for i in items]),1),
         round(np.mean([i['oos_wr']  for i in items]),1),
         best['label'],best['oos_n']]
    fmts=[None,None,'+0.0%','+0.0%','0.0%',None,None]
    for ci,(v,fmt) in enumerate(zip(row,fmts),1):
        c=ws_s.cell(sr,ci)
        bc(c,v/100 if fmt and isinstance(v,(int,float)) else v, L if ci in(1,6) else C)
        if fmt and isinstance(v,(int,float)): c.number_format=fmt
        if ci==3: c.fill=GREEN_L if row[2]>0 else RED_L
        if ci==4: c.fill=GREEN_L if row[3]>0 else RED_L
    sr+=1
ws_s.column_dimensions['A'].width=12
ws_s.column_dimensions['B'].width=6
for ci in range(3,6): ws_s.column_dimensions[get_column_letter(ci)].width=13
ws_s.column_dimensions['F'].width=55
ws_s.column_dimensions['G'].width=8
ws_s.freeze_panes='A2'

# ── Per-league sheets ──────────────────────────────────────────────────────────
STATIC=['Ніша','Side','WF Avg ROI%','WF WR%','WF Win%','W+/tot',
        'OOS n','OOS WR%','OOS ROI%']
N_S=len(STATIC)

for lg in LEAGUES:
    items=catalog[lg]
    if not items: continue
    lname=LSHORT[lg]
    ws=wb.create_sheet(lname)
    total_c=N_S+len(WINDOWS)*2

    # Row 1: static headers
    for ci,h in enumerate(STATIC,1): hc(ws.cell(1,ci),h)
    # Row 1: WF window group headers
    for wi,(wlbl,_,_) in enumerate(WINDOWS):
        cs=N_S+wi*2+1
        ws.merge_cells(start_row=1,start_column=cs,end_row=1,end_column=cs+1)
        hc(ws.cell(1,cs),wlbl)
    # Row 2: sub-headers
    for ci in range(1,N_S+1): hc(ws.cell(2,ci),'',fill=SUBHDR)
    for wi in range(len(WINDOWS)):
        cs=N_S+wi*2+1
        hc(ws.cell(2,cs),'ROI%',fill=SUBHDR)
        hc(ws.cell(2,cs+1),'n',fill=SUBHDR)

    items_sorted=sorted(items,key=lambda x:-x['oos_roi'])
    for er,niche in enumerate(items_sorted,3):
        row_fill=GREEN_L if niche['oos_roi']>=10 else YELLOW if niche['oos_roi']>0 else RED_L
        static_vals=[
            niche['label'],niche['side'],
            niche['avg_roi']/100,niche['avg_wr']/100,
            niche['win_pct']/100,
            f"{niche['n_windows']}/{len(WINDOWS)}",
            niche['oos_n'],niche['oos_wr']/100,niche['oos_roi']/100
        ]
        sfmts=[None,None,'+0.0%;-0.0%','0.0%','0%',None,None,'0.0%','+0.0%;-0.0%']
        for ci,(v,fmt) in enumerate(zip(static_vals,sfmts),1):
            c=ws.cell(er,ci)
            bc(c,v,L if ci in(1,2) else C,row_fill)
            if fmt and isinstance(v,(int,float)): c.number_format=fmt
            if ci==3:
                c.fill=GREEN_L if niche['avg_roi']>0 else RED_L
                c.font=Font(name='Arial',size=9,bold=niche['avg_roi']>15)
            if ci==9:
                c.fill=GREEN_L if niche['oos_roi']>0 else RED_L
                c.font=Font(name='Arial',size=9,bold=True)

        for wi,(wlbl,_,_) in enumerate(WINDOWS):
            cs=N_S+wi*2+1
            wd=niche['win_detail'].get(wlbl)
            if wd:
                cr=ws.cell(er,cs)
                cr.value=wd['roi']/100; cr.number_format='+0.0%;-0.0%'
                cr.font=Font(name='Arial',size=9,bold=abs(wd['roi'])>20)
                cr.alignment=C; cr.border=tb()
                cr.fill=GREEN_L if wd['roi']>0 else RED_L
                cn=ws.cell(er,cs+1); bc(cn,wd['n'],fill=row_fill)
            else:
                bc(ws.cell(er,cs),'—',fill=GREY)
                bc(ws.cell(er,cs+1),'—',fill=GREY)

    ws.column_dimensions['A'].width=52
    ws.column_dimensions['B'].width=6
    for ci in range(3,N_S+1): ws.column_dimensions[get_column_letter(ci)].width=10
    for wi in range(len(WINDOWS)):
        cs=N_S+wi*2+1
        ws.column_dimensions[get_column_letter(cs)].width=8
        ws.column_dimensions[get_column_letter(cs+1)].width=4
    ws.freeze_panes='A3'
    ws.row_dimensions[1].height=20; ws.row_dimensions[2].height=18

outpath=os.path.join(os.path.dirname(__file__),'working_niches.xlsx')
wb.save(outpath)
print(f"\nSaved: {outpath}  ({os.path.getsize(outpath)//1024}KB)")
print("Done!")
