"""
BETA/expanding_wf_v3.py

Два каталоги:
  High volume: avg test n >= 8 матчів на квартал (надійно)
  Low volume:  avg test n < 8  (потенційні gems, менша впевненість)

Без мінімального n — тестуємо всі 44,800 комбінацій.
Стабільність: >= 3/4 тестових фолдів прибуткові.
Final OOS: 2025-11 → 2026-04.
"""
import sys, os, warnings
warnings.filterwarnings('ignore')
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

import numpy as np
import pandas as pd
from itertools import product
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as BdrSide
from openpyxl.utils import get_column_letter

from BETA.data.extract import load_all
from BETA.data.features import build_feature_matrix

# ── Config ────────────────────────────────────────────────────────────────────
MIN_TRAIN_N      = 1     # мінімум матчів у тренуванні (майже без обмежень)
MIN_TEST_N       = 1     # мінімум матчів у тесті
MIN_STABLE_FOLDS = 3     # прибуткові в >= 3 з 4 фолдів
HIGH_VOL_N       = 8     # avg test n >= 8 → high volume
MIN_OOS_N        = 1     # мінімум матчів у фінальному OOS

FOLDS = [
    ('Fold3', pd.Timestamp('2023-08-01'), pd.Timestamp('2024-10-31'),
              pd.Timestamp('2024-11-01'), pd.Timestamp('2025-01-31')),
    ('Fold4', pd.Timestamp('2023-08-01'), pd.Timestamp('2025-01-31'),
              pd.Timestamp('2025-02-01'), pd.Timestamp('2025-04-30')),
    ('Fold5', pd.Timestamp('2023-08-01'), pd.Timestamp('2025-04-30'),
              pd.Timestamp('2025-05-01'), pd.Timestamp('2025-07-31')),
    ('Fold6', pd.Timestamp('2023-08-01'), pd.Timestamp('2025-07-31'),
              pd.Timestamp('2025-08-01'), pd.Timestamp('2025-10-31')),
]
FOLD_LABELS = [f[0] for f in FOLDS]
N_FOLDS = len(FOLDS)
FINAL_OOS_START = pd.Timestamp('2025-11-01')

# ── Load ──────────────────────────────────────────────────────────────────────
print("Loading data...")
matches, stats, odds_data, injuries = load_all()
df = build_feature_matrix(matches, stats, odds_data, injuries)
df = df[df['home_odds_val'].notna() & df['away_odds_val'].notna()].copy()
df['home_win'] = (df['result'] == 'H').astype(int)
df['away_win']  = (df['result'] == 'A').astype(int)

oos_df = df[df['date'] >= FINAL_OOS_START].copy()
print(f"Total: {len(df)} | OOS: {len(oos_df)} ({oos_df['date'].min().date()} – {oos_df['date'].max().date()})")

LEAGUES = [
    'Premier League','Bundesliga','Serie A','La Liga','Ligue 1',
    'Primeira Liga','Serie B','Eredivisie','Jupiler Pro League','Champions League'
]
LSHORT = {
    'Premier League':'EPL','Bundesliga':'Bundesliga','Serie A':'Serie A',
    'La Liga':'La Liga','Ligue 1':'Ligue 1','Primeira Liga':'Portugal',
    'Serie B':'Serie B','Eredivisie':'Eredivisie',
    'Jupiler Pro League':'Jupiler','Champions League':'UCL'
}

fold_slices = {}
for fname, tr_s, tr_e, te_s, te_e in FOLDS:
    fold_slices[fname] = {
        'train': df[(df['date'] >= tr_s) & (df['date'] <= tr_e)],
        'test':  df[(df['date'] >= te_s) & (df['date'] <= te_e)],
    }

ODDS_RANGES = [(1.30,1.55),(1.55,1.80),(1.70,2.00),(1.80,2.20),(2.00,2.50),(2.20,2.80),(2.50,3.50)]
XG_THRS    = [0.0,1.0,1.2,1.5,1.8]
ELO_H_THRS = [0,30,75,150]
ELO_A_THRS = [0,-30,-75,-150]
FORM_THRS  = [0.0,1.5,1.8,2.2]
MKT_H_THRS = [0.0,0.45,0.50,0.55]
MKT_A_THRS = [0.0,0.35,0.40,0.45]

print(f"Total combinations: {2*len(ODDS_RANGES)*len(XG_THRS)*len(ELO_H_THRS)*len(FORM_THRS)*len(MKT_H_THRS) * len(LEAGUES):,}")

def apply_mask(d, side, lo, hi, xg_t, elo_t, form_t, mkt_t):
    oc  = 'home_odds_val'   if side == 'home' else 'away_odds_val'
    xgc = 'xg_ratio_home_5' if side == 'home' else 'xg_ratio_away_5'
    fc  = 'home_pts_5'      if side == 'home' else 'away_pts_5'
    mc  = 'mkt_home_prob'   if side == 'home' else 'mkt_away_prob'
    m = (d[oc] >= lo) & (d[oc] < hi)
    if xg_t > 0:                        m &= (d[xgc].fillna(0) >= xg_t)
    if side == 'home' and elo_t > 0:    m &= (d['elo_diff'].fillna(0) >= elo_t)
    if side == 'away' and elo_t < 0:    m &= (d['elo_diff'].fillna(0) <= elo_t)
    if form_t > 0:                      m &= (d[fc].fillna(0) >= form_t)
    if mkt_t > 0:                       m &= (d[mc].fillna(0) >= mkt_t)
    return d[m]

# ── SWEEP ─────────────────────────────────────────────────────────────────────
print("\nRunning full sweep...")
results = {}

for side in ('home', 'away'):
    et = ELO_H_THRS if side == 'home' else ELO_A_THRS
    mt = MKT_H_THRS if side == 'home' else MKT_A_THRS
    wc = 'home_win'      if side == 'home' else 'away_win'
    oc = 'home_odds_val' if side == 'home' else 'away_odds_val'
    print(f"  Side: {side}...")

    for lo, hi in ODDS_RANGES:
        oos_pre = apply_mask(oos_df, side, lo, hi, 0, 0, 0, 0)

        for xg_t, elo_t, form_t, mkt_t in product(XG_THRS, et, FORM_THRS, mt):
            parts = [f'{side}[{lo},{hi})']
            if xg_t > 0:                     parts.append(f'xg>={xg_t}')
            if side == 'home' and elo_t > 0: parts.append(f'elo>={elo_t}')
            if side == 'away' and elo_t < 0: parts.append(f'elo<={elo_t}')
            if form_t > 0:                   parts.append(f'form>={form_t}')
            if mkt_t > 0:                    parts.append(f'mkt>={mkt_t}')
            label = ' '.join(parts)

            for lg in LEAGUES:
                key = (lg, label)
                fold_data = {}

                for fname, _, _, _, _ in FOLDS:
                    fs = fold_slices[fname]
                    tr = apply_mask(fs['train'], side, lo, hi, xg_t, elo_t, form_t, mkt_t)
                    tr = tr[tr['league_name'] == lg]
                    if len(tr) < MIN_TRAIN_N: continue

                    te = apply_mask(fs['test'], side, lo, hi, xg_t, elo_t, form_t, mkt_t)
                    te = te[te['league_name'] == lg]
                    if len(te) < MIN_TEST_N: continue

                    test_roi  = (te[wc]*(te[oc]-1)-(1-te[wc])).mean()*100
                    test_wr   = te[wc].mean()*100
                    train_roi = (tr[wc]*(tr[oc]-1)-(1-tr[wc])).mean()*100

                    fold_data[fname] = {
                        'train_n':   len(tr),
                        'train_roi': round(train_roi, 1),
                        'test_n':    len(te),
                        'test_roi':  round(test_roi, 1),
                        'test_wr':   round(test_wr, 1),
                    }

                if len(fold_data) < 2: continue

                sl_oos = apply_mask(oos_pre, side, lo, hi, xg_t, elo_t, form_t, mkt_t)
                sl_oos = sl_oos[sl_oos['league_name'] == lg]
                oos_stats = None
                if len(sl_oos) >= MIN_OOS_N:
                    oos_roi = (sl_oos[wc]*(sl_oos[oc]-1)-(1-sl_oos[wc])).mean()*100
                    oos_stats = {
                        'n':   len(sl_oos),
                        'roi': round(oos_roi, 1),
                        'wr':  round(sl_oos[wc].mean()*100, 1),
                    }

                results[key] = {
                    'meta': {
                        'label': label, 'side': side, 'lo': lo, 'hi': hi,
                        'xg_t': xg_t, 'elo_t': elo_t, 'form_t': form_t, 'mkt_t': mkt_t,
                        'lg': lg, 'lshort': LSHORT[lg],
                    },
                    'folds': fold_data,
                    'oos':   oos_stats,
                }

print(f"Combinations with data: {len(results):,}")

# ── Stability + volume classification ─────────────────────────────────────────
all_entries = []
for key, data in results.items():
    folds = data['folds']
    n_sel  = len(folds)
    n_prof = sum(1 for f in folds.values() if f['test_roi'] > 0)
    if n_prof < MIN_STABLE_FOLDS: continue

    avg_test_roi  = np.mean([f['test_roi']  for f in folds.values()])
    avg_train_roi = np.mean([f['train_roi'] for f in folds.values()])
    avg_test_wr   = np.mean([f['test_wr']   for f in folds.values()])
    avg_test_n    = np.mean([f['test_n']    for f in folds.values()])

    volume = 'high' if avg_test_n >= HIGH_VOL_N else 'low'

    all_entries.append({
        **data['meta'],
        'n_folds':      n_sel,
        'n_profitable': n_prof,
        'stability':    round(n_prof / n_sel * 100, 0),
        'avg_train_roi':round(avg_train_roi, 1),
        'avg_test_roi': round(avg_test_roi, 1),
        'avg_test_wr':  round(avg_test_wr, 1),
        'avg_test_n':   round(avg_test_n, 1),
        'volume':       volume,
        'folds':        folds,
        'oos':          data['oos'],
    })

all_entries.sort(key=lambda x: (-x['n_profitable'], -x['avg_test_roi']))
high_vol = [e for e in all_entries if e['volume'] == 'high']
low_vol  = [e for e in all_entries if e['volume'] == 'low']
print(f"\nStable niches total: {len(all_entries)}")
print(f"  High volume (avg n>={HIGH_VOL_N}): {len(high_vol)}")
print(f"  Low volume  (avg n< {HIGH_VOL_N}): {len(low_vol)}")

# ── Dedup ─────────────────────────────────────────────────────────────────────
def is_subset(a, b, side):
    if b['xg_t'] < a['xg_t'] or b['form_t'] < a['form_t'] or b['mkt_t'] < a['mkt_t']:
        return False
    return b['elo_t'] >= a['elo_t'] if side == 'home' else b['elo_t'] <= a['elo_t']

def dedup(filters):
    filters = sorted(filters, key=lambda x: (-x['n_profitable'], -x['avg_test_roi']))
    kept = []
    for f in filters:
        if not any(is_subset(k, f, f['side']) and k['n_profitable'] >= f['n_profitable'] for k in kept):
            kept.append(f)
    return kept

def build_catalog(entries):
    catalog = {}
    for lg in LEAGUES:
        lshort = LSHORT[lg]
        items = [e for e in entries if e['lg'] == lg]
        groups = defaultdict(list)
        for f in items:
            groups[(f['side'], f['lo'], f['hi'])].append(f)
        catalog[lshort] = []
        for group in groups.values():
            catalog[lshort].extend(dedup(group))
    return catalog

cat_high = build_catalog(high_vol)
cat_low  = build_catalog(low_vol)

n_high = sum(len(v) for v in cat_high.values())
n_low  = sum(len(v) for v in cat_low.values())
print(f"\nAfter dedup:")
print(f"  High volume: {n_high} niches")
print(f"  Low volume:  {n_low} niches")

print("\nHigh volume per league:")
for lshort, items in sorted(cat_high.items(), key=lambda x: -len(x[1])):
    if not items: continue
    oos_pos = [i for i in items if i['oos'] and i['oos']['roi'] > 0]
    oos_all = [i for i in items if i['oos']]
    print(f"  {lshort:<12} {len(items):>2} ніш  OOS+: {len(oos_pos)}/{len(oos_all)}")

print("\nLow volume per league:")
for lshort, items in sorted(cat_low.items(), key=lambda x: -len(x[1])):
    if not items: continue
    oos_pos = [i for i in items if i['oos'] and i['oos']['roi'] > 0]
    oos_all = [i for i in items if i['oos']]
    print(f"  {lshort:<12} {len(items):>2} ніш  OOS+: {len(oos_pos)}/{len(oos_all)}")

# ── Excel ─────────────────────────────────────────────────────────────────────
print("\nBuilding Excel...")

GREEN_L  = PatternFill('solid', start_color='D5F5E3')
GREEN_D  = PatternFill('solid', start_color='A9DFBF')
RED_L    = PatternFill('solid', start_color='FADBD8')
GREY     = PatternFill('solid', start_color='F2F3F4')
DARK     = PatternFill('solid', start_color='1B2631')
SUBHDR   = PatternFill('solid', start_color='5D6D7E')
YELLOW   = PatternFill('solid', start_color='FEF9E7')
ORANGE   = PatternFill('solid', start_color='FAD7A0')
BLUE_L   = PatternFill('solid', start_color='D6EAF8')

HDR_F = Font(name='Arial', bold=True, color='FFFFFF', size=9)
C = Alignment(horizontal='center', vertical='center')
L = Alignment(horizontal='left',   vertical='center')

def tb():
    s = BdrSide(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def hc(c, v, fill=DARK):
    c.value=v; c.font=HDR_F; c.fill=fill; c.alignment=C; c.border=tb()

def bc(c, v, align=C, fill=None, bold=False):
    c.value=v; c.font=Font(name='Arial', size=9, bold=bold)
    c.alignment=align; c.border=tb()
    if fill: c.fill=fill

wb = Workbook(); wb.remove(wb.active)

# ── SUMMARY ───────────────────────────────────────────────────────────────────
ws_s = wb.create_sheet('SUMMARY')
ws_s['A1'] = 'HIGH VOLUME (avg test n ≥ 8)'
ws_s['A1'].font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
ws_s['A1'].fill = PatternFill('solid', start_color='1A5276')
ws_s['A1'].alignment = L
ws_s.merge_cells('A1:G1')

SHDR = ['Ліга','Ніш','Avg Test n','Avg Test ROI%','Stability%','OOS підтверджено','Найкраща ніша']
for ci, h in enumerate(SHDR, 1): hc(ws_s.cell(2, ci), h)
sr = 3
for lshort, items in sorted(cat_high.items(), key=lambda x: -len(x[1])):
    if not items: continue
    oos_pos = [i for i in items if i['oos'] and i['oos']['roi'] > 0]
    oos_all = [i for i in items if i['oos']]
    best = max(items, key=lambda x: x['avg_test_roi'])
    row = [lshort, len(items),
           round(np.mean([i['avg_test_n']   for i in items]), 1),
           round(np.mean([i['avg_test_roi'] for i in items]), 1),
           round(np.mean([i['stability']    for i in items]), 0),
           f"{len(oos_pos)}/{len(oos_all)}",
           best['label']]
    fmts = [None,None,None,'+0.0%','0%',None,None]
    for ci,(v,fmt) in enumerate(zip(row,fmts),1):
        c=ws_s.cell(sr,ci)
        bc(c, v/100 if fmt and isinstance(v,(int,float)) else v, L if ci in(1,7) else C)
        if fmt and isinstance(v,(int,float)): c.number_format=fmt
        if ci==4: c.fill=GREEN_L if row[3]>0 else RED_L
    sr+=1

# Low volume section
sr += 1
ws_s.cell(sr,1).value='LOW VOLUME (avg test n < 8) — потенційні gems, менша впевненість'
ws_s.cell(sr,1).font=Font(name='Arial',bold=True,size=11,color='FFFFFF')
ws_s.cell(sr,1).fill=PatternFill('solid',start_color='7D6608')
ws_s.cell(sr,1).alignment=L
ws_s.merge_cells(f'A{sr}:G{sr}')
sr+=1

for ci,h in enumerate(SHDR,1): hc(ws_s.cell(sr,ci),h,fill=PatternFill('solid',start_color='9A7D0A'))
sr+=1
for lshort,items in sorted(cat_low.items(),key=lambda x:-len(x[1])):
    if not items: continue
    oos_pos=[i for i in items if i['oos'] and i['oos']['roi']>0]
    oos_all=[i for i in items if i['oos']]
    best=max(items,key=lambda x:x['avg_test_roi'])
    row=[lshort,len(items),
         round(np.mean([i['avg_test_n']   for i in items]),1),
         round(np.mean([i['avg_test_roi'] for i in items]),1),
         round(np.mean([i['stability']    for i in items]),0),
         f"{len(oos_pos)}/{len(oos_all)}",best['label']]
    fmts=[None,None,None,'+0.0%','0%',None,None]
    for ci,(v,fmt) in enumerate(zip(row,fmts),1):
        c=ws_s.cell(sr,ci)
        bc(c,v/100 if fmt and isinstance(v,(int,float)) else v,L if ci in(1,7) else C,YELLOW)
        if fmt and isinstance(v,(int,float)): c.number_format=fmt
        if ci==4: c.fill=GREEN_L if row[3]>0 else RED_L
    sr+=1

ws_s.column_dimensions['A'].width=12; ws_s.column_dimensions['B'].width=6
ws_s.column_dimensions['C'].width=11; ws_s.column_dimensions['D'].width=13
ws_s.column_dimensions['E'].width=12; ws_s.column_dimensions['F'].width=16
ws_s.column_dimensions['G'].width=55
ws_s.freeze_panes='A3'

# ── Per-league sheets (high + low разом, з розділювачем) ──────────────────────
STATIC = ['Ніша','Side','Volume','Prof/Folds','Stability%',
          'Avg Test n','Avg Test ROI%','Avg Test WR%',
          'OOS n','OOS WR%','OOS ROI%']
N_S = len(STATIC)

all_leagues = set(list(cat_high.keys()) + list(cat_low.keys()))
for lshort in [LSHORT[lg] for lg in LEAGUES]:
    h_items = cat_high.get(lshort, [])
    l_items = cat_low.get(lshort, [])
    if not h_items and not l_items: continue

    ws = wb.create_sheet(lshort)
    total_c = N_S + N_FOLDS*2

    for ci,h in enumerate(STATIC,1): hc(ws.cell(1,ci),h)
    for fi,(fname,_,_,ts,te) in enumerate(FOLDS):
        cs=N_S+fi*2+1
        ws.merge_cells(start_row=1,start_column=cs,end_row=1,end_column=cs+1)
        hc(ws.cell(1,cs),f'{fname} ({ts.strftime("%b%y")}→{te.strftime("%b%y")})')
    for ci in range(1,N_S+1): hc(ws.cell(2,ci),'',fill=SUBHDR)
    for fi in range(N_FOLDS):
        cs=N_S+fi*2+1
        hc(ws.cell(2,cs),  'Tr ROI%',fill=SUBHDR)
        hc(ws.cell(2,cs+1),'Te ROI%',fill=SUBHDR)

    er = 3

    er_ref = [er]

    def write_section(items, section_label, title_fill):
        if not items: return
        ws.merge_cells(start_row=er_ref[0],start_column=1,end_row=er_ref[0],end_column=total_c)
        c=ws.cell(er_ref[0],1); c.value=section_label
        c.font=Font(name='Arial',bold=True,color='FFFFFF',size=10)
        c.fill=title_fill; c.alignment=L
        er_ref[0]+=1

        for niche in sorted(items,key=lambda x:(-x['n_profitable'],-x['avg_test_roi'])):
            oos=niche.get('oos')
            is_high=niche['volume']=='high'
            row_fill=GREEN_L if (niche['avg_test_roi']>0 and is_high) else \
                     GREEN_D if (niche['avg_test_roi']>0 and not is_high) else RED_L

            sv=[niche['label'],niche['side'],
                'High' if is_high else 'Low',
                f"{niche['n_profitable']}/{niche['n_folds']}",
                niche['stability']/100,
                round(niche['avg_test_n'],1),
                niche['avg_test_roi']/100,
                niche['avg_test_wr']/100,
                oos['n']      if oos else '—',
                oos['wr']/100 if oos else '—',
                oos['roi']/100 if oos else '—']
            sf=[None,None,None,None,'0%',None,'+0.0%;-0.0%','0.0%',None,'0.0%','+0.0%;-0.0%']
            for ci,(v,fmt) in enumerate(zip(sv,sf),1):
                c=ws.cell(er_ref[0],ci)
                bc(c,v,L if ci in(1,2,3,4) else C,row_fill)
                if fmt and isinstance(v,(int,float)): c.number_format=fmt
                if ci==5: c.fill=GREEN_L if niche['stability']>=75 else YELLOW
                if ci==7:
                    c.fill=GREEN_L if niche['avg_test_roi']>0 else RED_L
                    c.font=Font(name='Arial',size=9,bold=abs(niche['avg_test_roi'])>10)
                if ci==11 and oos:
                    c.fill=GREEN_L if oos['roi']>0 else RED_L
                    c.font=Font(name='Arial',size=9,bold=True)

            for fi,(fname,_,_,_,_) in enumerate(FOLDS):
                cs=N_S+fi*2+1
                fd=niche['folds'].get(fname)
                if fd:
                    for col_i,(val,is_test) in enumerate([(fd['train_roi'],False),(fd['test_roi'],True)]):
                        c=ws.cell(er_ref[0],cs+col_i)
                        c.value=val/100; c.number_format='+0.0%;-0.0%'
                        c.font=Font(name='Arial',size=9,bold=(is_test and abs(val)>15))
                        c.alignment=C; c.border=tb()
                        c.fill=GREEN_L if val>0 else RED_L
                else:
                    bc(ws.cell(er_ref[0],cs),'—',fill=GREY)
                    bc(ws.cell(er_ref[0],cs+1),'—',fill=GREY)
            er_ref[0]+=1

    write_section(h_items,'HIGH VOLUME (avg n>=8) — надійні ніші',PatternFill('solid',start_color='1A5276'))
    write_section(l_items,'LOW VOLUME (avg n<8) — потенційні gems',PatternFill('solid',start_color='7D6608'))
    er = er_ref[0]

    ws.column_dimensions['A'].width=52; ws.column_dimensions['B'].width=6
    ws.column_dimensions['C'].width=9
    for ci in range(4,N_S+1): ws.column_dimensions[get_column_letter(ci)].width=11
    for fi in range(N_FOLDS):
        cs=N_S+fi*2+1
        ws.column_dimensions[get_column_letter(cs)].width=9
        ws.column_dimensions[get_column_letter(cs+1)].width=9
    ws.freeze_panes='A3'
    ws.row_dimensions[1].height=28; ws.row_dimensions[2].height=18

outpath=os.path.join(os.path.dirname(__file__),'expanding_wf_v3_results.xlsx')
wb.save(outpath)
print(f"\nSaved: {outpath}  ({os.path.getsize(outpath)//1024}KB)")
print("\n✅ Done!")
